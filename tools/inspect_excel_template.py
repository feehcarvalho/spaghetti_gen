"""Inspeciona o template Excel sem alterar o arquivo original."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
FALLBACK_TEMPLATE = REPO_ROOT / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
REPORT_PATH = REPO_ROOT / "docs" / "EXCEL_TEMPLATE_DIAGNOSTICO.md"

NON_EMPTY_SCAN_ROWS = 15
FORMULA_SCAN_ROWS = 20
DISPLAY_CELL_LIMIT = 35
DISPLAY_FORMULA_LIMIT = 35
DISPLAY_MERGE_LIMIT = 25


@dataclass
class CellValue:
    coordinate: str
    value: str


@dataclass
class FormulaValue:
    coordinate: str
    formula: str


@dataclass
class SheetInspection:
    title: str
    dimension: str
    max_row: int
    max_column: int
    non_empty_first_rows: list[CellValue]
    formulas_first_rows: list[FormulaValue]
    merged_ranges: list[str]
    chart_count: int
    image_count: int
    table_count: int
    hyperlink_count: int
    comment_count: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Inspeciona um template Excel e gera diagnostico Markdown."
    )
    parser.add_argument(
        "template_path",
        nargs="?",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help="Caminho do template .xlsx. Padrao: data/templates/PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx",
    )
    return parser.parse_args()


def resolve_template_path(requested_path: Path) -> tuple[Path, str | None]:
    if requested_path.exists():
        return requested_path, None

    if requested_path == DEFAULT_TEMPLATE and FALLBACK_TEMPLATE.exists():
        return FALLBACK_TEMPLATE, (
            f"Arquivo padrao {DEFAULT_TEMPLATE.relative_to(REPO_ROOT)} nao encontrado; "
            f"usando fallback {FALLBACK_TEMPLATE.relative_to(REPO_ROOT)}."
        )

    raise FileNotFoundError(f"Template nao encontrado: {requested_path}")


def format_value(value: Any, limit: int = 180) -> str:
    if isinstance(value, datetime):
        text = value.isoformat(sep=" ")
    else:
        text = str(value)

    text = text.replace("\n", "\\n").strip()
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def cell_has_value(cell: Cell) -> bool:
    return cell.value is not None and str(cell.value).strip() != ""


def is_formula(cell: Cell) -> bool:
    return cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )


def collect_non_empty_cells(sheet: Worksheet) -> list[CellValue]:
    cells: list[CellValue] = []
    max_row = min(sheet.max_row, NON_EMPTY_SCAN_ROWS)

    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=sheet.max_column):
        for cell in row:
            if cell_has_value(cell):
                cells.append(CellValue(cell.coordinate, format_value(cell.value)))

    return cells


def collect_formulas(sheet: Worksheet) -> list[FormulaValue]:
    formulas: list[FormulaValue] = []
    max_row = min(sheet.max_row, FORMULA_SCAN_ROWS)

    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=sheet.max_column):
        for cell in row:
            if is_formula(cell):
                formulas.append(FormulaValue(cell.coordinate, format_value(cell.value, 260)))

    return formulas


def count_hyperlinks(sheet: Worksheet) -> int:
    count = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.hyperlink is not None:
                count += 1
    return count


def count_comments(sheet: Worksheet) -> int:
    count = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.comment is not None:
                count += 1
    return count


def inspect_sheet(sheet: Worksheet) -> SheetInspection:
    return SheetInspection(
        title=sheet.title,
        dimension=sheet.calculate_dimension(),
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        non_empty_first_rows=collect_non_empty_cells(sheet),
        formulas_first_rows=collect_formulas(sheet),
        merged_ranges=[str(cell_range) for cell_range in sheet.merged_cells.ranges],
        chart_count=len(getattr(sheet, "_charts", [])),
        image_count=len(getattr(sheet, "_images", [])),
        table_count=len(sheet.tables),
        hyperlink_count=count_hyperlinks(sheet),
        comment_count=count_comments(sheet),
    )


def detect_external_links(workbook: Any) -> list[str]:
    links = []
    for index, link in enumerate(getattr(workbook, "_external_links", []), start=1):
        target = getattr(getattr(link, "file_link", None), "Target", None)
        links.append(str(target or f"ExternalLink{index}"))
    return links


def find_formula_risks(inspections: list[SheetInspection]) -> dict[str, int]:
    return {item.title: len(item.formulas_first_rows) for item in inspections if item.formulas_first_rows}


def find_broken_reference_risks(inspections: list[SheetInspection]) -> dict[str, int]:
    risks: dict[str, int] = {}
    for item in inspections:
        count = sum(1 for formula in item.formulas_first_rows if "#REF!" in formula.formula)
        if count:
            risks[item.title] = count
    return risks


def truncate_list(values: list[Any], limit: int) -> tuple[list[Any], int]:
    if len(values) <= limit:
        return values, 0
    return values[:limit], len(values) - limit


def print_section_list(prefix: str, values: list[str], limit: int = DISPLAY_MERGE_LIMIT) -> None:
    displayed, remaining = truncate_list(values, limit)
    if not displayed:
        print(f"  {prefix}: nenhuma")
        return

    print(f"  {prefix}: {', '.join(displayed)}")
    if remaining:
        print(f"  {prefix}: ... +{remaining} itens")


def print_diagnostics(template_path: Path, note: str | None, inspections: list[SheetInspection]) -> None:
    print("Diagnostico do template Excel")
    print(f"Arquivo: {template_path}")
    if note:
        print(f"Observacao: {note}")
    print()

    print("Abas encontradas:")
    for index, item in enumerate(inspections, start=1):
        print(f"{index:02d}. {item.title} ({item.dimension})")
    print()

    for item in inspections:
        print(f"## {item.title}")
        print(f"- Dimensao usada: {item.dimension} ({item.max_row} linhas x {item.max_column} colunas)")

        cells, remaining_cells = truncate_list(item.non_empty_first_rows, DISPLAY_CELL_LIMIT)
        if cells:
            print(f"- Celulas nao vazias nas primeiras {NON_EMPTY_SCAN_ROWS} linhas:")
            for cell in cells:
                print(f"  - {cell.coordinate}: {cell.value}")
            if remaining_cells:
                print(f"  - ... +{remaining_cells} celulas")
        else:
            print(f"- Celulas nao vazias nas primeiras {NON_EMPTY_SCAN_ROWS} linhas: nenhuma")

        if item.formulas_first_rows:
            print(f"- Formulas nas primeiras {FORMULA_SCAN_ROWS} linhas:")
            formulas, remaining_formulas = truncate_list(
                item.formulas_first_rows,
                DISPLAY_FORMULA_LIMIT,
            )
            for formula in formulas:
                print(f"  - {formula.coordinate}: {formula.formula}")
            if remaining_formulas:
                print(f"  - ... +{remaining_formulas} formulas")
        else:
            print(f"- Formulas nas primeiras {FORMULA_SCAN_ROWS} linhas: nenhuma")

        print_section_list("- Celulas mescladas", item.merged_ranges)
        print(
            "- Objetos: "
            f"graficos={item.chart_count}, imagens={item.image_count}, "
            f"tabelas={item.table_count}, hyperlinks={item.hyperlink_count}, comentarios={item.comment_count}"
        )
        print()


def markdown_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(value) for value in row) + " |")
    return lines


def sheet_by_name(inspections: list[SheetInspection], name: str) -> SheetInspection | None:
    normalized = name.casefold()
    for item in inspections:
        if item.title.casefold() == normalized:
            return item
    return None


def cell_value_map(sheet: SheetInspection) -> dict[str, str]:
    return {cell.coordinate: cell.value for cell in sheet.non_empty_first_rows}


def build_sheet_overview(inspections: list[SheetInspection]) -> list[str]:
    rows = []
    for item in inspections:
        rows.append(
            [
                item.title,
                item.dimension,
                len(item.merged_ranges),
                len(item.formulas_first_rows),
                item.chart_count,
                item.image_count,
                item.hyperlink_count,
            ]
        )

    return markdown_table(
        ["Aba", "Dimensao usada", "Mesclas", "Formulas L1-L20", "Graficos", "Imagens", "Links"],
        rows,
    )


def build_candidates(inspections: list[SheetInspection]) -> list[str]:
    existing = {item.title.casefold(): item.title for item in inspections}
    rows: list[list[str]] = []

    if "análise".casefold() in existing:
        rows.append(
            [
                existing["análise".casefold()],
                "Alta",
                "Cabecalho, tabela de microetapas, totais AV/NAV/D.",
                "Escrever apenas celulas mapeadas e preservar formatacao.",
            ]
        )
    if "melhorias".casefold() in existing:
        rows.append(
            [
                existing["melhorias".casefold()],
                "Alta",
                "Tabela de desperdicios, sugestoes e recomendacoes gerais.",
                "Limitar escrita aos campos de dados e nao mexer em estruturas.",
            ]
        )
    if "diagrama de espaguete".casefold() in existing:
        rows.append(
            [
                existing["diagrama de espaguete".casefold()],
                "Media",
                "Possivel insercao futura de imagem/rota spaghetti.",
                "Exige cuidado por alta quantidade de celulas mescladas.",
            ]
        )

    rows.append(
        [
            "Standard*, Worktable, graficos e validacao",
            "Baixa/nao escrever",
            "Abas de apoio, padrao e calculos existentes.",
            "Preservar integralmente ate haver mapeamento especifico.",
        ]
    )

    return markdown_table(["Aba", "Prioridade", "Uso candidato", "Restricao"], rows)


def build_analysis_mapping(sheet: SheetInspection | None) -> list[str]:
    if sheet is None:
        return ["Aba `ANÁLISE` nao encontrada no workbook."]

    values = cell_value_map(sheet)
    rows = [
        ["A1", "Titulo da analise", values.get("A1", "")],
        ["B3", "Departamento/area", values.get("B3", "")],
        ["E3", "Emitido por/responsavel", values.get("E3", "")],
        ["H3", "Data", values.get("H3", "")],
        ["B4", "Posto", values.get("B4", "")],
        ["E4", "Processo", values.get("E4", "")],
        ["H4", "Takt", values.get("H4", "")],
        ["B5", "Ciclo observado", values.get("B5", "")],
        ["E5", "Total AV", values.get("E5", "")],
        ["G5", "Total NAV", values.get("G5", "")],
        ["I5", "Total D", values.get("I5", "")],
        ["A6:I6", "Cabecalho da tabela de microetapas", "Nº, etapa, inicio, fim, duracao, classificacao, justificativa, ferramenta"],
        ["A7:I59", "Dados atuais de microetapas", "Faixa observada no template; escrever somente apos copia do arquivo."],
    ]

    lines = markdown_table(["Celula/faixa", "Campo candidato", "Valor observado"], rows)
    lines.append("")
    lines.append(
        f"Mesclas relevantes na aba: {len(sheet.merged_ranges)} intervalo(s). "
        "A escrita deve mirar a celula superior esquerda de cada mescla quando aplicavel."
    )
    return lines


def build_improvements_mapping(sheet: SheetInspection | None) -> list[str]:
    if sheet is None:
        return ["Aba `MELHORIAS` nao encontrada no workbook."]

    values = cell_value_map(sheet)
    rows = [
        ["A1", "Titulo da aba", values.get("A1", "")],
        ["B3", "Ciclo observado", values.get("B3", "")],
        ["E3", "Takt", values.get("E3", "")],
        ["H3", "Folga vs takt", values.get("H3", "")],
        ["B4", "Tempo D", values.get("B4", "")],
        ["E4", "Leitura/diagnostico", values.get("E4", "")],
        ["A6:H6", "Cabecalho da tabela de melhorias", "Etapa, inicio, fim, duracao, desperdicio, tipo, sugestao, prioridade"],
        ["A7:H13", "Dados atuais de melhorias", "Faixa inicial para sugestoes relacionadas a desperdicios."],
        ["A14:H18", "Recomendacoes gerais", "Bloco textual ja existente no template."],
    ]

    lines = markdown_table(["Celula/faixa", "Campo candidato", "Valor observado"], rows)
    lines.append("")
    lines.append(
        f"Mesclas relevantes na aba: {len(sheet.merged_ranges)} intervalo(s). "
        "Evitar inserir linhas antes de confirmar impacto visual."
    )
    return lines


def build_risks(
    inspections: list[SheetInspection],
    external_links: list[str],
    workbook_defined_names: list[str],
) -> list[str]:
    formula_risks = find_formula_risks(inspections)
    broken_ref_risks = find_broken_reference_risks(inspections)
    chart_sheets = [item.title for item in inspections if item.chart_count]
    image_sheets = [item.title for item in inspections if item.image_count]
    heavy_merged = [f"{item.title} ({len(item.merged_ranges)})" for item in inspections if item.merged_ranges]
    hyperlink_sheets = [f"{item.title} ({item.hyperlink_count})" for item in inspections if item.hyperlink_count]

    lines = [
        "- Fórmulas: abrir com `data_only=False` preserva as fórmulas carregadas. "
        f"Foram encontradas fórmulas nas primeiras {FORMULA_SCAN_ROWS} linhas em "
        f"{len(formula_risks)} aba(s): {', '.join(formula_risks) if formula_risks else 'nenhuma nas linhas inspecionadas'}.",
        "- Células mescladas: há mesclas em várias abas; escrever fora da célula superior esquerda de uma mescla pode falhar ou corromper o layout.",
        f"- Abas com muitas mesclas: {', '.join(heavy_merged) if heavy_merged else 'nenhuma'}.",
        f"- Gráficos: {', '.join(chart_sheets) if chart_sheets else 'nenhum gráfico carregado pelo openpyxl'}.",
        f"- Imagens: {', '.join(image_sheets) if image_sheets else 'nenhuma imagem carregada pelo openpyxl'}.",
        f"- Links externos: {', '.join(external_links) if external_links else 'nenhum link externo detectado via openpyxl'}.",
        f"- Hyperlinks em células: {', '.join(hyperlink_sheets) if hyperlink_sheets else 'nenhum hyperlink detectado'}.",
        f"- Nomes definidos: {', '.join(workbook_defined_names) if workbook_defined_names else 'nenhum nome definido detectado'}.",
        "- Referências quebradas: "
        f"{', '.join(f'{sheet} ({count})' for sheet, count in broken_ref_risks.items()) if broken_ref_risks else 'nenhuma referencia #REF! nas primeiras 20 linhas inspecionadas'}. "
        "O inspetor não altera nem recalcula fórmulas.",
    ]
    return lines


def write_report(
    template_path: Path,
    note: str | None,
    inspections: list[SheetInspection],
    external_links: list[str],
    defined_names: list[str],
) -> None:
    analysis_sheet = sheet_by_name(inspections, "ANÁLISE")
    improvements_sheet = sheet_by_name(inspections, "MELHORIAS")
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines: list[str] = [
        "# Diagnóstico Técnico do Template Excel",
        "",
        f"- Gerado em: {generated_at}",
        f"- Arquivo inspecionado: `{template_path.relative_to(REPO_ROOT) if template_path.is_relative_to(REPO_ROOT) else template_path}`",
        "- Modo de leitura: `openpyxl.load_workbook(..., data_only=False, keep_links=True)`",
        "- Garantia operacional: o script não chama `save()` e não modifica o template.",
    ]

    if note:
        lines.extend(["", f"> Observação: {note}"])

    lines.extend(["", "## Abas Existentes", ""])
    lines.extend(build_sheet_overview(inspections))

    lines.extend(["", "## Abas Candidatas a Preenchimento Automático", ""])
    lines.extend(build_candidates(inspections))

    lines.extend(["", "## Mapeamento Inicial da Aba ANÁLISE", ""])
    lines.extend(build_analysis_mapping(analysis_sheet))

    lines.extend(["", "## Mapeamento Inicial da Aba MELHORIAS", ""])
    lines.extend(build_improvements_mapping(improvements_sheet))

    lines.extend(["", "## Riscos Técnicos", ""])
    lines.extend(build_risks(inspections, external_links, defined_names))

    lines.extend(
        [
            "",
            "## Recomendação de Implementação",
            "",
            "- Nunca recriar o workbook do zero.",
            "- Sempre copiar o template para `data/outputs/` e editar somente a cópia.",
            "- Escrever apenas células/faixas mapeadas nas abas `ANÁLISE` e `MELHORIAS`.",
            "- Preservar fórmulas, células mescladas, gráficos, estilos, validações, imagens e abas de apoio.",
            "- Não usar pandas para escrita Excel neste fluxo; usar `openpyxl` sobre uma cópia do template.",
            "- Validar o JSON com `OperationalAnalysis` antes de qualquer preenchimento.",
        ]
    )

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    args = parse_args()

    try:
        template_path, note = resolve_template_path(args.template_path)
        workbook = load_workbook(
            template_path,
            read_only=False,
            data_only=False,
            keep_links=True,
            rich_text=True,
        )
        inspections = [inspect_sheet(sheet) for sheet in workbook.worksheets]
        external_links = detect_external_links(workbook)
        defined_names = sorted(str(name) for name in workbook.defined_names)

        print_diagnostics(template_path, note, inspections)
        write_report(template_path, note, inspections, external_links, defined_names)
        print(f"Relatorio Markdown gerado em: {REPORT_PATH}")
        return 0
    except Exception as exc:
        print(f"Erro ao inspecionar template: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
