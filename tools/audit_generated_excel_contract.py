from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

from app.analysis.operational_language_repair import validate_operational_text  # noqa: E402
from app.analysis.operational_language_repair import detect_generic_process_phrases  # noqa: E402
from app.analysis.quality_alerts import ALERT_HEADERS, ALERTS_SHEET_NAME  # noqa: E402
from app.excel.conversion_sheet_writer import CONVERSION_HEADERS  # noqa: E402


STANDARD_DATA_ROW = 9


def audit(path: Path) -> list[str]:
    errors: list[str] = []
    workbook = load_workbook(path, data_only=True)
    try:
        if not workbook.sheetnames or workbook.sheetnames[0] not in {"STANDARD_CONSOLIDADO", "STANDARD_COMPLETO"}:
            errors.append("ERRO: Primeira aba deve ser STANDARD_CONSOLIDADO ou STANDARD_COMPLETO.")
        if len(workbook.sheetnames) < 2 or workbook.sheetnames[1] != "Sheet2":
            errors.append("Sheet2 nao esta na segunda posicao.")
            return errors
        if len(workbook.sheetnames) < 3 or workbook.sheetnames[2] != "ENTENDIMENTO_CONVERSAO":
            errors.append("ERRO: ENTENDIMENTO_CONVERSAO nao esta na terceira posicao.")

        first = workbook[workbook.sheetnames[0]]
        errors.extend(_audit_first_standard_sheet(first, workbook))

        sheet2 = workbook["Sheet2"]
        headers = [sheet2.cell(row=1, column=column).value for column in range(1, 10)]
        if headers != CONVERSION_HEADERS:
            errors.append(f"Cabecalho Sheet2 invalido: {headers}")

        times = []
        activities = []
        for row in range(2, sheet2.max_row + 1):
            activity = sheet2.cell(row=row, column=2).value
            time_value = sheet2.cell(row=row, column=5).value
            if activity is None and time_value is None:
                continue
            if not isinstance(time_value, (int, float)):
                errors.append(f"Sheet2 linha {row}: timeOfEelement nao numerico.")
            else:
                times.append(float(time_value))
            text = str(activity or "").strip()
            activities.append(text)
            if "classificada como" in text.casefold() or "porque" in text.casefold():
                errors.append(f"Sheet2 linha {row}: activity parece justificativa.")
            for alert in validate_operational_text(text):
                errors.append(f"Sheet2 linha {row}: {alert}")
            for alert in detect_generic_process_phrases(text):
                errors.append(f"Sheet2 linha {row}: {alert}")

        if len(times) >= 2:
            running = 0.0
            for index, value in enumerate(times, start=2):
                running += value
                if abs(value - running) < 0.01 and index > 2:
                    errors.append(f"Sheet2 linha {index}: timeOfEelement parece acumulado.")

        repeated = {}
        for activity in activities:
            if not activity:
                continue
            key = " ".join(activity.casefold().split())
            repeated[key] = repeated.get(key, 0) + 1
        for activity, count in repeated.items():
            if count >= 3:
                errors.append(f"Repeticao excessiva de activity: '{activity}' ({count} vezes).")
        joined_activities = " ".join(activities).casefold()
        workbook_text = " ".join(_visible_sheet_text(sheet).casefold() for sheet in workbook.worksheets if sheet.sheet_state == "visible")
        if ("green box" in workbook_text or "caixa verde" in workbook_text) and "bluebox" in joined_activities:
            errors.append("ERRO: activity usa Bluebox apesar de contexto visivel indicar Green Box/caixa verde.")
        if "parafusadeira pneum" in joined_activities and "manual" in workbook_text and "parafusadeira pneum" not in workbook_text.replace(joined_activities, ""):
            errors.append("ERRO: activity cita pneumática sem evidencia fora da propria activity.")

        errors.extend(_audit_sps_alerts_sheet(workbook))

    finally:
        workbook.close()
    return errors


def _audit_sps_alerts_sheet(workbook) -> list[str]:
    errors: list[str] = []
    if ALERTS_SHEET_NAME not in workbook.sheetnames:
        errors.append(f"ERRO: Aba {ALERTS_SHEET_NAME} ausente.")
        return errors
    if workbook.sheetnames[-1] != ALERTS_SHEET_NAME:
        errors.append(f"ERRO: Aba {ALERTS_SHEET_NAME} deve ser a ultima aba.")

    sheet = workbook[ALERTS_SHEET_NAME]
    headers = [sheet.cell(row=4, column=column).value for column in range(1, len(ALERT_HEADERS) + 1)]
    if headers != ALERT_HEADERS:
        errors.append(f"ERRO: Cabecalhos de {ALERTS_SHEET_NAME} invalidos: {headers}")

    for row in range(5, sheet.max_row + 1):
        values = [sheet.cell(row=row, column=column).value for column in range(1, len(ALERT_HEADERS) + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        if str(values[2] or "") == "Sem alerta":
            continue
        required = {
            "severidade": values[1],
            "tipo": values[2],
            "descricao": values[4],
            "acao_recomendada": values[5],
        }
        for label, value in required.items():
            if value in (None, ""):
                errors.append(f"ERRO: {ALERTS_SHEET_NAME} linha {row} sem {label}.")
    return errors


def _audit_first_standard_sheet(first, workbook) -> list[str]:
    errors: list[str] = []
    if str(first["A1"].value or "").strip().casefold() == "sequencia":
        errors.append("ERRO: Primeira aba esta em formato de tabela simples, nao layout Scania.")
    visible_text = _visible_sheet_text(first)
    lowered = visible_text.casefold()
    if "scania" not in lowered:
        errors.append("ERRO: Primeira aba nao contem SCANIA.")
    if "for internal use only" not in lowered:
        errors.append("ERRO: Primeira aba nao contem For internal use only.")
    for term in ("departmento", "área", "area", "posição", "posicao", "página", "pagina", "emitido", "data", "edição", "edicao"):
        if term in lowered:
            break
    else:
        errors.append("ERRO: Primeira aba nao contem campos de metadados Standard.")
    for coordinate, label in (("D5", "departamento"), ("E5", "posicao"), ("D7", "emitido por"), ("E7", "data"), ("D8", "processo")):
        if first[coordinate].value in (None, ""):
            errors.append(f"ERRO: Primeira aba sem {label} atual em {coordinate}.")

    process = str(first["D8"].value or "")
    if "pmgs" not in process.casefold():
        errors.extend(_find_visible_old_pmgs_data(workbook))

    accumulated = 0.0
    found_steps = 0
    for row in range(STANDARD_DATA_ROW, first.max_row + 1):
        sequence = first.cell(row=row, column=1).value
        element = first.cell(row=row, column=7).value
        acc = first.cell(row=row, column=8).value
        activity = str(first.cell(row=row, column=4).value or "").strip()
        if sequence is None and element is None and acc is None and not activity:
            continue
        if not isinstance(sequence, (int, float)):
            continue
        found_steps += 1
        if "classificada como" in activity.casefold() or "porque" in activity.casefold():
            errors.append(f"Primeira aba linha {row}: coluna D parece justificativa.")
        if not isinstance(element, (int, float)):
            errors.append(f"Primeira aba linha {row}: tempo do elemento nao numerico.")
            continue
        accumulated += float(element)
        if not isinstance(acc, (int, float)) or abs(float(acc) - accumulated) > 0.01:
            errors.append(f"Primeira aba linha {row}: tempo acumulado incorreto.")
    if found_steps == 0:
        errors.append("ERRO: Primeira aba nao contem microetapas da analise.")
    return errors


def _visible_sheet_text(sheet) -> str:
    values = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def _find_visible_old_pmgs_data(workbook) -> list[str]:
    errors: list[str] = []
    forbidden = ("pmgs.p1", "pré montagem", "pre montagem", "pre-montagem", "grade superior")
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        text = _visible_sheet_text(sheet).casefold()
        for term in forbidden:
            if term in text:
                errors.append(f"ERRO: Dados antigos de PMGS encontrados em aba visivel '{sheet.title}'.")
                break
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description="Audita contrato Excel SPS/Sheet2.")
    parser.add_argument("excel_path", type=Path)
    args = parser.parse_args()
    errors = audit(args.excel_path)
    if errors:
        for error in errors:
            print(error)
        return 1
    print("Contrato Excel OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
