"""Testes do pipeline principal com provider mock."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

from app.main import run_analysis_pipeline
from app.schemas.analysis import AnalysisMetadata, OperationalAnalysis


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
FALLBACK_TEMPLATE = REPO_ROOT / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"


def resolve_template() -> Path:
    return DEFAULT_TEMPLATE if DEFAULT_TEMPLATE.exists() else FALLBACK_TEMPLATE


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def test_pipeline_mock_generates_excel_json_without_video_or_api_key():
    template_path = resolve_template()
    output_path = REPO_ROOT / "data" / "outputs" / "test_pipeline_mock.xlsx"
    json_path = output_path.with_suffix(".json")
    original_hash = file_hash(template_path)

    metadata = AnalysisMetadata(
        departamento="FUNCTION AREA 5",
        posto="PMGS.P1",
        processo="Pré montagem da grade superior (PMGS)",
        responsavel="MARIANE",
        data_analise="2026-05-05",
        takt_time_s=330.0,
    )

    result = run_analysis_pipeline(
        video_path=None,
        template_path=str(template_path),
        output_path=str(output_path),
        metadata=metadata,
        provider_name="mock",
        fill_standard=True,
    )

    assert Path(result) == output_path
    assert output_path.exists()
    assert json_path.exists()
    assert file_hash(template_path) == original_hash

    data = json.loads(json_path.read_text(encoding="utf-8"))
    analysis = OperationalAnalysis.model_validate(data)
    assert analysis.metadata.posto == "PMGS.P1"
    assert analysis.metadata.departamento == "FUNCTION AREA 5"
    assert analysis.metadata.takt_time_s == 330.0
    assert analysis.resumo_tempos.folga_vs_takt_s == 246.0

    workbook = load_workbook(output_path, data_only=False, keep_links=True)
    assert "ANÁLISE" in workbook.sheetnames
    assert "MELHORIAS" in workbook.sheetnames
    assert "Standard (1-10)" in workbook.sheetnames
    assert workbook["ANÁLISE"]["B3"].value == "FUNCTION AREA 5"
    assert workbook["ANÁLISE"]["H4"].value == "330s (05:30)"
    assert workbook["Standard (1-10)"]["A9"].value == 1
    assert workbook["Worktable"]["I5"].value == "=H5+1"
