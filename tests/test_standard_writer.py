"""Testes do preenchimento controlado das abas Standard."""

from __future__ import annotations

import hashlib
import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.excel.template_writer import write_analysis_to_template
from app.schemas.analysis import OperationalAnalysis


REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_JSON = REPO_ROOT / "data" / "outputs" / "sample_analysis_pmgs_p1.json"
DEFAULT_TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
FALLBACK_TEMPLATE = REPO_ROOT / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"


def resolve_template() -> Path:
    if DEFAULT_TEMPLATE.exists():
        return DEFAULT_TEMPLATE
    if FALLBACK_TEMPLATE.exists():
        return FALLBACK_TEMPLATE
    pytest.skip("Template Excel real nao encontrado no workspace")


def load_sample_analysis() -> OperationalAnalysis:
    data = json.loads(SAMPLE_JSON.read_text(encoding="utf-8"))
    return OperationalAnalysis.model_validate(data)


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def test_standard_is_not_filled_when_parameter_false():
    template_path = resolve_template()
    output_path = REPO_ROOT / "data" / "outputs" / "test_standard_false.xlsx"
    analysis = load_sample_analysis()
    original_wb = load_workbook(template_path, data_only=False, keep_links=True)
    original_values = {
        "A9": original_wb["Standard (1-10)"]["A9"].value,
        "B9": original_wb["Standard (1-10)"]["B9"].value,
        "D9": original_wb["Standard (1-10)"]["D9"].value,
        "G9": original_wb["Standard (1-10)"]["G9"].value,
    }

    write_analysis_to_template(
        analysis=analysis,
        template_path=str(template_path),
        output_path=str(output_path),
        fill_standard=False,
    )

    generated_wb = load_workbook(output_path, data_only=False, keep_links=True)
    for coordinate, value in original_values.items():
        assert generated_wb["Standard (1-10)"][coordinate].value == value


def test_standard_is_filled_when_cli_flag_is_used():
    template_path = resolve_template()
    output_path = REPO_ROOT / "data" / "outputs" / "test_standard_cli.xlsx"
    template_hash = file_hash(template_path)
    original_wb = load_workbook(template_path, data_only=False, keep_links=True)
    original_sheetnames = list(original_wb.sheetnames)
    preserved_formulas = {
        ("Standard (1-10)", "L10"): original_wb["Standard (1-10)"]["L10"].value,
        ("Standard (3-10)", "H30"): original_wb["Standard (3-10)"]["H30"].value,
        ("Worktable", "I5"): original_wb["Worktable"]["I5"].value,
    }

    result = subprocess.run(
        [
            sys.executable,
            "tools/generate_excel_from_json.py",
            str(SAMPLE_JSON),
            str(template_path),
            str(output_path),
            "--fill-standard",
        ],
        cwd=REPO_ROOT,
        text=True,
        capture_output=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert "Standard preenchido: sim" in result.stdout
    assert output_path.exists()
    assert file_hash(template_path) == template_hash

    generated_wb = load_workbook(output_path, data_only=False, keep_links=True)
    assert generated_wb.sheetnames[1] == "Sheet2"
    assert generated_wb.sheetnames[2] == "ENTENDIMENTO_CONVERSAO"
    assert [name for name in generated_wb.sheetnames if name in original_sheetnames] == original_sheetnames

    standard = generated_wb["Standard (1-10)"]
    assert standard["A9"].value == 1
    assert standard["B9"].value == "NAV"
    assert standard["C9"].value == 1
    assert standard["D9"].value == "Verificar WPO e confirmar produto/variante no monitor do posto"
    assert standard["E9"].value == "WPO"
    assert standard["G9"].value == 8
    assert standard["A18"].value == 10
    assert standard["D18"].value == "Posicionar VR no proximo posto para liberar a sequencia"

    for (sheet_name, coordinate), formula in preserved_formulas.items():
        assert generated_wb[sheet_name][coordinate].value == formula
