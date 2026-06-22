from __future__ import annotations

import json
import subprocess
import sys
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.analysis.quality_alerts import ALERTS_SHEET_NAME
from app.analysis.quality_gate import validate_analysis_quality
from app.analysis.sps_validator import assert_analysis_can_generate_excel
from app.excel.template_writer import write_analysis_to_template
from app.schemas.analysis import AnalysisMetadata, MicroStep, OperationalAnalysis, TimeSummary


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
FALLBACK_TEMPLATE = REPO_ROOT / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
OUTPUT_DIR = REPO_ROOT / "data" / "outputs" / "test_quality_alerts"


def resolve_template() -> Path:
    if DEFAULT_TEMPLATE.exists():
        return DEFAULT_TEMPLATE
    if FALLBACK_TEMPLATE.exists():
        return FALLBACK_TEMPLATE
    pytest.skip("Template Excel real nao encontrado no workspace")


def metadata() -> AnalysisMetadata:
    return AnalysisMetadata(
        departamento="FA",
        posto="5.2.6",
        processo="Montagem operacional teste",
        responsavel="SPS",
        data_analise="2026-06-07",
        takt_time_s=60,
    )


def step(
    number: int,
    text: str,
    *,
    start: float | None = None,
    confidence: float = 0.95,
    justification: str = "Etapa necessaria ao metodo observado no processo.",
    eixo: str | None = None,
    lado: str | None = None,
) -> MicroStep:
    start = float(number - 1) * 3.0 if start is None else start
    end = start + 3.0
    return MicroStep(
        numero=number,
        inicio_s=start,
        fim_s=end,
        duracao_s=end - start,
        inicio_formatado="00:00",
        fim_formatado="00:03",
        duracao_formatada="00:03",
        tempo_acumulado_s=end,
        tempo_acumulado_formatado="00:03",
        instrucao_operacional=text,
        instrucao_padrao=text,
        etapa_detalhada=text,
        descricao_tecnica_detalhada=text,
        interpretacao_de_processo=text,
        classificacao="NAV",
        justificativa_tecnica=justification,
        eixo=eixo,
        lado=lado,
        confianca=confidence,
    )


def analysis_with_steps(steps: list[MicroStep], alerts: list[str] | None = None) -> OperationalAnalysis:
    total = sum(item.duracao_s for item in steps)
    return OperationalAnalysis(
        metadata=metadata(),
        microetapas=steps,
        resumo_tempos=TimeSummary(
            av_s=0,
            nav_s=total,
            d_s=0,
            total_s=total,
            av_percent=0,
            nav_percent=100,
            d_percent=0,
        ),
        alertas_validacao=alerts or [],
    )


def read_workbook(path: Path):
    return load_workbook(path, data_only=False, keep_links=True)


def output_file(name: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / f"{Path(name).stem}_{uuid.uuid4().hex}{Path(name).suffix}"
    return path


def test_action_in_justification_alert_does_not_block_excel():
    bad = analysis_with_steps(
        [
            step(
                1,
                "Realizar atividade operacional.",
                justification="Fixar a porca no eixo com parafusadeira pneumatica.",
            )
        ]
    )
    gate = validate_analysis_quality(bad, None, "")
    assert not gate.passed
    assert gate.can_export

    output = output_file("acao_justificativa.xlsx")
    write_analysis_to_template(bad, str(resolve_template()), str(output), quality_alerts=gate.alerts)

    workbook = read_workbook(output)
    try:
        assert ALERTS_SHEET_NAME in workbook.sheetnames
        assert workbook.sheetnames[-1] == ALERTS_SHEET_NAME
        alert_types = {
            str(workbook[ALERTS_SHEET_NAME].cell(row=row, column=3).value or "")
            for row in range(5, workbook[ALERTS_SHEET_NAME].max_row + 1)
        }
        assert any("justificativa" in alert_type.casefold() for alert_type in alert_types)
    finally:
        workbook.close()


def test_fatal_missing_analysis_still_blocks_export():
    with pytest.raises(ValueError, match="analise valida"):
        assert_analysis_can_generate_excel(None)

    empty = OperationalAnalysis.model_construct(
        metadata=metadata(),
        microetapas=[],
        resumo_tempos=TimeSummary(av_s=0, nav_s=0, d_s=0, total_s=0, av_percent=0, nav_percent=0, d_percent=0),
        melhorias=[],
        recomendacoes_gerais=[],
        alertas_validacao=[],
        roteiro_operacional=[],
    )
    with pytest.raises(ValueError, match="microetapas validas"):
        assert_analysis_can_generate_excel(empty)


def test_low_confidence_alert_is_written_to_last_sheet():
    analysis = analysis_with_steps([step(1, "Posicionar o componente no ponto de montagem.", confidence=0.55)])
    output = output_file("baixa_confianca.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    workbook = read_workbook(output)
    try:
        sheet = workbook[ALERTS_SHEET_NAME]
        assert workbook.sheetnames[-1] == ALERTS_SHEET_NAME
        descriptions = [sheet.cell(row=row, column=5).value for row in range(5, sheet.max_row + 1)]
        assert any("Confianca" in str(description) for description in descriptions)
    finally:
        workbook.close()


def test_uncertain_axis_or_side_alert_is_written():
    analysis = analysis_with_steps(
        [step(1, "Posicionar a ferramenta no ponto de trabalho confirmado.")],
        alerts=["Microetapa 1: Eixo/lado nao confirmado no video ou nas memorias."],
    )
    output = output_file("eixo_lado_alerta.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    workbook = read_workbook(output)
    try:
        sheet = workbook[ALERTS_SHEET_NAME]
        descriptions = [sheet.cell(row=row, column=5).value for row in range(5, sheet.max_row + 1)]
        assert any("Eixo/lado" in str(description) for description in descriptions)
    finally:
        workbook.close()


def test_sheet2_stays_second_and_alerts_sheet_is_last():
    analysis = analysis_with_steps(
        [
            step(1, "Pegar o componente na bancada indicada."),
            step(2, "Posicionar o componente no conjunto confirmado."),
        ]
    )
    output = output_file("ordem_abas.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    workbook = read_workbook(output)
    try:
        assert workbook.sheetnames[1] == "Sheet2"
        assert workbook.sheetnames[-1] == ALERTS_SHEET_NAME
    finally:
        workbook.close()


def test_generate_excel_from_recovery_script():
    analysis = analysis_with_steps(
        [step(1, "Posicionar a ferramenta no ponto de trabalho confirmado.")],
        alerts=["Microetapa 1: Ferramenta nao confirmada no video; validar no gemba."],
    )
    recovery_dir = OUTPUT_DIR / f"recovery_{uuid.uuid4().hex}"
    recovery_dir.mkdir(parents=True, exist_ok=True)
    recovery = recovery_dir / "analise_preservada_teste.json"
    recovery.write_text(
        json.dumps({"erro_exportacao": "quality gate", "analysis": analysis.model_dump(mode="json")}),
        encoding="utf-8",
    )

    completed = subprocess.run(
        [
            sys.executable,
            str(REPO_ROOT / "tools" / "generate_excel_from_recovery.py"),
            str(recovery),
            "--output-dir",
            str(recovery_dir),
        ],
        cwd=REPO_ROOT,
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    generated = list(recovery_dir.glob("analise_preservada_teste_*.xlsx"))
    assert generated
    workbook = read_workbook(generated[0])
    try:
        assert workbook.sheetnames[-1] == ALERTS_SHEET_NAME
    finally:
        workbook.close()


def test_ui_message_is_warning_not_blocking_error():
    source = (REPO_ROOT / "app" / "ui" / "streamlit_app.py").read_text(encoding="utf-8")
    assert "A planilha foi gerada com alertas de validacao SPS" in source
    assert ALERTS_SHEET_NAME in source
