from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.excel.template_writer import write_analysis_to_template
from app.schemas.analysis import AnalysisMetadata, MicroStep, OperationalAnalysis, TimeSummary


REPO_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"


def _step(number: int, start: float, end: float, text: str, classification: str = "NAV") -> MicroStep:
    return MicroStep(
        numero=number,
        inicio_s=start,
        fim_s=end,
        duracao_s=end - start,
        inicio_formatado="00:00",
        fim_formatado="00:00",
        duracao_formatada="00:00",
        etapa_detalhada=text,
        instrucao_operacional=text,
        classificacao=classification,
        justificativa_tecnica="Etapa necessaria ao metodo observado.",
        ferramenta_observacao="Referencia operacional",
        confianca=0.95,
    )


def _non_pmgs_analysis() -> OperationalAnalysis:
    steps = [
        _step(1, 0, 2, "Preparar o conjunto roda/pneu no ponto de abastecimento."),
        _step(2, 5, 9, "Deslocar o conjunto roda/pneu até o eixo indicado."),
        _step(3, 20, 23, "Conferir o posicionamento do conjunto antes da próxima etapa.", "AV"),
    ]
    summary = TimeSummary(
        av_s=3,
        nav_s=6,
        d_s=0,
        total_s=9,
        av_percent=33.3,
        nav_percent=66.7,
        d_percent=0,
        folga_vs_takt_s=51,
    )
    return OperationalAnalysis(
        metadata=AnalysisMetadata(
            departamento="Função 5",
            posto="5.2.6",
            processo="Oitão XT-8X2",
            responsavel="Rafael",
            data_analise="2026-05-28",
            takt_time_s=60,
        ),
        microetapas=steps,
        resumo_tempos=summary,
    )


def _export(tmp_name: str) -> Path:
    output = REPO_ROOT / "data" / "outputs" / tmp_name
    write_analysis_to_template(
        _non_pmgs_analysis(),
        str(TEMPLATE),
        str(output),
        fill_standard=True,
        standard_export_mode="standard_consolidado",
    )
    return output


def test_first_sheet_is_standard_scania_layout_not_simple_table():
    output = _export("test_standard_layout_export.xlsx")
    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        assert workbook.sheetnames[:3] == ["STANDARD_CONSOLIDADO", "Sheet2", "ENTENDIMENTO_CONVERSAO"]
        assert sheet["A1"].value == "SCANIA"
        assert "For internal use only" in sheet["A3"].value
        assert sheet["A1"].value != "Sequencia"
        assert "A4:A8" in {str(item) for item in sheet.merged_cells.ranges}
        assert sheet.column_dimensions["A"].width == pytest.approx(
            load_workbook(TEMPLATE, data_only=False, keep_links=True)["Standard (1-10)"].column_dimensions["A"].width
        )
    finally:
        workbook.close()


def test_first_sheet_standard_metadata_and_times():
    output = _export("test_standard_layout_metadata.xlsx")
    workbook = load_workbook(output, data_only=True)
    try:
        sheet = workbook["STANDARD_CONSOLIDADO"]
        assert sheet["D5"].value == "Função 5"
        assert sheet["E5"].value == "5.2.6"
        assert sheet["D7"].value == "Rafael"
        assert sheet["E7"].value == "2026-05-28"
        assert sheet["D8"].value == "Oitão XT-8X2"
        assert [sheet.cell(row=row, column=7).value for row in range(9, 12)] == [2, 4, 3]
        assert [sheet.cell(row=row, column=8).value for row in range(9, 12)] == [2, 6, 9]
        assert sheet["D9"].value.startswith("Preparar")
    finally:
        workbook.close()


def test_no_old_pmgs_data_visible_for_non_pmgs_process():
    output = _export("test_no_old_pmgs_visible.xlsx")
    workbook = load_workbook(output, data_only=True)
    try:
        forbidden = ("pmgs.p1", "pré montagem", "pre montagem", "grade superior")
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                continue
            text = " ".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None).casefold()
            assert not any(term in text for term in forbidden), sheet.title
        assert workbook["Standard (1-10)"].sheet_state == "hidden"
    finally:
        workbook.close()


def test_sheet2_still_correct_after_standard_layout_export():
    output = _export("test_sheet2_still_correct_after_standard.xlsx")
    workbook = load_workbook(output, data_only=True)
    try:
        sheet = workbook["Sheet2"]
        headers = [sheet.cell(row=1, column=column).value for column in range(1, 10)]
        assert headers == [
            "id_AvNavD",
            "activity",
            "reminder",
            "id_safe_icon",
            "timeOfElement",
            "type_document",
            "id_takt",
            "id_symbol",
            "title",
        ]
        assert sheet.cell(row=2, column=5).value == 2
        assert sheet.cell(row=3, column=5).value == 4
        assert sheet.cell(row=2, column=3).value is None
        assert sheet.cell(row=2, column=6).value == "SPS"
    finally:
        workbook.close()
