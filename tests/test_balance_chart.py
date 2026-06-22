"""Testes da geracao e insercao do grafico de balanceamento."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.excel.balance_chart import (
    DEFAULT_BALANCE_CHART_SHEET,
    generate_balance_chart_image,
    insert_balance_chart,
)
from app.schemas.analysis import OperationalAnalysis


REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_JSON = REPO_ROOT / "data" / "outputs" / "sample_analysis_pmgs_p1.json"
TEST_OUTPUT_DIR = REPO_ROOT / "data" / "outputs"


def load_sample_analysis() -> OperationalAnalysis:
    data = json.loads(SAMPLE_JSON.read_text(encoding="utf-8"))
    return OperationalAnalysis.model_validate(data)


def test_generate_balance_chart_image_creates_png():
    analysis = load_sample_analysis()
    output_png = TEST_OUTPUT_DIR / "balance_chart.png"

    result = generate_balance_chart_image(analysis, str(output_png))

    assert Path(result) == output_png
    assert output_png.exists()
    assert output_png.stat().st_size > 1000


def test_insert_balance_chart_keeps_workbook_openable():
    analysis = load_sample_analysis()
    image_path = TEST_OUTPUT_DIR / "balance_chart_insert.png"
    workbook_path = TEST_OUTPUT_DIR / "workbook.xlsx"

    generate_balance_chart_image(analysis, str(image_path))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_BALANCE_CHART_SHEET
    sheet["A1"] = "conteudo preservado"
    workbook.save(workbook_path)
    workbook.close()

    insert_balance_chart(str(workbook_path), str(image_path))

    reopened = load_workbook(workbook_path, data_only=False, keep_links=True)
    assert DEFAULT_BALANCE_CHART_SHEET in reopened.sheetnames
    assert reopened[DEFAULT_BALANCE_CHART_SHEET]["A1"].value == "conteudo preservado"
    assert len(reopened[DEFAULT_BALANCE_CHART_SHEET]._images) == 1
    reopened.close()


def test_insert_balance_chart_missing_sheet_does_not_fail():
    analysis = load_sample_analysis()
    image_path = TEST_OUTPUT_DIR / "balance_chart_missing_sheet.png"
    workbook_path = TEST_OUTPUT_DIR / "workbook_sem_aba.xlsx"

    generate_balance_chart_image(analysis, str(image_path))

    workbook = Workbook()
    workbook.active.title = "Outra aba"
    workbook.save(workbook_path)
    workbook.close()

    insert_balance_chart(str(workbook_path), str(image_path))

    reopened = load_workbook(workbook_path, data_only=False, keep_links=True)
    assert reopened.sheetnames == ["Outra aba"]
    reopened.close()
