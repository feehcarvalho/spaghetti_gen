from __future__ import annotations

import inspect
import json
import subprocess
import sys
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.analysis.quality_alerts import ALERTS_SHEET_NAME
from app.analysis.sps_validator import assert_analysis_can_generate_excel
from app.excel.template_writer import write_analysis_to_template
from app.schemas.analysis import OperationalAnalysis, TimeSummary
from tests.test_export_with_quality_alerts import (
    REPO_ROOT,
    analysis_with_steps,
    metadata,
    output_file,
    read_workbook,
    resolve_template,
    step,
)


def _alert_descriptions(path: Path) -> list[str]:
    workbook = read_workbook(path)
    try:
        sheet = workbook[ALERTS_SHEET_NAME]
        return [str(sheet.cell(row=row, column=5).value or "") for row in range(5, sheet.max_row + 1)]
    finally:
        workbook.close()


def test_force_export_with_d_step_without_improvement_generates_excel():
    analysis = analysis_with_steps(
        [step(31, "Aguardar liberação do sistema no ponto observado.")],
        alerts=["Microetapa D 31 sem melhoria/alerta vinculado."],
    )
    output = output_file("force_d_sem_melhoria.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    assert output.exists()
    assert any("sem melhoria" in description.casefold() for description in _alert_descriptions(output))


def test_force_export_with_action_in_justification_generates_excel():
    analysis = analysis_with_steps(
        [
            step(
                99,
                "Realizar atividade operacional.",
                justification="Fixar a porca no eixo com parafusadeira pneumática.",
            )
        ]
    )
    output = output_file("force_acao_justificativa.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    assert output.exists()
    assert any("justificativa" in description.casefold() for description in _alert_descriptions(output))


def test_force_export_with_conflicting_tool_generates_excel():
    analysis = analysis_with_steps(
        [step(15, "Fixar as porcas com a parafusadeira pneumática.")],
        alerts=["Microetapa 15: ferramenta pneumática conflitante com método manual observado/contextual."],
    )
    output = output_file("force_ferramenta_conflitante.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    assert output.exists()
    assert any("conflitante" in description.casefold() for description in _alert_descriptions(output))


def test_none_analysis_blocks_as_fatal_error():
    with pytest.raises(ValueError, match="analise valida"):
        assert_analysis_can_generate_excel(None)


def test_empty_microsteps_blocks_as_fatal_error():
    empty = OperationalAnalysis.model_construct(
        metadata=metadata(),
        microetapas=[],
        resumo_tempos=TimeSummary(av_s=0, nav_s=0, d_s=0, total_s=0, av_percent=0, nav_percent=0, d_percent=0),
        melhorias=[],
        recomendacoes_gerais=[],
        alertas_validacao=[],
        alertas_validacao_sps=[],
        roteiro_operacional=[],
    )

    with pytest.raises(ValueError, match="microetapas validas"):
        assert_analysis_can_generate_excel(empty)


def test_sheet2_remains_second_when_export_has_alerts():
    analysis = analysis_with_steps(
        [step(1, "Posicionar o componente no conjunto confirmado.")],
        alerts=["Microetapa 1: eixo/lado não confirmado no vídeo."],
    )
    output = output_file("force_sheet2_segunda.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        assert workbook.sheetnames[1] == "Sheet2"
    finally:
        workbook.close()


def test_alerts_sheet_is_last_when_export_has_alerts():
    analysis = analysis_with_steps(
        [step(1, "Posicionar o componente no conjunto confirmado.")],
        alerts=["Microetapa 1: quantidade não confirmada no vídeo."],
    )
    output = output_file("force_alertas_ultima.xlsx")

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        assert workbook.sheetnames[-1] == ALERTS_SHEET_NAME
    finally:
        workbook.close()


def test_generate_excel_from_recovery_script_generates_excel():
    analysis = analysis_with_steps(
        [step(1, "Posicionar a ferramenta no ponto de trabalho confirmado.")],
        alerts=["Microetapa 1: Ferramenta não confirmada no vídeo; validar no gemba."],
    )
    recovery_dir = REPO_ROOT / "data" / "outputs" / "test_force_recovery" / uuid.uuid4().hex
    recovery_dir.mkdir(parents=True, exist_ok=True)
    recovery = recovery_dir / "analise_preservada_teste.json"
    recovery.write_text(
        json.dumps({"erro_exportacao": "quality gate", "analysis": analysis.model_dump(mode="json")}),
        encoding="utf-8",
    )

    completed = subprocess.run(
        [
            sys.executable,
            str(REPO_ROOT / "tools" / "generate_excel_from_recovery.py"),
            str(recovery),
            "--output-dir",
            str(recovery_dir),
        ],
        cwd=REPO_ROOT,
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert list(recovery_dir.glob("analise_preservada_teste_*.xlsx"))


def test_export_function_does_not_raise_for_sps_alerts():
    source = inspect.getsource(write_analysis_to_template)
    assert "block_on_quality=quality_gate_blocks_excel and not effective_force_export" in source
    assert "raise ValueError(\"Excel bloqueado pelo quality gate SPS" not in source


def test_ui_warning_message_replaces_blocking_quality_error():
    source = (REPO_ROOT / "app" / "ui" / "streamlit_app.py").read_text(encoding="utf-8")
    assert "A planilha foi gerada com alertas de validacao SPS" in source
    assert "ALERTAS_VALIDACAO_SPS" in source
