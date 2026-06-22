from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.analysis.activity_text import get_microstep_activity_text
from app.analysis.export_preparer import prepare_analysis_for_export
from app.analysis.microstep_consolidator import consolidate_redundant_microsteps, should_merge_microsteps
from app.analysis.operational_language_repair import repair_activity_text, validate_operational_text
from app.analysis.quality_gate import (
    detect_action_hidden_in_justification,
    detect_unsupported_tool_or_method_claims,
)
from app.analysis.time_auditor import audit_and_recalculate_times, calculate_av_nav_d_totals
from app.excel.conversion_sheet_writer import CONVERSION_HEADERS, ensure_conversion_sheets
from app.excel.standard_writer import write_standard_consolidado_sheet
from app.knowledge.xlsx_memory_reader import read_xlsx_memory
from app.schemas.analysis import AnalysisMetadata, MicroStep, OperationalAnalysis, TimeSummary


REPO_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"


def _metadata() -> AnalysisMetadata:
    return AnalysisMetadata(
        departamento="FA",
        posto="5.2.6",
        processo="Processo operacional teste",
        responsavel="SPS",
        data_analise="2026-05-27",
        takt_time_s=60,
    )


def _step(number: int, start: float, end: float, text: str, cls: str = "NAV", **kwargs) -> MicroStep:
    return MicroStep(
        numero=number,
        inicio_s=start,
        fim_s=end,
        duracao_s=end - start,
        inicio_formatado="00:00",
        fim_formatado="00:00",
        duracao_formatada="00:00",
        tempo_acumulado_s=end,
        etapa_detalhada=text,
        instrucao_operacional=kwargs.pop("instrucao_operacional", text),
        classificacao=cls,
        justificativa_tecnica=kwargs.pop(
            "justificativa_tecnica",
            "Etapa classificada porque e necessaria ao metodo observado.",
        ),
        confianca=kwargs.pop("confianca", 0.9),
        baixa_confianca_motivo=kwargs.pop("baixa_confianca_motivo", None),
        **kwargs,
    )


def _analysis(steps: list[MicroStep]) -> OperationalAnalysis:
    summary = TimeSummary(av_s=0, nav_s=sum(s.duracao_s for s in steps), d_s=0, total_s=sum(s.duracao_s for s in steps), av_percent=0, nav_percent=100, d_percent=0)
    return OperationalAnalysis(metadata=_metadata(), microetapas=steps, resumo_tempos=summary)


def test_time_auditor_uses_element_and_progressive_accumulated():
    analysis = _analysis([
        _step(1, 0, 2, "Pegar o componente no ponto indicado."),
        _step(2, 5, 9, "Posicionar o componente na bancada."),
        _step(3, 20, 23, "Conferir o componente montado."),
    ])
    audited = audit_and_recalculate_times(analysis)
    assert [s.duracao_s for s in audited.microetapas] == [2, 4, 3]
    assert [s.tempo_acumulado_s for s in audited.microetapas] == [2, 6, 9]
    summary = calculate_av_nav_d_totals(audited.microetapas)
    assert summary.total_s == pytest.approx(summary.av_s + summary.nav_s + summary.d_s)


def test_activity_never_uses_justification():
    step = _step(
        1,
        0,
        2,
        "Posicionar a parafusadeira pneumática na porca do segundo eixo LD.",
        justificativa_tecnica="Etapa classificada como NAV porque prepara a remoção da porca, mas não transforma diretamente o produto.",
    )
    assert get_microstep_activity_text(step).startswith("Posicionar a parafusadeira")

    bad = _analysis([
        _step(
            1,
            0,
            2,
            "Posicionar ferramenta no ponto indicado.",
            justificativa_tecnica="Fixar a porca no eixo com parafusadeira pneumática.",
        )
    ])
    assert detect_action_hidden_in_justification(bad)


def test_sheet2_contract_fixed():
    analysis = audit_and_recalculate_times(_analysis([
        _step(1, 0, 2, "Fixar o componente no conjunto.", "AV"),
        _step(2, 5, 9, "Posicionar a ferramenta de apoio no ponto de controle.", "NAV"),
        _step(3, 20, 23, "Aguardar liberação do sistema.", "D"),
    ]))
    workbook = Workbook()
    workbook.active.title = "STANDARD_TBONE"
    ensure_conversion_sheets(workbook, analysis)
    assert workbook.sheetnames[1] == "Sheet2"
    sheet = workbook["Sheet2"]
    assert [sheet.cell(1, c).value for c in range(1, 10)] == CONVERSION_HEADERS
    assert [sheet.cell(r, 1).value for r in range(2, 5)] == [211, 212, 213]
    assert [sheet.cell(r, 5).value for r in range(2, 5)] == [2, 4, 3]
    assert sheet.cell(3, 5).value != 6
    assert sheet.max_column == 9
    assert [sheet.cell(r, 3).value for r in range(2, 5)] == [None, None, None]
    assert [sheet.cell(r, 4).value for r in range(2, 5)] == [None, None, None]
    assert [sheet.cell(r, 6).value for r in range(2, 5)] == ["SPS", "SPS", "SPS"]
    assert [sheet.cell(r, 7).value for r in range(2, 5)] == [60, 60, 60]
    assert [sheet.cell(r, 8).value for r in range(2, 5)] == [None, None, None]
    assert "classificada como" not in sheet.cell(2, 2).value.casefold()
    assert sheet.cell(2, 9).value == analysis.metadata.processo


def test_standard_consolidated_times():
    analysis = audit_and_recalculate_times(_analysis([
        _step(1, 0, 2, "Pegar o componente no ponto indicado."),
        _step(2, 5, 9, "Posicionar o componente no conjunto."),
        _step(3, 20, 23, "Conferir o componente montado."),
    ]))
    workbook = load_workbook(TEMPLATE, data_only=False, keep_links=True)
    write_standard_consolidado_sheet(workbook, analysis)
    sheet = workbook["STANDARD_CONSOLIDADO"]
    assert [sheet.cell(r, 7).value for r in range(9, 12)] == [2, 4, 3]
    assert [sheet.cell(r, 8).value for r in range(9, 12)] == [2, 6, 9]


def test_repetition_consolidation_preserves_total_and_evidence():
    analysis = audit_and_recalculate_times(_analysis([
        _step(1, 0, 1, "Pegar o pneu.", peca_componente="pneu", evidencia_visual="pega"),
        _step(2, 1, 2, "Movimentar o pneu.", peca_componente="pneu", evidencia_visual="move"),
        _step(3, 2, 4, "Levar o pneu até o eixo indicado.", peca_componente="pneu", evidencia_visual="leva"),
    ]))
    assert should_merge_microsteps(analysis.microetapas[0], analysis.microetapas[1])
    consolidated = consolidate_redundant_microsteps(analysis)
    assert len(consolidated.microetapas) == 1
    assert consolidated.resumo_tempos.total_s == pytest.approx(analysis.resumo_tempos.total_s)
    assert "pega" in (consolidated.microetapas[0].evidencia_visual or "")


def test_unsupported_tool_claims_and_context_allowance():
    analysis = _analysis([_step(1, 0, 2, "Fixar as porcas com a parafusadeira pneumática.")])
    assert detect_unsupported_tool_or_method_claims(analysis, "")
    assert not detect_unsupported_tool_or_method_claims(analysis, "Memoria confirma parafusadeira pneumática nesta etapa.")
    manual = prepare_analysis_for_export(analysis, "instalação manual observada")
    assert "parafusadeira pneum" not in get_microstep_activity_text(manual.microetapas[0]).casefold()


def test_xlsx_memory_reader_reads_multiple_sheets():
    path = Path("data/outputs/test_xlsx_memory_reader_memoria.xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Processo"
    workbook.active["A1"] = "Descrição do processo de montagem"
    sheet = workbook.create_sheet("Mapeamento")
    sheet["A1"] = "Sheet2 id_AvNavD activity timeOfElement"
    workbook.save(path)
    content = read_xlsx_memory(str(path))
    assert content.sheet_names == ["Processo", "Mapeamento"]
    assert content.process_descriptions
    assert content.mapping_rules


def test_operational_tone_preserve_reject_and_rewrite():
    good = "Rolar o pneu até o eixo indicado."
    assert validate_operational_text(good) == []
    assert repair_activity_text(good) == good
    bureaucratic = "Realizar a movimentação do componente até o ponto de aplicação."
    assert any("burocrática" in alert for alert in validate_operational_text(bureaucratic))
    assert repair_activity_text(bureaucratic, context={"componente": "pneu", "local": "eixo indicado"}) == "Rolar o pneu até o eixo indicado."
    colloquial = "Bota o pneu lá no eixo."
    assert validate_operational_text(colloquial)
    assert repair_activity_text(colloquial) == "Colocar o pneu no eixo indicado."
