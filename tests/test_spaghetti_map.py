"""Testes do mapa de espaguete."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.schemas.analysis import OperationalAnalysis, SpaghettiData, SpaghettiMove
from app.spaghetti.map_generator import (
    DEFAULT_SPAGHETTI_SHEET,
    generate_spaghetti_map_image,
    insert_spaghetti_map,
    load_layout,
)


REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_JSON = REPO_ROOT / "data" / "outputs" / "sample_analysis_pmgs_p1.json"
LAYOUT_JSON = REPO_ROOT / "data" / "layouts" / "PMGS.P1.json"
TEST_OUTPUT_DIR = REPO_ROOT / "data" / "outputs"


def load_sample_analysis() -> OperationalAnalysis:
    data = json.loads(SAMPLE_JSON.read_text(encoding="utf-8"))
    return OperationalAnalysis.model_validate(data)


def analysis_with_moves(moves: list[SpaghettiMove]) -> OperationalAnalysis:
    analysis = load_sample_analysis()
    spaghetti = SpaghettiData(
        layout_id="PMGS.P1",
        pontos=[],
        movimentos=moves,
        total_passos_estimados=sum(move.passos_estimados or 0 for move in moves),
        distancia_total_m=sum(move.distancia_m or 0 for move in moves),
    )
    return analysis.model_copy(update={"spaghetti": spaghetti, "alertas_validacao": []})


def test_load_layout_reads_example_layout():
    layout = load_layout(str(LAYOUT_JSON))

    assert layout["layout_id"] == "PMGS.P1"
    assert "produto" in layout["locais"]
    assert layout["locais"]["wpo"]["x"] == 1


def test_generate_spaghetti_map_image_draws_valid_moves():
    analysis = analysis_with_moves(
        [
            SpaghettiMove(
                ordem=1,
                origem="wpo",
                destino="produto",
                passos_estimados=4,
                distancia_m=2.0,
                motivo="Consultar WPO e retornar ao produto",
            ),
            SpaghettiMove(
                ordem=2,
                origem="produto",
                destino="carrinho_farol",
                passos_estimados=6,
                distancia_m=3.0,
                motivo="Buscar farol",
            ),
        ]
    )
    output_png = TEST_OUTPUT_DIR / "test_spaghetti_map.png"

    result = generate_spaghetti_map_image(analysis, str(LAYOUT_JSON), str(output_png))

    assert Path(result) == output_png
    assert output_png.exists()
    assert output_png.stat().st_size > 1000
    assert analysis.alertas_validacao == []


def test_generate_spaghetti_map_image_warns_and_skips_invalid_move():
    analysis = analysis_with_moves(
        [
            SpaghettiMove(
                ordem=1,
                origem="local_inexistente",
                destino="produto",
                passos_estimados=3,
                motivo="Movimento invalido para teste",
            )
        ]
    )
    output_png = TEST_OUTPUT_DIR / "test_spaghetti_invalid.png"

    generate_spaghetti_map_image(analysis, str(LAYOUT_JSON), str(output_png))

    assert output_png.exists()
    assert any("Movimento spaghetti 1 ignorado" in alert for alert in analysis.alertas_validacao)


def test_insert_spaghetti_map_keeps_workbook_valid():
    analysis = analysis_with_moves(
        [
            SpaghettiMove(
                ordem=1,
                origem="vr_pega",
                destino="vr_montagem",
                passos_estimados=2,
                motivo="Posicionar VR",
            )
        ]
    )
    image_path = TEST_OUTPUT_DIR / "test_spaghetti_insert.png"
    workbook_path = TEST_OUTPUT_DIR / "test_spaghetti_workbook.xlsx"

    generate_spaghetti_map_image(analysis, str(LAYOUT_JSON), str(image_path))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_SPAGHETTI_SHEET
    sheet["A1"] = "conteudo preservado"
    workbook.save(workbook_path)
    workbook.close()

    insert_spaghetti_map(str(workbook_path), str(image_path))

    reopened = load_workbook(workbook_path, data_only=False, keep_links=True)
    assert reopened[DEFAULT_SPAGHETTI_SHEET]["A1"].value == "conteudo preservado"
    assert len(reopened[DEFAULT_SPAGHETTI_SHEET]._images) == 1
    reopened.close()
