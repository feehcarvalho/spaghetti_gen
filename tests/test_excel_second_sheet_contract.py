from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.analysis.time_auditor import recalculate_microstep_times
from app.excel.conversion_sheet_writer import CONVERSION_HEADERS
from app.excel.template_writer import write_analysis_to_template
from app.schemas.analysis import OperationalAnalysis


REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_JSON = REPO_ROOT / "data" / "outputs" / "sample_analysis_pmgs_p1.json"
DEFAULT_TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
FALLBACK_TEMPLATE = REPO_ROOT / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"


def resolve_template() -> Path:
    if DEFAULT_TEMPLATE.exists():
        return DEFAULT_TEMPLATE
    if FALLBACK_TEMPLATE.exists():
        return FALLBACK_TEMPLATE
    pytest.skip("Template Excel real nao encontrado no workspace")


def load_sample_analysis() -> OperationalAnalysis:
    return OperationalAnalysis.model_validate(json.loads(SAMPLE_JSON.read_text(encoding="utf-8")))


def test_sheet2_is_second_sheet_with_exact_headers():
    output = REPO_ROOT / "data" / "outputs" / "test_sheet2_contract.xlsx"
    analysis = load_sample_analysis()

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        assert workbook.sheetnames[1] == "Sheet2"
        assert workbook.sheetnames[2] == "ENTENDIMENTO_CONVERSAO"
        sheet = workbook["Sheet2"]
        headers = [sheet.cell(row=1, column=column).value for column in range(1, 10)]
        assert headers == CONVERSION_HEADERS
        assert sheet.max_column == 9
        assert not sheet.merged_cells.ranges
        assert not getattr(sheet, "_images", [])
        assert not getattr(sheet, "_charts", [])
    finally:
        workbook.close()


def test_sheet2_maps_av_nav_d_and_element_time():
    output = REPO_ROOT / "data" / "outputs" / "test_sheet2_values.xlsx"
    analysis = recalculate_microstep_times(load_sample_analysis())

    write_analysis_to_template(analysis, str(resolve_template()), str(output))

    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        sheet = workbook["Sheet2"]
        headers = [sheet.cell(row=1, column=column).value for column in range(1, 10)]
        for row_index, step in enumerate(analysis.microetapas, start=2):
            assert sheet.cell(row=row_index, column=1).value == {"AV": 211, "NAV": 212, "D": 213}[step.classificacao]
            assert sheet.cell(row=row_index, column=5).value == pytest.approx(step.duracao_s)
            assert "timeOfElement" == sheet.cell(row=1, column=5).value
            assert sheet.cell(row=row_index, column=3).value is None
            assert sheet.cell(row=row_index, column=4).value is None
            assert sheet.cell(row=row_index, column=6).value == "SPS"
            assert sheet.cell(row=row_index, column=7).value == analysis.metadata.takt_time_s
            assert sheet.cell(row=row_index, column=8).value is None
            assert sheet.cell(row=row_index, column=9).value == analysis.metadata.processo
        forbidden_headers = {"inicio_s", "fim_s", "duracao_formatada", "justificativa_tecnica", "melhoria"}
        assert forbidden_headers.isdisjoint(set(headers))
    finally:
        workbook.close()
