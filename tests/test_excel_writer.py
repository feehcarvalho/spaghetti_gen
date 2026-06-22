"""Testes do preenchimento do template Excel."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.excel.template_writer import write_analysis_to_template
from app.schemas.analysis import OperationalAnalysis


REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_JSON = REPO_ROOT / "data" / "outputs" / "sample_analysis_pmgs_p1.json"
DEFAULT_TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
FALLBACK_TEMPLATE = REPO_ROOT / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"


def resolve_template() -> Path:
    if DEFAULT_TEMPLATE.exists():
        return DEFAULT_TEMPLATE
    if FALLBACK_TEMPLATE.exists():
        return FALLBACK_TEMPLATE
    pytest.skip("Template Excel real nao encontrado no workspace")


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_sample_analysis() -> OperationalAnalysis:
    data = json.loads(SAMPLE_JSON.read_text(encoding="utf-8"))
    return OperationalAnalysis.model_validate(data)


def test_write_analysis_to_template_preserves_workbook_and_fills_sheets():
    template_path = resolve_template()
    output_path = REPO_ROOT / "data" / "outputs" / "test_excel_writer_output.xlsx"
    analysis = load_sample_analysis()
    original_hash = file_hash(template_path)
    original_wb = load_workbook(template_path, data_only=False, keep_links=True)
    original_sheetnames = list(original_wb.sheetnames)
    preserved_formulas = {
        ("Standard (1-10)", "L10"): original_wb["Standard (1-10)"]["L10"].value,
        ("Worktable", "I5"): original_wb["Worktable"]["I5"].value,
        ("Gráfico Workload", "E3"): original_wb["Gráfico Workload"]["E3"].value,
        ("Diagrama de Espaguete", "C3"): original_wb["Diagrama de Espaguete"]["C3"].value,
    }

    result = write_analysis_to_template(
        analysis=analysis,
        template_path=str(template_path),
        output_path=str(output_path),
    )

    assert Path(result) == output_path
    assert output_path.exists()
    assert file_hash(template_path) == original_hash

    generated_wb = load_workbook(output_path, data_only=False, keep_links=True)
    assert generated_wb.sheetnames[1] == "Sheet2"
    assert generated_wb.sheetnames[2] == "ENTENDIMENTO_CONVERSAO"
    assert [name for name in generated_wb.sheetnames if name in original_sheetnames] == original_sheetnames

    analysis_sheet = generated_wb["ANÁLISE"]
    assert analysis_sheet["A1"].value == "ANÁLISE DETALHADA DO VÍDEO – PMGS.P1"
    assert analysis_sheet["B3"].value == "PMGS"
    assert analysis_sheet["E3"].value == "Engenharia de Processos"
    assert analysis_sheet["A7"].value == 1
    assert analysis_sheet["B7"].value == "Verificar WPO e confirmar produto/variante no monitor do posto"
    assert analysis_sheet["G7"].value == "NAV"
    assert analysis_sheet["A16"].value == 10
    assert analysis_sheet["B16"].value == "Posicionar VR no proximo posto para liberar a sequencia"

    improvements_sheet = generated_wb["MELHORIAS"]
    assert improvements_sheet["B3"].value == "84s (01:24)"
    assert improvements_sheet["B4"].value == "15s (17.9%)"
    assert improvements_sheet["A7"].value == 3
    assert improvements_sheet["A8"].value == 9
    assert improvements_sheet["H8"].value == "Média"
    assert improvements_sheet["A14"].value == "RECOMENDAÇÕES GERAIS"
    assert improvements_sheet["A15"].value.startswith("• ")

    for (sheet_name, coordinate), formula in preserved_formulas.items():
        assert generated_wb[sheet_name][coordinate].value == formula


def test_write_analysis_to_template_includes_validation_warnings_sheet():
    template_path = resolve_template()
    output_path = REPO_ROOT / "data" / "outputs" / "test_excel_writer_warnings_output.xlsx"
    analysis = load_sample_analysis()
    analysis.microetapas[0].confianca = 0.65

    result = write_analysis_to_template(
        analysis=analysis,
        template_path=str(template_path),
        output_path=str(output_path),
    )

    generated_wb = load_workbook(result, data_only=False, keep_links=True)
    assert "Avisos_Validacao" in generated_wb.sheetnames
    warnings_sheet = generated_wb["Avisos_Validacao"]
    assert warnings_sheet["A2"].value == 1
    assert warnings_sheet["C2"].value == "confianca"
    assert warnings_sheet["D2"].value == "Baixa - validar no gemba"
    assert "Confiança abaixo de 0.7" in warnings_sheet["E2"].value
