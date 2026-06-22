from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.analysis.time_auditor import recalculate_microstep_times
from app.excel.template_writer import write_analysis_to_template
from app.schemas.analysis import OperationalAnalysis


REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_JSON = REPO_ROOT / "data" / "outputs" / "sample_analysis_pmgs_p1.json"
DEFAULT_TEMPLATE = REPO_ROOT / "data" / "templates" / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
FALLBACK_TEMPLATE = REPO_ROOT / "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"


def resolve_template() -> Path:
    if DEFAULT_TEMPLATE.exists():
        return DEFAULT_TEMPLATE
    if FALLBACK_TEMPLATE.exists():
        return FALLBACK_TEMPLATE
    pytest.skip("Template Excel real nao encontrado no workspace")


def load_sample_analysis() -> OperationalAnalysis:
    return OperationalAnalysis.model_validate(json.loads(SAMPLE_JSON.read_text(encoding="utf-8")))


def test_standard_consolidado_is_first_sheet_and_continuous():
    output = REPO_ROOT / "data" / "outputs" / "test_standard_consolidado.xlsx"
    template = resolve_template()
    analysis = recalculate_microstep_times(load_sample_analysis())

    write_analysis_to_template(
        analysis,
        str(template),
        str(output),
        fill_standard=True,
        standard_export_mode="standard_consolidado",
    )

    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        assert workbook.sheetnames[0] == "STANDARD_CONSOLIDADO"
        assert workbook.sheetnames[1] == "Sheet2"
        sheet = workbook["STANDARD_CONSOLIDADO"]
        assert sheet["A1"].value == "SCANIA"
        assert "For internal use only" in sheet["A3"].value
        assert sheet["A9"].value == 1
        assert sheet["A18"].value == 10
        assert sheet["G9"].value == pytest.approx(analysis.microetapas[0].duracao_s)
        assert sheet["H9"].value == pytest.approx(analysis.microetapas[0].tempo_acumulado_s)
        assert sheet["H18"].value == pytest.approx(analysis.microetapas[9].tempo_acumulado_s)
    finally:
        workbook.close()


def test_standard_consolidado_does_not_fill_legacy_standard_sheets():
    output = REPO_ROOT / "data" / "outputs" / "test_standard_consolidado_no_split.xlsx"
    template = resolve_template()
    analysis = load_sample_analysis()
    original = load_workbook(template, data_only=False, keep_links=True)
    original_a9 = original["Standard (1-10)"]["A9"].value
    original.close()

    write_analysis_to_template(
        analysis,
        str(template),
        str(output),
        fill_standard=True,
        standard_export_mode="standard_consolidado",
    )

    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        assert workbook["Standard (1-10)"]["A9"].value == original_a9
        assert workbook["Standard (1-10)"].sheet_state == "hidden"
    finally:
        workbook.close()
