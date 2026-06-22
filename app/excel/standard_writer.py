"""Preenchimento controlado das abas Standard do template SPS."""

from __future__ import annotations

import logging
import re
from copy import copy
from datetime import time
from typing import Iterable

from openpyxl.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.analysis.export_steps import NormalizedExportStep, build_normalized_export_steps
from app.schemas.analysis import OperationalAnalysis


logger = logging.getLogger(__name__)

STANDARD_PREFIX = "Standard"
STANDARD_TEMPLATE_SHEET = "Standard (1-10)"
STANDARD_CONSOLIDATED_SHEET = "STANDARD_CONSOLIDADO"
FIRST_DATA_ROW = 9
MAPPED_COLUMNS = (1, 2, 3, 4, 5, 7)


def _natural_standard_key(sheet_name: str) -> tuple[int, str]:
    match = re.search(r"\((\d+)", sheet_name)
    if match:
        return int(match.group(1)), sheet_name
    return 9999, sheet_name


def _standard_sheets(workbook: Workbook) -> list[Worksheet]:
    names = sorted(
        (name for name in workbook.sheetnames if name.startswith(STANDARD_PREFIX)),
        key=_natural_standard_key,
    )
    return [workbook[name] for name in names]


def _writable_cell(sheet: Worksheet, row: int, column: int):
    cell = sheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return None
    return cell


def _set_if_writable(sheet: Worksheet, coordinate: str, value) -> None:
    cell = sheet[coordinate]
    if not isinstance(cell, MergedCell):
        cell.value = value


def _excel_time_from_seconds(seconds: float | None):
    if seconds is None:
        return None

    total_seconds = int(round(seconds))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    remaining_seconds = total_seconds % 60

    if hours >= 24:
        return seconds / 86400

    return time(hour=hours, minute=minutes, second=remaining_seconds)


def _date_value(value: str):
    return value


def _find_data_end_row(sheet: Worksheet) -> int:
    merged_starts = []
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, _ = merged_range.bounds
        if min_row >= FIRST_DATA_ROW and min_col <= 1 and max_col >= 3:
            merged_starts.append(min_row)

    if merged_starts:
        return min(merged_starts) - 1

    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        first_value = sheet.cell(row=row, column=1).value
        if isinstance(first_value, str) and first_value.strip().casefold().startswith("work load"):
            return row - 1

    return sheet.max_row


def _data_rows(sheet: Worksheet) -> range:
    end_row = _find_data_end_row(sheet)
    if end_row < FIRST_DATA_ROW:
        return range(FIRST_DATA_ROW, FIRST_DATA_ROW)
    return range(FIRST_DATA_ROW, end_row + 1)


def _clear_mapped_rows(sheet: Worksheet, rows: Iterable[int]) -> None:
    for row in rows:
        for column in MAPPED_COLUMNS:
            cell = _writable_cell(sheet, row, column)
            if cell is not None:
                cell.value = None


def _write_metadata(sheet: Worksheet, analysis: OperationalAnalysis) -> None:
    metadata = analysis.metadata
    _set_if_writable(sheet, "D5", metadata.departamento)
    _set_if_writable(sheet, "E5", metadata.posto)
    _set_if_writable(sheet, "D7", metadata.responsavel)
    _set_if_writable(sheet, "E7", _date_value(metadata.data_analise))
    _set_if_writable(sheet, "D8", metadata.processo)
    _set_if_writable(sheet, "F2", _excel_time_from_seconds(metadata.takt_time_s))
    _set_if_writable(sheet, "F5", "1/1")
    _set_if_writable(sheet, "F7", "1")


def _write_step(sheet: Worksheet, row: int, export_step: NormalizedExportStep, write_accumulated: bool = False) -> None:
    step = export_step.microstep
    values = {
        1: export_step.sequence,
        2: export_step.classification,
        3: step.numero,
        4: export_step.activity,
        5: _build_step_reminder(step),
        6: "",
        7: export_step.duration_s,
    }
    if write_accumulated:
        values[8] = step.tempo_acumulado_s
    for column, value in values.items():
        cell = _writable_cell(sheet, row, column)
        if cell is not None:
            cell.value = value


def write_standard_sheets(workbook: Workbook, analysis: OperationalAnalysis) -> None:
    """Preenche as abas Standard existentes com microetapas da analise."""

    sheets = _standard_sheets(workbook)
    if not sheets:
        logger.warning("Nenhuma aba Standard encontrada no workbook.")
        return
    export_steps = build_normalized_export_steps(analysis)

    for sheet in sheets:
        rows = list(_data_rows(sheet))
        _write_metadata(sheet, analysis)
        _clear_mapped_rows(sheet, rows)

    step_index = 0
    for sheet in sheets:
        for row in _data_rows(sheet):
            if step_index >= len(export_steps):
                break

            _write_step(sheet, row, export_steps[step_index])
            step_index += 1

        if step_index >= len(export_steps):
            break

    if step_index < len(export_steps):
        logger.warning(
            "Sem abas Standard suficientes: %s de %s microetapas foram escritas.",
            step_index,
            len(export_steps),
        )


def write_standard_consolidado_sheet(workbook: Workbook, analysis: OperationalAnalysis) -> None:
    """Write all microsteps in a complete Scania Standard layout."""

    export_steps = build_normalized_export_steps(analysis)
    sheet = create_standard_complete_sheet_from_template(workbook)
    _move_sheet_to_first(workbook, sheet)
    _write_metadata(sheet, analysis)
    _ensure_data_capacity(sheet, len(export_steps))
    _unmerge_data_area(sheet, FIRST_DATA_ROW, FIRST_DATA_ROW + len(export_steps) - 1)
    rows = list(range(FIRST_DATA_ROW, FIRST_DATA_ROW + len(export_steps)))
    _clear_standard_data_area(sheet, FIRST_DATA_ROW, _find_data_end_row(sheet))

    for row_index, export_step in zip(rows, export_steps):
        _write_step(sheet, row_index, export_step, write_accumulated=True)

    _write_complementary_info(sheet, analysis, FIRST_DATA_ROW + len(export_steps) + 2)
    _hide_legacy_standard_sheets(workbook)


def create_standard_complete_sheet_from_template(
    workbook: Workbook,
    template_sheet_name: str = STANDARD_TEMPLATE_SHEET,
    output_sheet_name: str = STANDARD_CONSOLIDATED_SHEET,
) -> Worksheet:
    """Create the consolidated Standard by cloning the official template layout."""

    if template_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Template sem aba Standard base: {template_sheet_name}")
    if output_sheet_name in workbook.sheetnames:
        del workbook[output_sheet_name]
    source = workbook[template_sheet_name]
    sheet = workbook.copy_worksheet(source)
    sheet.title = output_sheet_name
    _move_sheet_to_first(workbook, sheet)
    return sheet


def _move_sheet_to_first(workbook: Workbook, sheet: Worksheet) -> None:
    sheets = workbook._sheets
    if sheet in sheets:
        sheets.remove(sheet)
    sheets.insert(0, sheet)


def _ensure_data_capacity(sheet: Worksheet, step_count: int) -> None:
    current_capacity = max(0, _find_data_end_row(sheet) - FIRST_DATA_ROW + 1)
    if step_count <= current_capacity:
        return
    insert_at = _find_data_end_row(sheet) + 1
    rows_to_add = step_count - current_capacity
    template_row = insert_at - 1
    sheet.insert_rows(insert_at, rows_to_add)
    for row in range(insert_at, insert_at + rows_to_add):
        _copy_row_style(sheet, template_row, row)


def _unmerge_data_area(sheet: Worksheet, start_row: int, end_row: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        intersects_rows = max_row >= start_row and min_row <= end_row
        intersects_columns = max_col >= 1 and min_col <= 8
        if intersects_rows and intersects_columns:
            sheet.unmerge_cells(str(merged_range))


def _copy_row_style(sheet: Worksheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if isinstance(target, MergedCell):
            continue
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _clear_standard_data_area(sheet: Worksheet, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = _writable_cell(sheet, row, column)
            if cell is not None:
                cell.value = None


def _write_complementary_info(sheet: Worksheet, analysis: OperationalAnalysis, start_row: int) -> None:
    if start_row > sheet.max_row:
        sheet.insert_rows(start_row, 8)
        _copy_row_style(sheet, max(FIRST_DATA_ROW, start_row - 1), start_row)
    rows = [
        ("INFORMACOES COMPLEMENTARES", ""),
        ("Video analisado", analysis.metadata.fonte_video or ""),
        ("Usuario/login responsavel", analysis.metadata.usuario_login or analysis.metadata.usuario_nome or ""),
        ("Memorias utilizadas", _joined_memories(analysis)),
        ("Alertas de baixa confianca", _low_confidence_alerts(analysis)),
        ("Validacao gemba/SPS", "Obrigatoria antes de oficializar como padrao."),
    ]
    for offset, (label, value) in enumerate(rows):
        row = start_row + offset
        for column in range(1, min(sheet.max_column, 8) + 1):
            cell = _writable_cell(sheet, row, column)
            if cell is not None:
                cell.value = None
        _set_row_value(sheet, row, 1, label)
        _set_row_value(sheet, row, 4, value)


def _set_row_value(sheet: Worksheet, row: int, column: int, value) -> None:
    cell = _writable_cell(sheet, row, column)
    if cell is not None:
        cell.value = value


def _joined_memories(analysis: OperationalAnalysis) -> str:
    memories = []
    for step in analysis.microetapas:
        for item in step.memoria_utilizada:
            if item not in memories:
                memories.append(item)
    return " / ".join(memories[:8])[:400]


def _low_confidence_alerts(analysis: OperationalAnalysis) -> str:
    alerts = [
        f"Etapa {step.numero}: {step.baixa_confianca_motivo}"
        for step in analysis.microetapas
        if step.baixa_confianca_motivo
    ]
    return " / ".join(alerts[:5])[:500]


def _hide_legacy_standard_sheets(workbook: Workbook) -> None:
    for sheet in _standard_sheets(workbook):
        if sheet.title != STANDARD_CONSOLIDATED_SHEET:
            sheet.sheet_state = "hidden"


def _build_step_reminder(step) -> str:
    parts = [
        step.ferramenta_observacao,
        step.peca_componente,
        step.dispositivo,
        step.eixo,
        step.lado,
        ", ".join(step.memoria_utilizada[:2]) if step.memoria_utilizada else None,
    ]
    return " / ".join(part.strip() for part in parts if part and part.strip())[:220]


def _clear_all_values(sheet: Worksheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None
