"""Sheet2 conversion contract for the logistics automation app."""

from __future__ import annotations

import os

from openpyxl.cell.cell import MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.analysis.export_steps import NormalizedExportStep, build_normalized_export_steps
from app.schemas.analysis import OperationalAnalysis


CONVERSION_SHEET_NAME = "Sheet2"
UNDERSTANDING_SHEET_NAME = "ENTENDIMENTO_CONVERSAO"
CONVERSION_HEADERS = [
    "id_AvNavD",
    "activity",
    "reminder",
    "id_safe_icon",
    "timeOfElement",
    "type_document",
    "id_takt",
    "id_symbol",
    "title",
]
AV_NAV_D_IDS = {"AV": 211, "NAV": 212, "D": 213}
DEFAULT_TYPE_DOCUMENT = "SPS"


def ensure_conversion_sheets(workbook: Workbook, analysis: OperationalAnalysis) -> None:
    """Write Sheet2 as the second tab and explanatory conversion notes as the third tab."""

    export_steps = build_normalized_export_steps(analysis)

    conversion_sheet = _replace_sheet(workbook, CONVERSION_SHEET_NAME, 1)
    _write_conversion_sheet(conversion_sheet, analysis, export_steps)

    understanding_sheet = _get_or_create_sheet(workbook, UNDERSTANDING_SHEET_NAME)
    _move_sheet_to_index(workbook, understanding_sheet, 2)
    _write_understanding_sheet(understanding_sheet)


def validate_conversion_sheet_contract(workbook: Workbook, analysis: OperationalAnalysis) -> list[str]:
    """Validate the Sheet2 integration contract before saving the workbook."""

    errors: list[str] = []
    if len(workbook.sheetnames) < 2 or workbook.sheetnames[1] != CONVERSION_SHEET_NAME:
        errors.append("Sheet2 deve estar na segunda posicao do workbook.")
        return errors

    sheet = workbook[CONVERSION_SHEET_NAME]
    if sheet.max_column != len(CONVERSION_HEADERS):
        errors.append(f"Sheet2 deve ter exatamente {len(CONVERSION_HEADERS)} colunas.")
    if sheet.merged_cells.ranges:
        errors.append("Sheet2 nao deve conter celulas mescladas.")
    if getattr(sheet, "_images", None):
        errors.append("Sheet2 nao deve conter imagens.")
    if getattr(sheet, "_charts", None):
        errors.append("Sheet2 nao deve conter graficos.")

    headers = [sheet.cell(row=1, column=column).value for column in range(1, len(CONVERSION_HEADERS) + 1)]
    if headers != CONVERSION_HEADERS:
        errors.append(f"Cabecalhos da Sheet2 invalidos: {headers}")

    export_steps = build_normalized_export_steps(analysis)
    expected_last_row = len(export_steps) + 1
    if sheet.max_row != expected_last_row:
        errors.append(f"Sheet2 deve conter cabecalho e {len(export_steps)} linha(s) de microetapa.")

    av_nav_d_ids = _configured_av_nav_d_ids()
    for row_index, export_step in enumerate(export_steps, start=2):
        expected_id = av_nav_d_ids.get(export_step.classification)
        if sheet.cell(row=row_index, column=1).value != expected_id:
            errors.append(f"Sheet2 linha {row_index}: id_AvNavD invalido.")
        if sheet.cell(row=row_index, column=2).value != export_step.activity:
            errors.append(f"Sheet2 linha {row_index}: activity invalida.")
        for column in (3, 4, 8):
            if not _is_blank(sheet.cell(row=row_index, column=column).value):
                errors.append(f"Sheet2 linha {row_index}: coluna {column} deve ficar vazia.")
        time_value = sheet.cell(row=row_index, column=5).value
        if not isinstance(time_value, (int, float)):
            errors.append(f"Sheet2 linha {row_index}: timeOfElement deve ser numerico.")
        elif abs(float(time_value) - export_step.duration_s) > 0.01:
            errors.append(f"Sheet2 linha {row_index}: timeOfElement deve usar tempo individual.")
        if sheet.cell(row=row_index, column=6).value != _type_document_value():
            errors.append(f"Sheet2 linha {row_index}: type_document invalido.")
        if sheet.cell(row=row_index, column=7).value != _takt_value(analysis):
            errors.append(f"Sheet2 linha {row_index}: id_takt invalido.")
        if sheet.cell(row=row_index, column=9).value != _title_value(analysis):
            errors.append(f"Sheet2 linha {row_index}: title invalido.")

    return errors


def _get_or_create_sheet(workbook: Workbook, title: str) -> Worksheet:
    if title in workbook.sheetnames:
        return workbook[title]
    return workbook.create_sheet(title=title)


def _replace_sheet(workbook: Workbook, title: str, index: int) -> Worksheet:
    if title in workbook.sheetnames:
        del workbook[title]
    return workbook.create_sheet(title=title, index=min(index, len(workbook._sheets)))


def _move_sheet_to_index(workbook: Workbook, sheet: Worksheet, index: int) -> None:
    sheets = workbook._sheets  # openpyxl keeps worksheet order here.
    if sheet in sheets:
        sheets.remove(sheet)
    sheets.insert(min(index, len(sheets)), sheet)


def _write_conversion_sheet(
    sheet: Worksheet,
    analysis: OperationalAnalysis,
    export_steps: list[NormalizedExportStep],
) -> None:
    for column, header in enumerate(CONVERSION_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)

    av_nav_d_ids = _configured_av_nav_d_ids()
    for row_index, export_step in enumerate(export_steps, start=2):
        values = [
            av_nav_d_ids[export_step.classification],
            export_step.activity,
            None,
            _optional_env_value("SHEET2_ID_SAFE_ICON"),
            export_step.duration_s,
            _type_document_value(),
            _takt_value(analysis),
            _optional_env_value("SHEET2_ID_SYMBOL"),
            _title_value(analysis),
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)


def _configured_av_nav_d_ids() -> dict[str, int]:
    return {
        "AV": _read_int_env("SHEET2_ID_AV", AV_NAV_D_IDS["AV"]),
        "NAV": _read_int_env("SHEET2_ID_NAV", AV_NAV_D_IDS["NAV"]),
        "D": _read_int_env("SHEET2_ID_D", AV_NAV_D_IDS["D"]),
    }


def _read_int_env(name: str, default: int) -> int:
    value = os.getenv(name)
    if value is None or not value.strip():
        return default
    try:
        return int(value)
    except ValueError:
        return default


def _optional_env_value(name: str) -> str | None:
    value = os.getenv(name)
    if value is None:
        return None
    text = value.strip()
    return text or None


def _type_document_value() -> str:
    return os.getenv("SHEET2_TYPE_DOCUMENT", DEFAULT_TYPE_DOCUMENT).strip() or DEFAULT_TYPE_DOCUMENT


def _takt_value(analysis: OperationalAnalysis) -> int | float | None:
    value = analysis.metadata.takt_time_s
    if value is None:
        return None
    numeric = float(value)
    return int(numeric) if numeric.is_integer() else numeric


def _title_value(analysis: OperationalAnalysis) -> str:
    return analysis.metadata.processo


def _write_understanding_sheet(sheet: Worksheet) -> None:
    _clear_sheet_values(sheet)
    rows = [
        ["ENTENDIMENTO DA CONVERSAO"],
        ["A segunda aba Sheet2 existe para integracao com o app da automacao logistica."],
        ["O app deve ler a segunda aba do workbook; por isso Sheet2 e mantida na posicao 2."],
        [""],
        ["Mapeamento id_AvNavD"],
        [f"AV = {AV_NAV_D_IDS['AV']}"],
        [f"NAV = {AV_NAV_D_IDS['NAV']}"],
        [f"D = {AV_NAV_D_IDS['D']}"],
        [""],
        ["Campo activity: instrucao operacional da microetapa no padrao Scania."],
        ["Campo reminder: vazio por contrato tecnico; nao recebe inicio, fim, acumulado ou justificativa."],
        ["Campo timeOfElement: tempo individual do elemento em segundos; nao usar tempo acumulado."],
        ["Campo type_document: SPS por padrao, configuravel por SHEET2_TYPE_DOCUMENT."],
        ["Campo id_takt: takt time dos metadados da analise."],
        [""],
        ["Pendencias para validacao da automacao logistica"],
        ["Confirmar IDs definitivos de AV/NAV/D quando necessario."],
        ["Confirmar se id_safe_icon deve ficar vazio ou receber ID unico."],
        ["Confirmar se id_symbol deve ser fixo ou variavel."],
        ["Confirmar se o app exige o nome Sheet2 ou apenas a posicao da aba."],
    ]
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column, value=value)


def _clear_sheet_values(sheet: Worksheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def _is_blank(value) -> bool:
    return value is None or value == ""
