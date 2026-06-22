"""Geracao e insercao do grafico de balanceamento AV/NAV/D."""

from __future__ import annotations

import logging
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from app.schemas.analysis import OperationalAnalysis


logger = logging.getLogger(__name__)

DEFAULT_BALANCE_CHART_SHEET = "Gráfico Balanceamento x Volume"
DEFAULT_BALANCE_CHART_ANCHOR = "B4"
DEFAULT_BALANCE_CHART_WIDTH = 760
DEFAULT_BALANCE_CHART_HEIGHT = 430


def _format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.1f}"


def _chart_anchor() -> str:
    return os.getenv("BALANCE_CHART_ANCHOR", DEFAULT_BALANCE_CHART_ANCHOR)


def generate_balance_chart_image(analysis: OperationalAnalysis, output_png: str) -> str:
    """Gera PNG com uma coluna empilhada AV/NAV/D em segundos."""

    output = Path(output_png)
    output.parent.mkdir(parents=True, exist_ok=True)
    matplotlib_config_dir = output.parent / "matplotlib_cache"
    matplotlib_config_dir.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(matplotlib_config_dir))

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    summary = analysis.resumo_tempos
    values = [summary.av_s, summary.nav_s, summary.d_s]
    percents = [summary.av_percent, summary.nav_percent, summary.d_percent]
    labels = ["AV", "NAV", "D"]
    colors = ["#2E7D32", "#F9A825", "#C62828"]

    fig, ax = plt.subplots(figsize=(7.2, 4.3), dpi=150)
    bottom = 0.0
    for label, value, percent, color in zip(labels, values, percents, colors, strict=True):
        ax.bar(
            ["Ciclo observado"],
            [value],
            bottom=bottom,
            color=color,
            width=0.42,
            label=f"{label} - {_format_number(value)}s ({percent:.1f}%)",
        )
        if value > 0:
            ax.text(
                0,
                bottom + value / 2,
                f"{label}\n{_format_number(value)}s\n{percent:.1f}%",
                ha="center",
                va="center",
                color="white" if label in {"AV", "D"} else "#222222",
                fontsize=9,
                fontweight="bold",
            )
        bottom += value

    takt_time_s = analysis.metadata.takt_time_s
    if takt_time_s is not None:
        ax.axhline(
            takt_time_s,
            color="#1565C0",
            linestyle="--",
            linewidth=1.4,
            label=f"Takt - {_format_number(takt_time_s)}s",
        )

    upper_limit = max(summary.total_s, takt_time_s or 0, 1) * 1.18
    ax.set_ylim(0, upper_limit)
    ax.set_title("Balanceamento AV / NAV / D", fontsize=14, fontweight="bold")
    ax.set_ylabel("Tempo (s)")
    ax.grid(axis="y", linestyle=":", linewidth=0.8, alpha=0.5)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    takt_text = "sem takt" if takt_time_s is None else f"Takt: {_format_number(takt_time_s)}s"
    ax.text(
        0.98,
        0.97,
        f"Total do ciclo: {_format_number(summary.total_s)}s\n{takt_text}",
        transform=ax.transAxes,
        ha="right",
        va="top",
        fontsize=9,
        bbox={"boxstyle": "round,pad=0.35", "facecolor": "#F7F7F7", "edgecolor": "#CCCCCC"},
    )
    ax.legend(loc="upper left", bbox_to_anchor=(1.01, 1.0), frameon=False)
    fig.tight_layout()
    fig.savefig(output, format="png", bbox_inches="tight")
    plt.close(fig)

    return str(output)


def insert_balance_chart(
    workbook_path: str,
    image_path: str,
    sheet_name: str = DEFAULT_BALANCE_CHART_SHEET,
) -> None:
    """Insere o PNG do balanceamento em uma aba existente, sem apagar objetos."""

    workbook_file = Path(workbook_path)
    image_file = Path(image_path)

    if not workbook_file.exists():
        raise FileNotFoundError(f"Workbook nao encontrado: {workbook_file}")
    if not image_file.exists():
        raise FileNotFoundError(f"Imagem do grafico nao encontrada: {image_file}")

    workbook = load_workbook(workbook_file, data_only=False, keep_links=True)
    if sheet_name not in workbook.sheetnames:
        logger.warning("Aba %s nao encontrada; grafico de balanceamento nao inserido.", sheet_name)
        workbook.close()
        return

    sheet = workbook[sheet_name]
    image = ExcelImage(str(image_file))
    image.width = DEFAULT_BALANCE_CHART_WIDTH
    image.height = DEFAULT_BALANCE_CHART_HEIGHT
    sheet.add_image(image, _chart_anchor())
    workbook.save(workbook_file)
    workbook.close()
