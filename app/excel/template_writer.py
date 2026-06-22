"""Preenchimento do template Excel SPS a partir de uma analise validada."""

from __future__ import annotations

import logging
import os
from collections import Counter
from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.analysis.activity_text import get_microstep_activity_text
from app.analysis.export_preparer import prepare_analysis_for_export
from app.analysis.quality_alerts import ALERT_HEADERS, ALERTS_SHEET_NAME, build_quality_alert_rows
from app.analysis.schema_compat import (
    build_validation_warnings_for_analysis,
    normalize_analysis_payload_for_current_schema,
)
from app.analysis.sps_validator import assert_analysis_can_generate_excel
from app.excel.balance_chart import generate_balance_chart_image, insert_balance_chart
from app.excel.conversion_sheet_writer import ensure_conversion_sheets, validate_conversion_sheet_contract
from app.excel.standard_writer import write_standard_consolidado_sheet, write_standard_sheets
from app.schemas.analysis import OperationalAnalysis
from app.spaghetti.map_generator import (
    generate_spaghetti_map_image,
    insert_spaghetti_map,
    insert_layout_image,
)
from app.utils.time_utils import seconds_to_mmss


logger = logging.getLogger(__name__)


ANALYSIS_SHEET = "ANÁLISE"
IMPROVEMENTS_SHEET = "MELHORIAS"
VALIDATION_WARNINGS_SHEET = "Avisos_Validacao"
ANALYSIS_FIRST_DATA_ROW = 7
ANALYSIS_LAST_COLUMN = 9
IMPROVEMENTS_FIRST_DATA_ROW = 7
IMPROVEMENTS_LAST_COLUMN = 8
RECOMMENDATIONS_HEADER_ROW = 14
RECOMMENDATIONS_FIRST_ROW = 15
FORCE_EXCEL_EXPORT_WITH_ALERTS = True


def _format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.1f}"


def _format_seconds(value: float | None) -> str:
    if value is None:
        return ""
    return f"{_format_number(value)}s ({seconds_to_mmss(value)})"


def _format_delta_seconds(value: float | None) -> str:
    if value is None:
        return ""

    sign = "+" if value >= 0 else "-"
    absolute = abs(value)
    return f"{sign}{_format_number(absolute)}s"


def _format_seconds_percent(seconds: float, percent: float) -> str:
    return f"{_format_number(seconds)}s ({percent:.1f}%)"


def _env_flag(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().casefold() in {"1", "true", "yes", "sim", "on"}


def _collect_quality_alerts_for_export(
    analysis: OperationalAnalysis,
    quality_alerts: list[str] | None = None,
) -> list[str]:
    """Collect SPS quality findings as export alerts, never as blockers."""

    alerts: list[str] = []
    for alert in quality_alerts or []:
        if alert and alert not in alerts:
            alerts.append(alert)
    try:
        from app.analysis.quality_gate import validate_analysis_quality

        gate = validate_analysis_quality(analysis, None, None)
        for alert in gate.alerts:
            if alert and alert not in alerts:
                alerts.append(alert)
    except Exception as exc:
        message = f"Quality gate SPS nao pode ser coletado para aba de alertas: {exc}"
        if message not in alerts:
            alerts.append(message)
    return alerts


def _writable_cell(sheet: Worksheet, row: int, column: int):
    cell = sheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return None
    return cell


def _set_cell(sheet: Worksheet, coordinate: str, value) -> None:
    cell = sheet[coordinate]
    if isinstance(cell, MergedCell):
        raise ValueError(f"Celula {sheet.title}!{coordinate} faz parte de mescla e nao pode ser escrita diretamente")
    cell.value = value


def _clear_region(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> None:
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            cell = _writable_cell(sheet, row, column)
            if cell is not None:
                cell.value = None


def _find_header_column(sheet: Worksheet, header_row: int, expected_terms: tuple[str, ...]) -> int | None:
    for cell in sheet[header_row]:
        value = str(cell.value or "").casefold()
        if value and all(term.casefold() in value for term in expected_terms):
            return cell.column
    return None


def _find_or_create_header_column(sheet: Worksheet, header_row: int, header_name: str) -> int:
    existing = _find_header_column(sheet, header_row, (header_name,))
    if existing is not None:
        return existing

    last_column = max((cell.column for cell in sheet[header_row] if cell.value is not None), default=0)
    target_column = last_column + 1
    cell = _writable_cell(sheet, header_row, target_column)
    if cell is None:
        raise ValueError(f"Nao foi possivel criar cabecalho '{header_name}' na aba {sheet.title}")
    cell.value = header_name
    return target_column


def _clear_sheet(sheet: Worksheet) -> None:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _coalesce_cycle_seconds(analysis: OperationalAnalysis) -> float:
    if analysis.metadata.ciclo_observado_s is not None:
        return analysis.metadata.ciclo_observado_s
    return analysis.resumo_tempos.total_s


def _build_improvement_reading(analysis: OperationalAnalysis) -> str:
    summary = analysis.resumo_tempos
    if summary.d_s <= 0 or not analysis.melhorias:
        return "Sem desperdícios D registrados nesta análise; manter verificação no gemba."

    waste_types = Counter(item.tipo_desperdicio for item in analysis.melhorias)
    main_types = ", ".join(item for item, _ in waste_types.most_common(3))
    high_priority = sum(1 for item in analysis.melhorias if item.prioridade == "Alta")

    priority_text = (
        f"{high_priority} sugestão(ões) de prioridade Alta"
        if high_priority
        else "sem sugestão de prioridade Alta"
    )
    return (
        f"Desperdícios D somam {_format_seconds_percent(summary.d_s, summary.d_percent)} "
        f"do ciclo, com concentração em: {main_types}. Há {priority_text}; validar ações no gemba."
    )


def _write_analysis_sheet(sheet: Worksheet, analysis: OperationalAnalysis) -> None:
    metadata = analysis.metadata
    summary = analysis.resumo_tempos
    cycle_seconds = _coalesce_cycle_seconds(analysis)
    max_data_row = max(sheet.max_row, ANALYSIS_FIRST_DATA_ROW + len(analysis.microetapas) - 1)
    accumulated_column = _find_header_column(sheet, 6, ("acumul",))
    if accumulated_column is not None and accumulated_column <= ANALYSIS_LAST_COLUMN:
        accumulated_column = None

    _set_cell(sheet, "A1", f"ANÁLISE DETALHADA DO VÍDEO – {metadata.posto}")
    _set_cell(sheet, "B3", metadata.departamento)
    _set_cell(sheet, "E3", metadata.responsavel)
    _set_cell(sheet, "H3", metadata.data_analise)
    _set_cell(sheet, "B4", metadata.posto)
    _set_cell(sheet, "E4", metadata.processo)
    _set_cell(sheet, "H4", _format_seconds(metadata.takt_time_s))
    _set_cell(sheet, "B5", _format_seconds(cycle_seconds))
    _set_cell(sheet, "E5", _format_seconds_percent(summary.av_s, summary.av_percent))
    _set_cell(sheet, "G5", _format_seconds_percent(summary.nav_s, summary.nav_percent))
    _set_cell(sheet, "I5", _format_seconds_percent(summary.d_s, summary.d_percent))

    _clear_region(
        sheet,
        ANALYSIS_FIRST_DATA_ROW,
        max_data_row,
        1,
        ANALYSIS_LAST_COLUMN,
    )
    if accumulated_column is not None:
        _clear_region(
            sheet,
            ANALYSIS_FIRST_DATA_ROW,
            max_data_row,
            accumulated_column,
            accumulated_column,
        )

    conf_column = _find_or_create_header_column(sheet, 6, "Confiança")
    status_column = _find_or_create_header_column(sheet, 6, "Status de validação")
    aviso_column = _find_or_create_header_column(sheet, 6, "Aviso")
    gemba_column = _find_or_create_header_column(sheet, 6, "Validar no gemba?")
    motivo_column = _find_or_create_header_column(sheet, 6, "Motivo / observação")
    extra_columns = [conf_column, status_column, aviso_column, gemba_column, motivo_column]

    for column in extra_columns:
        _clear_region(
            sheet,
            ANALYSIS_FIRST_DATA_ROW,
            max_data_row,
            column,
            column,
        )

    for row_index, step in enumerate(analysis.microetapas, start=ANALYSIS_FIRST_DATA_ROW):
        row_values = (
            step.numero,
            get_microstep_activity_text(step),
            step.inicio_formatado,
            step.fim_formatado,
            step.duracao_s,
            step.duracao_formatada,
            step.classificacao,
            step.justificativa_tecnica,
            step.evidencia_observavel or step.ferramenta_observacao or "",
        )
        for column_index, value in enumerate(row_values, start=1):
            cell = _writable_cell(sheet, row_index, column_index)
            if cell is not None:
                cell.value = value
        if accumulated_column is not None:
            cell = _writable_cell(sheet, row_index, accumulated_column)
            if cell is not None:
                cell.value = step.tempo_acumulado_formatado or step.tempo_acumulado_s or ""

        confidence_value = step.confianca
        status_value = "Validar no gemba" if confidence_value < 0.7 else ("Revisar" if confidence_value < 0.9 else "Ok")
        aviso_value = "Baixa confiança" if confidence_value < 0.7 else ""
        gemba_value = "Sim" if confidence_value < 0.7 else "Nao"
        motivo_value = step.baixa_confianca_motivo or ""

        for column_index, value in (
            (conf_column, confidence_value),
            (status_column, status_value),
            (aviso_column, aviso_value),
            (gemba_column, gemba_value),
            (motivo_column, motivo_value),
        ):
            cell = _writable_cell(sheet, row_index, column_index)
            if cell is not None:
                cell.value = value


def _write_improvements_sheet(sheet: Worksheet, analysis: OperationalAnalysis) -> None:
    metadata = analysis.metadata
    summary = analysis.resumo_tempos
    cycle_seconds = _coalesce_cycle_seconds(analysis)

    improvements_end_row = max(
        RECOMMENDATIONS_HEADER_ROW - 1,
        IMPROVEMENTS_FIRST_DATA_ROW + len(analysis.melhorias) - 1,
    )
    recommendations_end_row = max(
        sheet.max_row,
        RECOMMENDATIONS_FIRST_ROW + len(analysis.recomendacoes_gerais) - 1,
    )
    cause_column = _find_header_column(sheet, 6, ("causa",))
    if cause_column is not None and cause_column <= IMPROVEMENTS_LAST_COLUMN:
        cause_column = None
    validation_column = _find_header_column(sheet, 6, ("gemba",))
    if validation_column is None:
        validation_column = _find_header_column(sheet, 6, ("valid",))
    if validation_column is not None and validation_column <= IMPROVEMENTS_LAST_COLUMN:
        validation_column = None

    _set_cell(sheet, "B3", _format_seconds(cycle_seconds))
    _set_cell(sheet, "E3", _format_seconds(metadata.takt_time_s))
    _set_cell(sheet, "H3", _format_delta_seconds(summary.folga_vs_takt_s))
    _set_cell(sheet, "B4", _format_seconds_percent(summary.d_s, summary.d_percent))
    _set_cell(sheet, "E4", _build_improvement_reading(analysis))

    _clear_region(
        sheet,
        IMPROVEMENTS_FIRST_DATA_ROW,
        improvements_end_row,
        1,
        IMPROVEMENTS_LAST_COLUMN,
    )
    for optional_column in (cause_column, validation_column):
        if optional_column is not None:
            _clear_region(
                sheet,
                IMPROVEMENTS_FIRST_DATA_ROW,
                improvements_end_row,
                optional_column,
                optional_column,
            )
    _clear_region(
        sheet,
        RECOMMENDATIONS_HEADER_ROW,
        recommendations_end_row,
        1,
        IMPROVEMENTS_LAST_COLUMN,
    )

    for row_index, suggestion in enumerate(analysis.melhorias, start=IMPROVEMENTS_FIRST_DATA_ROW):
        row_values = (
            suggestion.microetapa_numero,
            suggestion.inicio_formatado,
            suggestion.fim_formatado,
            suggestion.duracao_s,
            suggestion.descricao_desperdicio,
            suggestion.tipo_desperdicio,
            suggestion.sugestao_pratica,
            suggestion.prioridade,
        )
        for column_index, value in enumerate(row_values, start=1):
            cell = _writable_cell(sheet, row_index, column_index)
            if cell is not None:
                cell.value = "" if value is None else value
        if cause_column is not None:
            cell = _writable_cell(sheet, row_index, cause_column)
            if cell is not None:
                cell.value = suggestion.causa_observavel
        if validation_column is not None:
            cell = _writable_cell(sheet, row_index, validation_column)
            if cell is not None:
                cell.value = "Sim" if suggestion.requer_validacao_gemba else "Nao"

    _set_cell(sheet, "A14", "RECOMENDAÇÕES GERAIS")
    for row_index, recommendation in enumerate(
        analysis.recomendacoes_gerais,
        start=RECOMMENDATIONS_FIRST_ROW,
    ):
        cell = _writable_cell(sheet, row_index, 1)
        if cell is not None:
            cell.value = f"• {recommendation}"


def write_analysis_to_template(
    analysis: OperationalAnalysis,
    template_path: str,
    output_path: str,
    quality_alerts: list[str] | None = None,
    force_export_with_alerts: bool = True,
    fill_standard: bool = False,
    standard_export_mode: str = "standard_legacy",
    export_conversion_sheet: bool = True,
    insert_charts: bool = False,
    insert_spaghetti: bool = False,
    layout_path: str | None = None,
    layout_image_path: str | None = None,
) -> str:
    """Copia o template e preenche as abas ANÁLISE e MELHORIAS na cópia."""

    analysis = _rehydrate_analysis_for_current_schema(analysis)
    analysis = prepare_analysis_for_export(analysis)
    quality_alerts = _collect_quality_alerts_for_export(analysis, quality_alerts)
    if quality_alerts:
        alerts = list(analysis.alertas_validacao)
        for alert in quality_alerts:
            if alert not in alerts:
                alerts.append(alert)
        analysis = OperationalAnalysis.model_validate(analysis.model_dump() | {"alertas_validacao": alerts})
    warnings = build_validation_warnings_for_analysis(analysis)
    effective_force_export = bool(force_export_with_alerts) or _env_flag(
        "FORCE_EXCEL_EXPORT_WITH_ALERTS",
        FORCE_EXCEL_EXPORT_WITH_ALERTS,
    )
    quality_gate_blocks_excel = _env_flag("QUALITY_GATE_BLOCKS_EXCEL", False)
    assert_analysis_can_generate_excel(
        analysis,
        block_on_quality=quality_gate_blocks_excel and not effective_force_export,
    )

    template = Path(template_path)
    output = Path(output_path)

    if not template.exists():
        raise FileNotFoundError(f"Template nao encontrado: {template}")

    output.parent.mkdir(parents=True, exist_ok=True)
    if template.resolve() == output.resolve():
        raise ValueError("output_path nao pode ser igual ao template_path")

    copy2(template, output)

    workbook = load_workbook(output, data_only=False, keep_links=True)
    missing_sheets = [
        sheet_name
        for sheet_name in (ANALYSIS_SHEET, IMPROVEMENTS_SHEET)
        if sheet_name not in workbook.sheetnames
    ]
    if missing_sheets:
        workbook.close()
        raise ValueError(f"Template sem aba(s) obrigatoria(s): {', '.join(missing_sheets)}")

    _write_analysis_sheet(workbook[ANALYSIS_SHEET], analysis)
    _write_improvements_sheet(workbook[IMPROVEMENTS_SHEET], analysis)
    if fill_standard:
        if standard_export_mode == "standard_consolidado":
            write_standard_consolidado_sheet(workbook, analysis)
        else:
            write_standard_sheets(workbook, analysis)
    if export_conversion_sheet:
        ensure_conversion_sheets(workbook, analysis)
        conversion_errors = validate_conversion_sheet_contract(workbook, analysis)
        if conversion_errors:
            workbook.close()
            raise ValueError("Contrato Sheet2 invalido: " + " | ".join(conversion_errors))

    _write_validation_warnings_sheet(workbook, warnings)
    _write_sps_validation_alerts_sheet(workbook, analysis, quality_alerts)
    workbook.save(output)
    workbook.close()

    if insert_charts:
        chart_path = output.with_name(f"{output.stem}_balance_chart.png")
        generate_balance_chart_image(analysis, str(chart_path))
        insert_balance_chart(str(output), str(chart_path))

    if insert_spaghetti:
        if layout_path is None:
            if layout_image_path is not None:
                insert_layout_image(str(output), layout_image_path)
                message = (
                    "Mapa de espaguete solicitado. Layout JSON nao fornecido; "
                    "a imagem do posto foi inserida como referencia visual."
                )
                logger.warning(message)
                if message not in analysis.alertas_validacao:
                    analysis.alertas_validacao.append(message)
            else:
                message = "Mapa de espaguete solicitado, mas nenhum layout JSON ou imagem foi informado."
                logger.warning(message)
                if message not in analysis.alertas_validacao:
                    analysis.alertas_validacao.append(message)
        else:
            if layout_image_path is not None:
                insert_layout_image(str(output), layout_image_path)
                message = (
                    "Layout JSON fornecido e imagem do posto anexada como referencia visual."
                )
                if message not in analysis.alertas_validacao:
                    analysis.alertas_validacao.append(message)
            spaghetti_path = output.with_name(f"{output.stem}_spaghetti_map.png")
            generate_spaghetti_map_image(analysis, layout_path, str(spaghetti_path))
            insert_spaghetti_map(str(output), str(spaghetti_path))

    _refresh_sps_validation_alerts_sheet(output, analysis, quality_alerts)
    return str(output)


def export_analysis_with_warnings(
    analysis: OperationalAnalysis,
    template_path: str,
    output_path: str,
    quality_alerts: list[str] | None = None,
    force_export: bool = True,
    **kwargs,
) -> str:
    """Generate Excel while preserving SPS quality alerts for manual review."""

    return write_analysis_to_template(
        analysis=analysis,
        template_path=template_path,
        output_path=output_path,
        quality_alerts=quality_alerts,
        force_export_with_alerts=force_export,
        **kwargs,
    )


def _write_sps_validation_alerts_sheet(
    workbook: object,
    analysis: OperationalAnalysis,
    quality_alerts: list[str] | None = None,
) -> None:
    if ALERTS_SHEET_NAME in workbook.sheetnames:
        del workbook[ALERTS_SHEET_NAME]
    sheet = workbook.create_sheet(title=ALERTS_SHEET_NAME)

    title_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_font = Font(bold=True, color="1F1F1F")
    body_alignment = Alignment(wrap_text=True, vertical="top")

    sheet["A1"] = "ALERTAS DE VALIDACAO SPS"
    sheet["A1"].font = title_font
    sheet["A1"].fill = title_fill
    sheet["A2"] = (
        "Esta planilha foi gerada mesmo com alertas. Revisar manualmente antes de utilizar "
        "como padrao oficial."
    )
    sheet["A2"].alignment = body_alignment

    header_row = 4
    for column, header in enumerate(ALERT_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = body_alignment

    rows = build_quality_alert_rows(analysis, quality_alerts)
    if not rows:
        rows = [
            {
                "numero": 1,
                "severidade": "Informativa",
                "tipo": "Sem alerta",
                "microetapa": "",
                "descricao": "Nenhum alerta critico registrado.",
                "acao_recomendada": "Manter revisao manual habitual antes de oficializar o padrao.",
                "requer_validacao_gemba": "Nao",
                "status": "Concluido",
            }
        ]

    for row_index, alert in enumerate(rows, start=header_row + 1):
        values = (
            alert.get("numero"),
            alert.get("severidade"),
            alert.get("tipo"),
            alert.get("microetapa"),
            alert.get("descricao"),
            alert.get("acao_recomendada"),
            alert.get("requer_validacao_gemba"),
            alert.get("status"),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = body_alignment

    widths = {
        "A": 8,
        "B": 16,
        "C": 28,
        "D": 14,
        "E": 70,
        "F": 70,
        "G": 24,
        "H": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A5"


def _refresh_sps_validation_alerts_sheet(
    output: Path,
    analysis: OperationalAnalysis,
    quality_alerts: list[str] | None = None,
) -> None:
    workbook = load_workbook(output, data_only=False, keep_links=True)
    try:
        _write_sps_validation_alerts_sheet(workbook, analysis, quality_alerts)
        workbook.save(output)
    finally:
        workbook.close()


def _write_validation_warnings_sheet(workbook: object, warnings: list[dict[str, Any]]) -> None:
    if VALIDATION_WARNINGS_SHEET in workbook.sheetnames:
        sheet = workbook[VALIDATION_WARNINGS_SHEET]
        _clear_sheet(sheet)
    else:
        sheet = workbook.create_sheet(title=VALIDATION_WARNINGS_SHEET)

    headers = [
        "Número",
        "Etapa",
        "Campo",
        "Severidade",
        "Mensagem",
        "Ação recomendada",
        "Confiança",
        "Validado no gemba?",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = _writable_cell(sheet, 1, col_index)
        if cell is not None:
            cell.value = header

    for row_index, warning in enumerate(warnings, start=2):
        values = (
            warning.get("numero"),
            warning.get("etapa"),
            warning.get("campo"),
            warning.get("severidade"),
            warning.get("mensagem"),
            warning.get("acao_recomendada"),
            warning.get("confianca"),
            warning.get("validacao_gemba"),
        )
        for col_index, value in enumerate(values, start=1):
            cell = _writable_cell(sheet, row_index, col_index)
            if cell is not None:
                cell.value = value


def _rehydrate_analysis_for_current_schema(analysis: OperationalAnalysis) -> OperationalAnalysis:
    """Accept analyses created before optional schema fields were added."""

    return OperationalAnalysis.model_validate(
        normalize_analysis_payload_for_current_schema(analysis)
    )
