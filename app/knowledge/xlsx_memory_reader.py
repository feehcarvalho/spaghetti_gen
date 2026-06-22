"""Structured reader for XLSX files used as SPS memories."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class XlsxMemoryContent:
    workbook_name: str
    sheet_names: list[str]
    process_descriptions: list[str] = field(default_factory=list)
    terminology_terms: list[str] = field(default_factory=list)
    mapping_rules: list[str] = field(default_factory=list)
    operational_examples: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    source_cells: list[str] = field(default_factory=list)

    def to_context_text(self) -> str:
        blocks = [
            f"MEMORIA XLSX: {self.workbook_name}",
            f"Abas: {', '.join(self.sheet_names)}",
        ]
        for title, values in (
            ("Descricoes de processo", self.process_descriptions),
            ("Termos/nomenclatura", self.terminology_terms),
            ("Regras de mapeamento/conversao", self.mapping_rules),
            ("Exemplos operacionais", self.operational_examples),
            ("Avisos", self.warnings),
            ("Celulas fonte", self.source_cells),
        ):
            if values:
                blocks.append(title + ":\n" + "\n".join(f"- {item}" for item in values[:40]))
        return "\n".join(blocks)


def read_xlsx_memory(file_path: str) -> XlsxMemoryContent:
    path = Path(file_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        process: list[str] = []
        terminology: list[str] = []
        mappings: list[str] = []
        examples: list[str] = []
        warnings: list[str] = []
        cells: list[str] = []

        for sheet in workbook.worksheets:
            lowered_title = sheet.title.casefold()
            for row in sheet.iter_rows(max_row=80):
                values = [str(cell.value).strip() for cell in row if cell.value is not None and str(cell.value).strip()]
                if not values:
                    continue
                line = " | ".join(values)
                first_coordinate = next(cell.coordinate for cell in row if cell.value is not None and str(cell.value).strip())
                source = f"{sheet.title}!{first_coordinate}: {line[:240]}"
                cells.append(source)
                lowered = f"{lowered_title} {line.casefold()}"
                if any(key in lowered for key in ("processo", "metodo", "operação", "operacao", "descrição", "descricao")):
                    process.append(line)
                if any(key in lowered for key in ("termo", "nomenclatura", "ferramenta", "dispositivo", "peça", "peca")):
                    terminology.append(line)
                if any(key in lowered for key in ("mapeamento", "convers", "sheet2", "id_avnavd", "activity", "timeofeelement")):
                    mappings.append(line)
                if any(key in lowered for key in ("exemplo", "microetapa", "instrução", "instrucao", "atividade")):
                    examples.append(line)
                if any(key in lowered for key in ("atenção", "atencao", "alerta", "risco", "segurança", "seguranca")):
                    warnings.append(line)

        return XlsxMemoryContent(
            workbook_name=path.name,
            sheet_names=sheet_names,
            process_descriptions=_unique(process),
            terminology_terms=_unique(terminology),
            mapping_rules=_unique(mappings),
            operational_examples=_unique(examples),
            warnings=_unique(warnings),
            source_cells=_unique(cells),
        )
    finally:
        workbook.close()


def _unique(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        if value not in result:
            result.append(value)
    return result
