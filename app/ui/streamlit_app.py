"""Interface Streamlit para execucao assistida da analise SPS."""

from __future__ import annotations

import json
import os
import re
import sys
import time
import zipfile
import base64
import hashlib
import importlib
import inspect
import mimetypes
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

# Force load .env variables at app startup
from dotenv import load_dotenv
REPO_ROOT = Path(__file__).resolve().parents[2]
load_dotenv(REPO_ROOT / ".env", override=True)
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    import streamlit as st
except ImportError:  # pragma: no cover - permite importar o modulo sem streamlit.
    st = None

from app.config import get_settings, get_default_template_path
# Always get fresh settings with updated environment variables
settings = get_settings()
from app.analysis.correction_flow import (
    append_correction_history,
    prepare_correction_rerun,
    save_correction_context_note,
    save_feedback_memory,
)
from app.analysis.sps_validator import assert_analysis_can_generate_excel
from app.excel.template_writer import write_analysis_to_template
from app.analysis.export_preparer import prepare_analysis_for_export
from app.analysis.quality_alerts import build_quality_alert_rows
from app.analysis.schema_compat import normalize_analysis_payload_for_current_schema
import app.main as app_main
from app.main import run_analysis_only
from app.schemas.analysis import AnalysisMetadata, OperationalAnalysis
from app.security.audit_logger import log_login_event
from app.ui.auth import render_login_page
from app.ui.review_editor import analysis_to_dataframe, dataframe_to_analysis
from app.video.frame_extractor import ExtractedFrame, extract_frames


TEMPLATE_FILENAME = "PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx"
DEFAULT_TEMPLATE_PATH = REPO_ROOT / "data" / "templates" / TEMPLATE_FILENAME
ROOT_TEMPLATE_FALLBACK = REPO_ROOT / TEMPLATE_FILENAME
VIDEOS_DIR = REPO_ROOT / "data" / "videos"
VIDEOS_UPLOAD_DIR = VIDEOS_DIR / "uploads"
FRAME_PREVIEW_DIR = REPO_ROOT / "data" / "frames" / "previews"
TEMPLATES_TEMP_DIR = REPO_ROOT / "data" / "templates_temp"
LAYOUTS_DIR = REPO_ROOT / "data" / "layouts"
LAYOUTS_TEMP_DIR = REPO_ROOT / "data" / "layouts_temp"
LAYOUTS_UPLOAD_DIR = LAYOUTS_DIR / "uploads"
OUTPUTS_DIR = REPO_ROOT / "data" / "outputs"
KNOWLEDGE_ROOT = REPO_ROOT / "data" / "knowledge_raw"
KNOWLEDGE_UPLOAD_DIR = KNOWLEDGE_ROOT / "uploads"
SESSION_KNOWLEDGE_DIR = OUTPUTS_DIR / "session_knowledge"
BACKGROUND_UPLOAD_DIR = REPO_ROOT / "data" / "backgrounds" / "uploads"
DEFAULT_BACKGROUND_IMAGE = Path.home() / "Downloads" / "19357-001.jpg"
VIDEO_EXTENSIONS = {".mp4", ".mov", ".avi", ".mkv"}
TEMPLATE_EXTENSIONS = {".xlsx"}
LAYOUT_EXTENSIONS = {".json"}
LAYOUT_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
KNOWLEDGE_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".txt", ".md", ".csv", ".png", ".jpg", ".jpeg"}
BACKGROUND_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
GENERATION_LOCK_MAX_AGE_S = 15 * 60
GENERATION_LOCK_DIR = Path(tempfile.gettempdir()) / "ia_sps_scania_locks"
DEFAULT_PROVIDER = "openai"
DEFAULT_EXTRACTION_FPS = 1.0
DEFAULT_ANALYSIS_MAX_FRAMES = 0
PREVIEW_MAX_FRAMES = 5


class GenerationAlreadyRunningError(RuntimeError):
    """Raised when another Excel generation holds the output lock."""


PRODUCTION_QUALITY_MODE = "Máxima qualidade / produção"
QUALITY_PRESETS = {
    "Máxima qualidade / produção": {
        "window_seconds": 12,
        "max_frames_per_window": 10,
        "detail_window": "auto",
        "reprocess_low_confidence": True,
    },
    "Equilibrada": {
        "window_seconds": 14,
        "max_frames_per_window": 8,
        "detail_window": "auto",
        "reprocess_low_confidence": True,
    },
    "Diagnóstico rápido": {
        "window_seconds": 20,
        "max_frames_per_window": 6,
        "detail_window": "auto",
        "reprocess_low_confidence": False,
    },
}
MOCK_DEMO_WARNING = "Modo demonstração/mock: esta análise é simulada e não representa vídeo real."
MOCK_VIDEO_ERROR_UI = "Modo mock não analisa vídeo real. Use provider openai para analisar o vídeo."
OPENAI_API_KEY_ERROR_UI = "OPENAI_API_KEY não configurada. Configure a chave para análise real de vídeo."
REQUIRED_TEMPLATE_SHEETS = {"ANÁLISE", "MELHORIAS"}


def _safe_filename(value: str) -> str:
    name = Path(value).name.strip() or "upload"
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", name)


def _safe_slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", value.strip()).strip("_")
    return slug.lower() or "analise"


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")


def _refresh_env() -> None:
    load_dotenv(REPO_ROOT / ".env", override=True)


def _has_openai_api_key() -> bool:
    _refresh_env()
    api_key = os.getenv("OPENAI_API_KEY")
    return bool(api_key and api_key.strip())


def _run_analysis_only_compat(**kwargs: Any) -> OperationalAnalysis:
    """Call run_analysis_only while tolerating stale Streamlit module cache."""

    func = run_analysis_only
    signature = inspect.signature(func)
    expected_optional = {"knowledge_paths", "resume_from_checkpoint"}
    if not expected_optional.issubset(signature.parameters):
        try:
            reloaded_main = importlib.reload(app_main)
            func = reloaded_main.run_analysis_only
            signature = inspect.signature(func)
        except Exception:
            pass

    accepted_kwargs = {key: value for key, value in kwargs.items() if key in signature.parameters}
    dropped = set(kwargs) - set(accepted_kwargs)
    if dropped & expected_optional and st is not None:
        st.warning(
            "O servidor Streamlit estava com parte do código antigo em memória. "
            "A análise continuará, mas reinicie o Streamlit para carregar todos os recursos novos."
        )
    return func(**accepted_kwargs)


def _session_value(key: str) -> Any:
    if st is None:
        return None
    try:
        return st.session_state.get(key)
    except Exception:
        return None


def _authenticated_responsavel_label() -> str:
    login = str(_session_value("auth_user_login") or "").strip()
    name = str(_session_value("auth_user_name") or "").strip()
    if name and login:
        return f"{name} ({login})"
    return login


def _render_authenticated_user_bar() -> None:
    login = str(_session_value("auth_user_login") or "").strip()
    name = str(_session_value("auth_user_name") or "").strip()
    area = str(_session_value("auth_user_area") or "").strip()
    if not login:
        return

    user_label = f"{name} ({login})" if name else login
    st.sidebar.caption(f"Usuário logado: {user_label}")
    if area:
        st.sidebar.caption(f"Area: {area}")

    if st.sidebar.button("Sair", key="auth_logout_button"):
        log_login_event(
            login,
            "logout",
            "Logout realizado pelo usuário.",
            {"nome": name, "area": area, "perfil": _session_value("auth_user_profile")},
        )
        for key in list(st.session_state.keys()):
            if str(key).startswith("auth_"):
                st.session_state.pop(key, None)
        st.rerun()


def _render_memory_sidebar() -> dict[str, Any]:
    """Renderiza a memoria auxiliar sem entrar no fluxo principal numerado."""

    payload = {
        "uploads": [],
        "note": "",
        "persist_for_future": False,
        "enabled": False,
    }
    with st.sidebar.expander("Memória da IA / adicionar conhecimento", expanded=False):
        st.caption(
            "Use esta área para adicionar documentos, fotos ou nomenclaturas que ajudem a IA a interpretar "
            "o processo. Esses arquivos complementam a memória interna usada na análise."
        )
        uploads = st.file_uploader(
            "Arquivos de memória/conhecimento",
            type=["pdf", "docx", "xlsx", "txt", "md", "csv", "png", "jpg", "jpeg"],
            accept_multiple_files=True,
            help="Aprendizado significa salvar conhecimento local/RAG; nao treina nem altera pesos do modelo OpenAI.",
            key="memory_knowledge_uploads_sidebar",
        ) or []
        note = st.text_area(
            "Observação de memória / nomenclatura",
            value=st.session_state.get("memory_note_current", ""),
            placeholder="Ex.: Neste posto, o dispositivo usado para transporte do pneu é chamado de VR de pneu.",
            height=90,
            key="memory_observation_text",
        )
        persist_for_future = st.checkbox(
            "Salvar como memória interna para análises futuras",
            value=bool(st.session_state.get("memory_persist_for_future", False)),
            key="memory_persist_checkbox",
        )
        if st.button("Adicionar à memória desta análise", key="memory_add_button"):
            st.session_state["memory_session_enabled"] = True
            st.session_state["memory_note_current"] = note
            st.session_state["memory_persist_for_future"] = persist_for_future
            st.success("Memória adicionada para esta análise.")

        if st.session_state.get("memory_session_enabled"):
            st.caption("Memória auxiliar ativa para a próxima análise.")
        if uploads:
            st.caption("Arquivos anexados: " + ", ".join(file.name for file in uploads))

    payload["uploads"] = uploads
    payload["note"] = note
    payload["persist_for_future"] = persist_for_future
    payload["enabled"] = bool(st.session_state.get("memory_session_enabled")) or bool(uploads) or bool(note.strip())
    return payload


def _image_data_url(path: Path | None) -> str | None:
    if path is None or not path.exists() or not path.is_file():
        return None
    mime_type = mimetypes.guess_type(path.name)[0] or "image/jpeg"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def _apply_brand_styles() -> None:
    """Aplica estilos CSS com contraste alto e paleta clara."""
    st.markdown(
        """
        <style>
            /* Fundo principal */
            .stApp {
                background: linear-gradient(135deg, #061B33 0%, #082B49 100%);
                color: #F8FAFC;
            }
            
            /* Container principal */
            .main .block-container {
                max-width: 1320px;
                padding-top: 2.2rem;
                padding-bottom: 3rem;
            }
            
            /* Sidebar */
            [data-testid="stSidebar"] {
                background: #061B33;
                border-right: 1px solid #7DB7E8;
            }
            
            /* Títulos e textos principais */
            h1, h2, h3 {
                color: #FFFFFF !important;
                font-weight: 700;
            }
            
            /* Labels com contraste alto */
            .stCheckbox label, 
            .stFileUploader label, 
            .stTextInput label, 
            .stNumberInput label, 
            .stTextArea label, 
            .stSelectbox label,
            label {
                color: #F8FAFC !important;
                font-weight: 600 !important;
            }
            
            /* Textos secundários/captions */
            .stCaption, .stHelp, small {
                color: #D8E4F2 !important;
            }
            
            /* Markdown e texto */
            .stMarkdown, .stText {
                color: #F8FAFC !important;
            }
            
            /* Hero section */
            .sps-hero {
                padding: 1.15rem 0 1.35rem;
                border-bottom: 1px solid #7DB7E8;
                margin-bottom: 1.1rem;
            }
            
            .sps-hero h1 {
                font-size: clamp(2.1rem, 4vw, 4.15rem);
                line-height: 1.02;
                margin: 0;
                letter-spacing: 0;
                color: #FFFFFF;
            }
            
            .sps-hero p {
                max-width: 760px;
                margin: 0.7rem 0 0;
                color: #D8E4F2;
                font-size: 1.02rem;
            }
            
            .sps-provider {
                display: inline-flex;
                align-items: center;
                gap: 0.45rem;
                margin-top: 0.95rem;
                color: #7DB7E8;
                font-size: 0.86rem;
                text-transform: uppercase;
                letter-spacing: 0.08em;
            }
            
            /* Cards e containers */
            [data-testid="stForm"], 
            [data-testid="stVerticalBlockBorderWrapper"] {
                background: #0E3558;
                border: 1px solid #7DB7E8;
                border-radius: 8px;
                padding: 1rem;
                box-shadow: 0 18px 45px rgba(0, 0, 0, 0.22);
            }
            
            /* Métricas */
            [data-testid="stMetric"] {
                background: #0E3558;
                border: 1px solid #7DB7E8;
                border-radius: 8px;
                padding: 0.75rem;
            }
            
            [data-testid="stMetric"] label {
                color: #D8E4F2 !important;
            }
            
            /* Botões */
            .stButton > button, 
            .stDownloadButton > button {
                border-radius: 6px;
                border: 1px solid #2F80ED;
                background: #2F80ED;
                color: #FFFFFF;
                font-weight: 600;
            }
            
            .stButton > button:hover, 
            .stDownloadButton > button:hover {
                border-color: #1C64C7;
                background: #1C64C7;
                color: #FFFFFF;
            }
            
            /* Inputs de texto */
            .stTextInput input, 
            .stNumberInput input, 
            .stDateInput input, 
            textarea,
            input[type="text"],
            input[type="number"],
            input[type="date"] {
                background-color: #FFFFFF !important;
                color: #0F172A !important;
                border-radius: 6px !important;
                border: 1px solid #7DB7E8 !important;
            }
            
            /* File uploader */
            [data-testid="stFileUploader"] section {
                background: #0E3558;
                border: 2px dashed #7DB7E8;
                border-radius: 8px;
            }
            
            [data-testid="stFileUploader"] label {
                color: #F8FAFC !important;
                font-weight: 600 !important;
            }
            
            /* Alertas e mensagens */
            .stAlert {
                background-color: #0E3558;
                border: 1px solid #7DB7E8;
                border-radius: 8px;
            }
            
            .stError, [data-testid="stAlert"] {
                color: #F8FAFC !important;
            }
            
            /* Success alerts */
            .stSuccess {
                background-color: rgba(46, 125, 50, 0.2);
                color: #A5D6A7 !important;
            }
            
            /* Warning alerts */
            .stWarning {
                background-color: rgba(251, 140, 0, 0.2);
                color: #FFB74D !important;
            }
            
            /* Info alerts */
            .stInfo {
                background-color: rgba(13, 71, 161, 0.2);
                color: #64B5F6 !important;
            }
            
            /* Datatable */
            .stDataFrame {
                background-color: #0E3558;
                border: 1px solid #7DB7E8;
            }
            
            /* Subheadings */
            .stSubheader {
                color: #FFFFFF !important;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _validate_upload_extension(filename: str, allowed_extensions: set[str]) -> None:
    suffix = Path(filename).suffix.lower()
    normalized = {item.lower() for item in allowed_extensions}
    if suffix not in normalized:
        expected = ", ".join(sorted(normalized))
        raise ValueError(f"Extensao de upload invalida para {filename}. Permitido: {expected}")


def _render_memory_upload_control() -> list[Any]:
    """Renderiza o anexo de memória como controle discreto no topo."""
    _, memory_col = st.columns([5.5, 1.7])
    memory_uploads: list[Any] = []
    uploader_kwargs = {
        "label": "Arquivos de memória/conhecimento",
        "type": ["pdf", "docx", "xlsx", "txt", "md", "csv"],
        "accept_multiple_files": True,
        "help": "Use apenas quando precisar ensinar a IA sobre um padrão, regra ou contexto novo.",
        "label_visibility": "collapsed",
        "key": "memory_knowledge_uploads",
    }

    with memory_col:
        if hasattr(st, "popover"):
            with st.popover("Memória da IA"):
                st.caption("Use apenas quando a IA precisar aprender algo novo para esta análise.")
                memory_uploads = st.file_uploader(**uploader_kwargs) or []
        else:
            with st.expander("Memória da IA", expanded=False):
                st.caption("Use apenas quando a IA precisar aprender algo novo para esta análise.")
                memory_uploads = st.file_uploader(**uploader_kwargs) or []

    return memory_uploads


def _save_uploaded_file(uploaded_file: Any, destination_dir: Path, allowed_extensions: set[str]) -> Path:
    _validate_upload_extension(uploaded_file.name, allowed_extensions)
    destination_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{_timestamp()}_{_safe_filename(uploaded_file.name)}"
    destination = destination_dir / filename
    destination.write_bytes(uploaded_file.getvalue())
    return destination


def save_uploaded_video(uploaded_file: Any) -> str:
    """Salva UploadedFile de video em disco antes de passar para OpenCV."""

    _validate_upload_extension(uploaded_file.name, VIDEO_EXTENSIONS)
    VIDEOS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    output_path = VIDEOS_UPLOAD_DIR / f"{_timestamp()}_{_safe_filename(uploaded_file.name)}"
    with output_path.open("wb") as file:
        file.write(bytes(uploaded_file.getbuffer()))

    if not output_path.exists() or output_path.stat().st_size <= 0:
        raise ValueError("Video enviado nao foi salvo corretamente.")

    return str(output_path)


def _file_ready_for_download(path: Path) -> bool:
    return path.exists() and path.is_file() and path.stat().st_size > 0


def _read_download_bytes(path: Path) -> bytes:
    if not _file_ready_for_download(path):
        raise FileNotFoundError(f"Arquivo nao esta disponivel para download: {path}")
    if path.suffix.lower() == ".xlsx":
        _validate_xlsx_integrity(path)
    with path.open("rb") as file:
        return file.read()


def _validate_xlsx_integrity(path: Path) -> None:
    if not _file_ready_for_download(path):
        raise FileNotFoundError(f"Arquivo Excel nao encontrado ou vazio: {path}")
    try:
        with zipfile.ZipFile(path) as workbook_zip:
            bad_member = workbook_zip.testzip()
    except zipfile.BadZipFile as exc:
        raise ValueError(f"Arquivo gerado nao e um .xlsx valido: {path}") from exc
    except Exception as exc:
        raise ValueError(f"Arquivo Excel corrompido: {path}. Detalhe: {exc}") from exc

    if bad_member is not None:
        raise ValueError(f"Arquivo Excel corrompido no item interno: {bad_member}")


def _workbook_sheetnames(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _validate_template_workbook(path: Path) -> None:
    _validate_xlsx_integrity(path)
    if not _file_ready_for_download(path):
        raise FileNotFoundError(f"Template Excel nao encontrado ou vazio: {path}")
    sheetnames = set(_workbook_sheetnames(path))
    missing = sorted(REQUIRED_TEMPLATE_SHEETS - sheetnames)
    if missing:
        raise ValueError(f"Template nao contem abas obrigatorias: {', '.join(missing)}")


def _validate_output_matches_template(output_path: Path, template_path: Path) -> None:
    _validate_xlsx_integrity(output_path)
    _validate_xlsx_integrity(template_path)
    template_sheets = _workbook_sheetnames(template_path)
    output_sheets = _workbook_sheetnames(output_path)
    allowed_added_sheets = {
        "Sheet2",
        "ENTENDIMENTO_CONVERSAO",
        "STANDARD_CONSOLIDADO",
        "Avisos_Validacao",
        "ALERTAS_VALIDACAO_SPS",
    }
    unexpected_sheets = [
        sheet for sheet in output_sheets if sheet not in template_sheets and sheet not in allowed_added_sheets
    ]
    preserved_template_order = [sheet for sheet in output_sheets if sheet in template_sheets]
    if unexpected_sheets or preserved_template_order != template_sheets:
        raise ValueError(
            "Planilha gerada nao preservou as abas do template padrao. "
            f"Template={template_sheets} | Gerada={output_sheets}"
        )
    if "Sheet2" in output_sheets and output_sheets[1] != "Sheet2":
        raise ValueError("Aba Sheet2 deve permanecer na segunda posicao para integracao.")


def _legacy_generation_lock_path(excel_path: Path) -> Path:
    return excel_path.with_name(f"{excel_path.name}.lock")


def _generation_lock_path(excel_path: Path) -> Path:
    try:
        lock_key = str(excel_path.resolve(strict=False)).casefold()
    except OSError:
        lock_key = str(excel_path.absolute()).casefold()
    digest = hashlib.sha256(lock_key.encode("utf-8", errors="ignore")).hexdigest()[:16]
    return GENERATION_LOCK_DIR / f"{_safe_slug(excel_path.stem)}_{digest}.lock"


def _unlink_lock_file(lock_path: Path, *, raise_on_error: bool) -> None:
    last_error: OSError | None = None
    for attempt in range(3):
        try:
            lock_path.unlink()
            return
        except FileNotFoundError:
            return
        except PermissionError as exc:
            last_error = exc
            if attempt < 2:
                time.sleep(0.1)
        except OSError as exc:
            last_error = exc
            break

    if raise_on_error and last_error is not None:
        raise RuntimeError(
            "Nao foi possivel limpar o arquivo de trava da geracao. "
            "Feche outra geracao em andamento ou reinicie o Streamlit e tente novamente."
        ) from last_error


def _cleanup_stale_lock(lock_path: Path, *, raise_on_error: bool) -> None:
    try:
        if lock_path.exists() and time.time() - lock_path.stat().st_mtime > GENERATION_LOCK_MAX_AGE_S:
            _unlink_lock_file(lock_path, raise_on_error=raise_on_error)
    except FileNotFoundError:
        return
    except OSError as exc:
        if raise_on_error:
            raise RuntimeError(
                "Nao foi possivel verificar o arquivo de trava da geracao. "
                "Feche outra geracao em andamento ou reinicie o Streamlit e tente novamente."
            ) from exc


def _cleanup_legacy_generation_lock(excel_path: Path) -> None:
    _cleanup_stale_lock(_legacy_generation_lock_path(excel_path), raise_on_error=False)


def _acquire_generation_lock(excel_path: Path) -> tuple[int, Path]:
    _cleanup_legacy_generation_lock(excel_path)
    lock_path = _generation_lock_path(excel_path)
    try:
        lock_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise RuntimeError(
            "Nao foi possivel preparar a pasta local de trava da geracao. "
            "Verifique permissoes da pasta temporaria do Windows e tente novamente."
        ) from exc

    _cleanup_stale_lock(lock_path, raise_on_error=True)

    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:
        raise GenerationAlreadyRunningError(
            "Geracao da planilha ja esta em andamento. Aguarde terminar antes de clicar novamente."
        ) from exc
    except PermissionError as exc:
        raise RuntimeError(
            "Sem permissao para criar o arquivo local de trava da geracao. "
            "Feche outra geracao em andamento ou reinicie o Streamlit e tente novamente."
        ) from exc

    try:
        os.write(fd, str(os.getpid()).encode("ascii", errors="ignore"))
    except OSError:
        os.close(fd)
        _unlink_lock_file(lock_path, raise_on_error=False)
        raise
    return fd, lock_path


def _release_generation_lock(fd: int, lock_path: Path) -> None:
    try:
        os.close(fd)
    except OSError:
        pass
    finally:
        if lock_path.exists():
            _unlink_lock_file(lock_path, raise_on_error=False)


def _is_generation_lock_active(excel_path: Path) -> bool:
    _cleanup_legacy_generation_lock(excel_path)
    lock_path = _generation_lock_path(excel_path)
    _cleanup_stale_lock(lock_path, raise_on_error=False)
    return lock_path.exists()


def _sync_generation_state_with_lock(review_state: dict[str, Any]) -> bool:
    output_path = review_state.get("output_path")
    active_lock = bool(output_path) and _is_generation_lock_active(Path(output_path))
    st.session_state["generating_excel"] = active_lock
    return active_lock


def _default_layout_path(posto: str) -> Path | None:
    candidate = LAYOUTS_DIR / f"{posto.strip()}.json"
    if candidate.exists():
        return candidate
    return None


def _extract_text_from_csv(path: Path) -> str:
    import csv

    lines: list[str] = []
    with path.open("r", encoding="utf-8", errors="replace") as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            if row:
                lines.append(" | ".join(str(cell) for cell in row))
    return "\n".join(lines)


def _extract_text_from_docx(path: Path) -> str | None:
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        with zipfile.ZipFile(path, "r") as archive:
            with archive.open("word/document.xml") as document_xml:
                tree = ET.parse(document_xml)
                root = tree.getroot()
                paragraphs = []
                for paragraph in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
                    texts = [node.text for node in paragraph.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t") if node.text]
                    if texts:
                        paragraphs.append("".join(texts))
                return "\n\n".join(paragraphs)
    except Exception:
        return None


def _extract_text_from_xlsx(path: Path) -> str | None:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        texts: list[str] = []
        for sheet in workbook.worksheets:
            texts.append(f"[Planilha: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) for cell in row if cell is not None]
                if row_values:
                    texts.append(" | ".join(row_values))
        workbook.close()
        return "\n".join(texts)
    except Exception:
        return None


def _extract_text_from_uploaded_knowledge_file(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix == ".csv":
        return _extract_text_from_csv(path)
    if suffix == ".docx":
        return _extract_text_from_docx(path)
    if suffix == ".xlsx":
        return _extract_text_from_xlsx(path)
    return None


def save_uploaded_knowledge_files(uploaded_files: list[Any], persist_for_future: bool = True) -> list[str]:
    """Salva arquivos de conhecimento/memória em disco e retorna lista de caminhos."""
    if not uploaded_files:
        return []
    
    root_dir = KNOWLEDGE_UPLOAD_DIR if persist_for_future else SESSION_KNOWLEDGE_DIR
    session_dir = root_dir / f"session_{_timestamp()}"
    session_dir.mkdir(parents=True, exist_ok=True)
    
    saved_paths: list[str] = []
    for uploaded_file in uploaded_files:
        try:
            _validate_upload_extension(uploaded_file.name, KNOWLEDGE_EXTENSIONS)
            filename = f"{_safe_filename(uploaded_file.name)}"
            destination = session_dir / filename
            destination.write_bytes(uploaded_file.getvalue())
            saved_paths.append(str(destination))

            extracted_text = _extract_text_from_uploaded_knowledge_file(destination)
            if extracted_text:
                sidecar = destination.with_suffix(destination.suffix + ".txt")
                sidecar.write_text(extracted_text, encoding="utf-8")
            elif destination.suffix.lower() in {".pdf", ".docx", ".xlsx"}:
                st.warning(
                    f"O arquivo {uploaded_file.name} foi salvo, mas a extração de texto é pendente para este formato. "
                    "A base será atualizada assim que o recurso estiver disponível."
                )
        except Exception as exc:
            st.warning(f"Nao foi possivel salvar arquivo {uploaded_file.name}: {exc}")
    
    return saved_paths


def save_manual_memory_note(note: str, persist_for_future: bool = False) -> str | None:
    """Salva observacao curta como memoria textual local para a analise."""
    cleaned = (note or "").strip()
    if not cleaned:
        return None
    api_key = os.getenv("OPENAI_API_KEY", "")
    if api_key and api_key in cleaned:
        cleaned = cleaned.replace(api_key, "[redacted]")

    root_dir = KNOWLEDGE_UPLOAD_DIR / "manual_notes" if persist_for_future else SESSION_KNOWLEDGE_DIR / "manual_notes"
    root_dir.mkdir(parents=True, exist_ok=True)
    login = str(_session_value("auth_user_login") or "usuario")
    destination = root_dir / f"memoria_manual_{_timestamp()}_{_safe_slug(login)}.md"
    destination.write_text(
        "# Memoria informada pelo usuario para apoiar a analise SPS\n\n"
        f"- data_hora: {datetime.now().isoformat(timespec='seconds')}\n"
        f"- usuario_login: {login}\n"
        "- tipo: nomenclatura/contexto operacional\n\n"
        "## Observacao\n\n"
        f"{cleaned}\n\n"
        "Observacao: este arquivo complementa o RAG local; nao treina nem altera pesos do modelo OpenAI.\n",
        encoding="utf-8",
    )
    return str(destination)


def save_uploaded_layout_image(uploaded_file: Any) -> str | None:
    """Salva imagem de layout/post para spaghetti e retorna o caminho."""
    if uploaded_file is None:
        return None
    
    try:
        _validate_upload_extension(uploaded_file.name, LAYOUT_IMAGE_EXTENSIONS)
        LAYOUTS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        filename = f"{_timestamp()}_{_safe_filename(uploaded_file.name)}"
        destination = LAYOUTS_UPLOAD_DIR / filename
        destination.write_bytes(uploaded_file.getvalue())
        return str(destination)
    except Exception as exc:
        st.warning(f"Nao foi possivel salvar imagem de layout: {exc}")
        return None


def _json_path_for_excel(excel_path: Path) -> Path:
    return excel_path.with_suffix(".json")


def _write_analysis_json(analysis: OperationalAnalysis, excel_path: Path) -> Path:
    analysis = prepare_analysis_for_export(analysis)
    json_path = _json_path_for_excel(excel_path)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(analysis.model_dump(mode="json"), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return json_path


def _read_analysis(json_path: Path) -> OperationalAnalysis:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    return OperationalAnalysis.model_validate(
        normalize_analysis_payload_for_current_schema(data)
    )


def _frame_preview_dir(video_path: Path) -> Path:
    return FRAME_PREVIEW_DIR / video_path.stem


def _extract_preview_frames(video_path: Path, fps: float, max_frames: int) -> list[ExtractedFrame]:
    preview_limit = min(max(max_frames, 1), 5)
    return extract_frames(
        video_path=str(video_path),
        output_dir=str(_frame_preview_dir(video_path)),
        fps=fps,
        max_frames=preview_limit,
    )


def _render_video_status() -> None:
    video_info = st.session_state.get("last_video_upload")
    if not video_info:
        return

    st.subheader("Validação técnica do vídeo")
    st.write(f"Nome: {video_info['name']}")
    st.write(f"Tamanho: {video_info['size_bytes']} bytes")
    st.write(f"Path salvo: `{video_info['path']}`")

    metrics = st.session_state.get("last_pipeline_metrics") or {}
    metadata = metrics.get("metadata") or {}
    windows = metrics.get("windows") or {}
    frames_info = metrics.get("frames") or {}
    if metadata or windows or frames_info:
        metric_cols = st.columns(4)
        metric_cols[0].metric("Duracao do video", f"{float(metadata.get('duration_s', 0.0)):.1f}s")
        metric_cols[1].metric("Frames extraidos", int(frames_info.get("count", 0)))
        metric_cols[2].metric("Janelas", int(windows.get("count", 0)))
        metric_cols[3].metric("Frames/janela", str(windows.get("max_frames_per_window", "-")))
        frames_per_window = windows.get("frames_per_window") or []
        if frames_per_window:
            st.caption(f"Frames enviados por janela: {frames_per_window}")

    frames = st.session_state.get("last_frame_preview") or []
    if not frames:
        st.info("Use a opção de extrair frames para validar a leitura técnica do vídeo.")
        return

    st.write(f"Frames extraídos para prévia: {len(frames)}")
    columns = st.columns(min(len(frames), 5))
    for column, frame in zip(columns, frames, strict=False):
        with column:
            st.image(frame["path"], caption=frame["timestamp_formatado"], use_container_width=True)


def _build_metadata(
    departamento: str,
    linha: str,
    bloco: str,
    posto: str,
    processo: str,
    responsavel: str,
    data_analise: date,
    takt_time_s: float,
    observacoes_usuario: str,
    fonte_video: str | None,
) -> AnalysisMetadata:
    usuario_login = _session_value("auth_user_login")
    usuario_nome = _session_value("auth_user_name")
    usuario_area = _session_value("auth_user_area")
    aceite_responsabilidade_em = _session_value("auth_accepted_responsibility_at")
    return AnalysisMetadata(
        departamento=departamento.strip(),
        linha=linha.strip() or None,
        bloco=bloco.strip() or None,
        posto=posto.strip(),
        processo=processo.strip(),
        responsavel=responsavel.strip(),
        data_analise=data_analise.isoformat(),
        takt_time_s=float(takt_time_s) if takt_time_s > 0 else None,
        ciclo_observado_s=None,
        fonte_video=fonte_video,
        observacoes_gerais=observacoes_usuario.strip() or None,
        usuario_login=str(usuario_login).strip() if usuario_login else None,
        usuario_nome=str(usuario_nome).strip() if usuario_nome else None,
        usuario_area=str(usuario_area).strip() if usuario_area else None,
        aceite_responsabilidade_em=str(aceite_responsabilidade_em).strip()
        if aceite_responsabilidade_em
        else None,
    )


def _compose_user_observations(
    observacoes_usuario: str,
    foco_inicio_processo: str = "",
    fim_esperado_processo: str = "",
    variante_veiculo: str = "",
    eixo_envolvido: str = "",
    lado_envolvido: str = "",
    observacoes_nomenclatura: str = "",
) -> str:
    parts: list[str] = []
    if observacoes_usuario.strip():
        parts.append(observacoes_usuario.strip())
    if foco_inicio_processo.strip():
        parts.append(f"Foco/inicio esperado do processo: {foco_inicio_processo.strip()}")
    if fim_esperado_processo.strip():
        parts.append(f"Fim esperado do processo: {fim_esperado_processo.strip()}")
    if variante_veiculo.strip():
        parts.append(f"Configuracao/variante do veiculo: {variante_veiculo.strip()}")
    if eixo_envolvido.strip():
        parts.append(f"Eixo envolvido: {eixo_envolvido.strip()}")
    if lado_envolvido.strip():
        parts.append(f"Lado envolvido: {lado_envolvido.strip()}")
    if observacoes_nomenclatura.strip():
        parts.append(f"Observacoes de nomenclatura: {observacoes_nomenclatura.strip()}")
    return "\n".join(parts)


def _format_seconds_percent(seconds: float, percent: float) -> str:
    return f"{seconds:.1f}s | {percent:.1f}%"


def _render_summary(analysis: OperationalAnalysis) -> None:
    summary = analysis.resumo_tempos
    col_av, col_nav, col_d, col_total, col_takt = st.columns(5)
    col_av.metric("AV", _format_seconds_percent(summary.av_s, summary.av_percent))
    col_nav.metric("NAV", _format_seconds_percent(summary.nav_s, summary.nav_percent))
    col_d.metric("D", _format_seconds_percent(summary.d_s, summary.d_percent))
    col_total.metric("Total", f"{summary.total_s:.1f}s")
    if summary.folga_vs_takt_s is None:
        col_takt.metric("Vs takt", "n/a")
    else:
        label = "Folga vs takt" if summary.folga_vs_takt_s >= 0 else "Estouro vs takt"
        col_takt.metric(label, f"{summary.folga_vs_takt_s:.1f}s")


def _render_alerts(analysis: OperationalAnalysis) -> None:
    if not analysis.alertas_validacao:
        return
    st.subheader("Alertas de validação")
    for alert in analysis.alertas_validacao:
        st.warning(alert)


def _has_quality_alerts_for_export(analysis: OperationalAnalysis) -> bool:
    try:
        prepared = prepare_analysis_for_export(analysis)
        rows = build_quality_alert_rows(prepared)
        from app.analysis.quality_gate import validate_analysis_quality

        gate = validate_analysis_quality(prepared, None, None)
        if gate.alerts:
            return True
    except Exception:
        return bool(analysis.alertas_validacao)
    return any(row.get("tipo") != "Sem alerta" for row in rows)


def _render_downloads(excel_path: Path, json_path: Path) -> None:
    st.subheader("Downloads")
    if not _file_ready_for_download(excel_path):
        st.error("Planilha nao foi gerada. Verifique erros no pipeline.")
        st.caption(f"Caminho esperado: {excel_path}")
        return
    if not _file_ready_for_download(json_path):
        st.error("JSON da analise nao foi gerado. Verifique erros no pipeline.")
        st.caption(f"Caminho esperado: {json_path}")
        return

    excel_bytes = _read_download_bytes(excel_path)
    json_bytes = _read_download_bytes(json_path)
    st.caption(f"Excel: {excel_path}")
    st.caption(f"JSON: {json_path}")
    _validate_xlsx_integrity(excel_path)
    template_path = st.session_state.get("last_template_path")
    if not template_path and DEFAULT_TEMPLATE_PATH.exists():
        template_path = str(DEFAULT_TEMPLATE_PATH)
    if template_path:
        try:
            _validate_output_matches_template(excel_path, Path(template_path))
        except Exception as exc:
            st.error(f"A planilha gerada nao confere com o template usado: {exc}")
            return

    col_excel, col_json = st.columns(2)
    with col_excel:
        st.download_button(
            "Baixar planilha Scania preenchida",
            data=excel_bytes,
            file_name=excel_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_excel_{excel_path.name}",
        )
    with col_json:
        st.download_button(
            "Baixar JSON da análise",
            data=json_bytes,
            file_name=json_path.name,
            mime="application/json",
            key=f"download_json_{json_path.name}",
        )


def _generate_excel_and_json_for_analysis(
    analysis: OperationalAnalysis,
    excel_path: Path,
    run_config: dict[str, Any],
) -> Path:
    template = run_config.get("template_path") or get_default_template_path()
    if template is None:
        raise FileNotFoundError("Template padrao Scania nao encontrado para gerar a versao corrigida.")
    template_path = Path(template)
    assert_analysis_can_generate_excel(analysis)
    generated_excel = write_analysis_to_template(
        analysis=analysis,
        template_path=str(template_path),
        output_path=str(excel_path),
        fill_standard=bool(run_config.get("fill_standard", True)),
        standard_export_mode=str(run_config.get("standard_export_mode", "standard_consolidado")),
        insert_charts=bool(run_config.get("insert_charts", False)),
        insert_spaghetti=bool(run_config.get("insert_spaghetti", False)),
        layout_path=run_config.get("layout_path"),
        layout_image_path=run_config.get("layout_image_path"),
    )
    generated_excel_path = Path(generated_excel)
    json_path = _write_analysis_json(analysis, generated_excel_path)
    _validate_output_matches_template(generated_excel_path, template_path)
    return json_path


def _validate_feedback_was_applied_lazy(
    analysis: OperationalAnalysis,
    feedback_text: str,
) -> list[str]:
    """Load feedback quality gate lazily so Streamlit startup is resilient."""

    try:
        from app.analysis import quality_gate
    except Exception as exc:
        return [f"Quality gate de feedback nao carregado: {exc}"]

    helper = getattr(quality_gate, "validate_feedback_was_applied", None)
    if helper is None:
        return []
    return helper(analysis, feedback_text)


def _handle_correction_rerun(
    *,
    analysis: OperationalAnalysis,
    excel_path: Path,
    json_path: Path,
    observation: str,
    save_as_memory: bool,
) -> None:
    cleaned_observation = observation.strip()
    if not cleaned_observation:
        st.warning("Informe uma observacao para orientar a correcao.")
        return

    video_path = st.session_state.get("last_video_path") or analysis.metadata.fonte_video
    if not video_path:
        st.error("Nao foi possivel localizar o video original salvo para refazer a analise.")
        return

    run_config = st.session_state.get("last_run_config") or {
        "template_path": str(get_default_template_path()),
        "fill_standard": True,
        "standard_export_mode": "standard_consolidado",
        "insert_charts": False,
        "insert_spaghetti": False,
        "layout_path": None,
        "layout_image_path": st.session_state.get("layout_image_path"),
    }
    history = list(st.session_state.get("correction_history") or [])
    plan = prepare_correction_rerun(
        excel_path,
        cleaned_observation,
        history=history,
        login=str(_session_value("auth_user_login") or ""),
        base_stem=st.session_state.get("analysis_output_base_stem"),
    )
    new_excel_path = Path(plan["output_path"])
    correction_context_path = save_correction_context_note(
        analysis,
        cleaned_observation,
        login=str(_session_value("auth_user_login") or ""),
    )
    feedback_path: Path | None = None
    if save_as_memory:
        feedback_path = save_feedback_memory(
            cleaned_observation,
            metadata=analysis.metadata,
            analysis_path=json_path,
            login=str(_session_value("auth_user_login") or ""),
        )

    corrected_metadata = analysis.metadata.model_copy(
        update={
            "observacoes_gerais": _compose_user_observations(
                analysis.metadata.observacoes_gerais or "",
                "Revisar analise anterior com base na observacao do usuario.",
                cleaned_observation,
            ),
            "fonte_video": str(video_path),
        }
    )
    knowledge_paths = list(st.session_state.get("last_knowledge_paths") or [])
    knowledge_paths.append(str(correction_context_path))
    if feedback_path:
        knowledge_paths.append(str(feedback_path))

    with st.spinner("Refazendo analise SPS com observacoes do usuario..."):
        corrected_analysis = _run_analysis_only_compat(
            video_path=str(video_path),
            output_path=str(new_excel_path),
            metadata=corrected_metadata,
            knowledge_root=str(KNOWLEDGE_ROOT),
            knowledge_paths=knowledge_paths,
            provider_name=DEFAULT_PROVIDER,
            fps=DEFAULT_EXTRACTION_FPS,
            max_frames=None,
            window_seconds=int(QUALITY_PRESETS[PRODUCTION_QUALITY_MODE]["window_seconds"]),
            max_frames_per_window=int(QUALITY_PRESETS[PRODUCTION_QUALITY_MODE]["max_frames_per_window"]),
            reprocess_low_confidence=True,
            detail_window=str(QUALITY_PRESETS[PRODUCTION_QUALITY_MODE]["detail_window"]),
            quality_mode=PRODUCTION_QUALITY_MODE,
            resume_from_checkpoint=False,
        )
        feedback_alerts = _validate_feedback_was_applied_lazy(corrected_analysis, cleaned_observation)
        if feedback_alerts:
            alerts = list(corrected_analysis.alertas_validacao)
            for alert in feedback_alerts:
                if alert not in alerts:
                    alerts.append(alert)
            corrected_analysis = OperationalAnalysis.model_validate(
                corrected_analysis.model_dump() | {"alertas_validacao": alerts}
            )
        new_json_path = _generate_excel_and_json_for_analysis(corrected_analysis, new_excel_path, run_config)

    st.session_state["correction_history"] = append_correction_history(
        history,
        observation=cleaned_observation,
        previous_excel_path=excel_path,
        previous_json_path=json_path,
        new_excel_path=new_excel_path,
        new_json_path=new_json_path,
        feedback_path=feedback_path,
    )
    st.session_state["last_pipeline_result"] = {
        "excel_path": str(new_excel_path),
        "json_path": str(new_json_path),
    }
    st.session_state["last_excel_path"] = str(new_excel_path)
    st.session_state["last_json_path"] = str(new_json_path)
    st.session_state["last_analysis"] = corrected_analysis.model_dump(mode="json")
    st.session_state["last_metadata"] = corrected_analysis.metadata.model_dump(mode="json")
    st.session_state["last_knowledge_paths"] = knowledge_paths
    st.success("Nova versao da analise gerada com observacoes.")
    st.rerun()


def _render_correction_panel(analysis: OperationalAnalysis, excel_path: Path, json_path: Path) -> None:
    st.subheader("Correção da análise")
    st.caption(
        "Use este campo para orientar a IA caso a análise não esteja aderente ao processo, "
        "à nomenclatura Scania ou ao padrão esperado."
    )
    observation = st.text_area(
        "Observações para correção",
        value=st.session_state.get("correction_observation_text", ""),
        placeholder=(
            "Ex.: O processo começa quando a parafusadeira pneumática é pega. "
            "A etapa inicial de deslocamento não deve ser considerada. Usar o termo VR de pneu, "
            "não equipamento. Reescrever as microetapas no modo imperativo do padrão Scania."
        ),
        height=110,
        key="correction_observation_text",
    )
    save_as_memory = st.checkbox(
        "Salvar observação como memória interna pendente de validação",
        value=False,
        key="correction_save_memory",
    )
    col_rerun, col_save, col_clear = st.columns(3)
    with col_rerun:
        if st.button("Refazer análise com observações", type="primary", key="correction_rerun_button"):
            _handle_correction_rerun(
                analysis=analysis,
                excel_path=excel_path,
                json_path=json_path,
                observation=observation,
                save_as_memory=save_as_memory,
            )
    with col_save:
        if st.button("Salvar observação como memória interna", key="correction_save_memory_button"):
            if observation.strip():
                path = save_feedback_memory(
                    observation,
                    metadata=analysis.metadata,
                    analysis_path=json_path,
                    login=str(_session_value("auth_user_login") or ""),
                )
                st.success(f"Memória salva como pendente de validação: {path}")
            else:
                st.warning("Informe uma observação antes de salvar como memória.")
    with col_clear:
        if st.button("Limpar observação", key="correction_clear_button"):
            st.session_state["correction_observation_text"] = ""
            st.rerun()

    history = st.session_state.get("correction_history") or []
    if history:
        with st.expander("Histórico de correções desta análise", expanded=False):
            st.dataframe(history, use_container_width=True, hide_index=True)


def _render_results(analysis: OperationalAnalysis, excel_path: Path, json_path: Path) -> None:
    st.subheader("Resumo AV/NAV/D")
    _render_summary(analysis)
    _render_alerts(analysis)
    if _has_quality_alerts_for_export(analysis):
        st.warning(
            "A planilha foi gerada com alertas de validacao SPS. Revise a ultima aba "
            "'ALERTAS_VALIDACAO_SPS' antes de utilizar o padrao."
        )

    st.subheader("Tabela da análise do processo")
    micro_rows = [
        {
            "Nº da etapa": step.numero,
            "Descrição técnica detalhada": step.etapa_detalhada,
            "Início observável": step.inicio_formatado,
            "Fim observável": step.fim_formatado,
            "Duração (s)": step.duracao_s,
            "Tempo acumulado": step.tempo_acumulado_formatado or "",
            "Classificação AV/NAV/D": step.classificacao,
            "Justificativa técnica": step.justificativa_tecnica,
            "Confiança": step.confianca,
            "Observação/ferramenta": step.ferramenta_observacao,
        }
        for step in analysis.microetapas
    ]
    st.dataframe(micro_rows, use_container_width=True, hide_index=True)

    st.subheader("Desperdícios")
    waste_rows = [
        {
            "Nº da etapa": step.numero,
            "Duração (s)": step.duracao_s,
            "Descrição": step.etapa_detalhada,
            "Evidência": step.evidencia_visual or "",
            "Justificativa": step.justificativa_tecnica,
        }
        for step in analysis.microetapas
        if step.classificacao == "D"
    ]
    if waste_rows:
        st.dataframe(waste_rows, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhuma microetapa classificada como desperdício D.")

    st.subheader("Melhorias")
    improvement_rows = [
        {
            "Microetapa": item.microetapa_numero,
            "Inicio": item.inicio_formatado,
            "Fim": item.fim_formatado,
            "Duracao_s": item.duracao_s,
            "Desperdicio": item.descricao_desperdicio,
            "Tipo": item.tipo_desperdicio,
            "Causa observavel": item.causa_observavel,
            "Sugestao": item.sugestao_pratica,
            "Prioridade": item.prioridade,
            "Validacao gemba/SPS": "Sim" if item.requer_validacao_gemba else "Nao",
        }
        for item in analysis.melhorias
    ]
    if improvement_rows:
        st.dataframe(improvement_rows, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhuma melhoria estruturada foi retornada para esta analise.")

    _render_correction_panel(analysis, excel_path, json_path)
    _render_downloads(excel_path, json_path)


def _generate_excel_from_review_state(analysis: OperationalAnalysis, review_state: dict[str, Any]) -> None:
    analysis = _rehydrate_analysis_for_export(analysis)
    assert_analysis_can_generate_excel(analysis)
    excel_path = Path(review_state["output_path"])
    template_path = Path(review_state["template_path"])
    temp_excel_path = excel_path.with_name(f"{excel_path.stem}_{_timestamp()}_building.xlsx")
    lock_fd, lock_path = _acquire_generation_lock(excel_path)
    try:
        generated_excel = write_analysis_to_template(
            analysis=analysis,
            template_path=str(template_path),
            output_path=str(temp_excel_path),
            fill_standard=review_state["fill_standard"],
            standard_export_mode=review_state.get("standard_export_mode", "standard_consolidado"),
            insert_charts=review_state["insert_charts"],
            insert_spaghetti=review_state["insert_spaghetti"],
            layout_path=review_state["layout_path"],
            layout_image_path=review_state.get("layout_image_path"),
        )
        generated_temp_path = Path(generated_excel)
        if not _file_ready_for_download(generated_temp_path):
            raise FileNotFoundError("Planilha temporaria nao foi gerada. Verifique erros no pipeline.")
        _validate_output_matches_template(generated_temp_path, template_path)
        generated_temp_path.replace(excel_path)
    except Exception:
        if temp_excel_path.exists():
            temp_excel_path.unlink()
        raise
    finally:
        _release_generation_lock(lock_fd, lock_path)

    json_path = _write_analysis_json(analysis, excel_path)
    if not _file_ready_for_download(excel_path):
        raise FileNotFoundError("Planilha nao foi gerada. Verifique erros no pipeline.")
    if not _file_ready_for_download(json_path):
        raise FileNotFoundError("JSON da analise nao foi gerado. Verifique erros no pipeline.")
    _validate_output_matches_template(excel_path, template_path)

    st.session_state["last_pipeline_result"] = {
        "excel_path": str(excel_path),
        "json_path": str(json_path),
    }
    st.session_state["last_template_path"] = review_state["template_path"]
    st.session_state["last_excel_path"] = str(excel_path)
    st.session_state["last_json_path"] = str(json_path)
    st.session_state["last_analysis"] = analysis.model_dump(mode="json")
    st.session_state["last_metadata"] = analysis.metadata.model_dump(mode="json")
    if review_state.get("video_path"):
        st.session_state["last_video_path"] = review_state["video_path"]
    st.session_state["last_knowledge_paths"] = review_state.get("knowledge_paths", [])
    st.session_state["last_run_config"] = {
        "template_path": review_state["template_path"],
        "fill_standard": review_state["fill_standard"],
        "standard_export_mode": review_state.get("standard_export_mode", "standard_consolidado"),
        "insert_charts": review_state["insert_charts"],
        "insert_spaghetti": review_state["insert_spaghetti"],
        "layout_path": review_state["layout_path"],
        "layout_image_path": review_state.get("layout_image_path"),
    }
    st.session_state.setdefault("analysis_output_base_stem", excel_path.stem)
    st.session_state.pop("pending_review", None)


def _rehydrate_analysis_for_export(analysis) -> OperationalAnalysis:
    """Recreate analysis with the current schema before review/export paths."""

    return OperationalAnalysis.model_validate(
        normalize_analysis_payload_for_current_schema(analysis)
    )


def _save_failed_export_recovery(analysis, review_state: dict[str, Any], exc: Exception) -> Path:
    """Persist current analysis if Excel generation fails, so it is not lost."""

    recovery_dir = Path("data/outputs/recovery")
    recovery_dir.mkdir(parents=True, exist_ok=True)
    path = recovery_dir / f"analise_preservada_{_timestamp()}.json"
    try:
        analysis_payload = _rehydrate_analysis_for_export(analysis).model_dump(mode="json")
    except Exception:
        analysis_payload = analysis.model_dump(mode="json") if hasattr(analysis, "model_dump") else {"analysis": str(analysis)}
    payload = {
        "erro_exportacao": str(exc),
        "review_state_keys": sorted(review_state.keys()),
        "analysis": analysis_payload,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _render_review_editor(review_state: dict[str, Any]) -> None:
    analysis = _rehydrate_analysis_for_export(review_state["analysis"])

    st.subheader("Tabela da análise do processo")
    _render_summary(analysis)
    _render_alerts(analysis)
    st.warning(
        "A exportacao CSV da tabela editavel nao e a planilha Scania. "
        "Para baixar o modelo padrao preenchido, use o botao de gerar planilha .xlsx."
    )
    st.caption("Edite tempos, classificacao e justificativas antes de gerar o Excel final.")

    generating_excel = _sync_generation_state_with_lock(review_state)
    if generating_excel:
        st.info("Geracao da planilha em andamento. Aguarde.")

    if st.button(
        "Gerar planilha Scania (.xlsx) com analise atual",
        type="primary",
        disabled=generating_excel,
    ):
        st.session_state["generating_excel"] = True
        try:
            with st.spinner("Gerando e validando planilha Scania .xlsx..."):
                _generate_excel_from_review_state(analysis, review_state)
            st.success("Planilha Scania .xlsx gerada com sucesso.")
            if _has_quality_alerts_for_export(analysis):
                st.warning(
                    "A planilha foi gerada com alertas de validacao SPS. Revise a ultima aba "
                    "'ALERTAS_VALIDACAO_SPS' antes de utilizar o padrao."
                )
            st.session_state["generating_excel"] = False
            st.rerun()
        except GenerationAlreadyRunningError as exc:
            st.warning(str(exc))
            return
        except Exception as exc:  # pragma: no cover - comportamento visual.
            recovery_path = _save_failed_export_recovery(analysis, review_state, exc)
            st.error(f"Nao foi possivel gerar a planilha Scania: {exc}")
            st.warning(f"A analise foi preservada em: {recovery_path}")
            return
        finally:
            _sync_generation_state_with_lock(review_state)

    editor_kwargs: dict[str, Any] = {
        "use_container_width": True,
        "hide_index": True,
        "num_rows": "fixed",
        "key": review_state["editor_key"],
    }
    if hasattr(st, "column_config"):
        editor_kwargs["column_config"] = {
            "numero": st.column_config.NumberColumn("Numero", disabled=True),
            "inicio_s": st.column_config.NumberColumn("Inicio (s)", step=0.1, format="%.1f"),
            "fim_s": st.column_config.NumberColumn("Fim (s)", step=0.1, format="%.1f"),
            "duracao_s": st.column_config.NumberColumn("Duracao (s)", step=0.1, format="%.1f"),
            "tempo_acumulado_s": st.column_config.NumberColumn("Tempo acumulado (s)", disabled=True),
            "tempo_acumulado_formatado": st.column_config.TextColumn("Tempo acumulado", disabled=True),
            "classificacao": st.column_config.SelectboxColumn("AV/NAV/D", options=["AV", "NAV", "D"]),
            "etapa_detalhada": st.column_config.TextColumn("Etapa detalhada"),
            "justificativa_tecnica": st.column_config.TextColumn("Justificativa"),
            "ferramenta_observacao": st.column_config.TextColumn("Ferramenta/observacao"),
            "confianca": st.column_config.NumberColumn("Confianca", disabled=True),
        }

    edited_df = st.data_editor(analysis_to_dataframe(analysis), **editor_kwargs)

    col_generate, col_discard = st.columns([1, 1])
    with col_generate:
        generate = st.button("Gerar Excel final revisado", type="primary", disabled=generating_excel)
    with col_discard:
        discard = st.button("Descartar revisao")

    if discard:
        st.session_state.pop("pending_review", None)
        st.info("Revisao descartada.")
        st.rerun()

    if not generate:
        return

    try:
        reviewed_analysis = dataframe_to_analysis(edited_df, analysis)
        st.session_state["generating_excel"] = True
        with st.spinner("Gerando e validando Excel final revisado..."):
            _generate_excel_from_review_state(reviewed_analysis, review_state)
        st.success("Excel final revisado gerado com sucesso.")
        if _has_quality_alerts_for_export(reviewed_analysis):
            st.warning(
                "A planilha foi gerada com alertas de validacao SPS. Revise a ultima aba "
                "'ALERTAS_VALIDACAO_SPS' antes de utilizar o padrao."
            )
        st.session_state["generating_excel"] = False
        st.rerun()
    except GenerationAlreadyRunningError as exc:
        st.warning(str(exc))
    except Exception as exc:  # pragma: no cover - comportamento visual.
        recovery_path = _save_failed_export_recovery(analysis, review_state, exc)
        st.error(f"Revisao invalida: {exc}")
        st.warning(f"A analise foi preservada em: {recovery_path}")
    finally:
        _sync_generation_state_with_lock(review_state)


def _render_app() -> None:
    st.set_page_config(page_title="SPS Process Analysis AI", layout="wide")
    if not st.session_state.get("auth_authenticated"):
        render_login_page()
        st.stop()

    _apply_brand_styles()
    _render_authenticated_user_bar()
    memory_payload = _render_memory_sidebar()

    st.markdown(
        """
        <section class="sps-hero">
            <h1>SPS Process Analysis AI</h1>
            <p>Análise operacional com IA para decomposição de microetapas, classificação AV/NAV/D, tempos, desperdícios, melhorias e preenchimento da planilha padrão Scania.</p>
            <div class="sps-provider">Provider ativo: OpenAI</div>
        </section>
        """,
        unsafe_allow_html=True,
    )

    default_template = get_default_template_path()
    if default_template is None:
        st.error("🚨 Template padrão Scania não encontrado em data/templates/. A aplicação não funcionará sem o arquivo: PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx")
        return
    else:
        st.caption(f"✓ Template padrão carregado: {default_template.name}")

    st.markdown("---")

    # Seção 1: Vídeo da operação
    st.subheader("1. Vídeo da operação")
    video_upload = st.file_uploader(
        "Vídeo de entrada",
        type=["mp4", "mov", "avi", "mkv"],
        help="O vídeo será salvo localmente, convertido em frames e enviado ao provider OpenAI para análise."
    )
    st.caption("Limite configurado: até 1024 MB por arquivo.")
    if video_upload is not None:
        st.caption(f"Arquivo selecionado: {video_upload.name} | tamanho aproximado: {len(video_upload.getbuffer()) / (1024 * 1024):.2f} MB")

    st.markdown("---")

    # Seção 2: Layout do posto para spaghetti
    st.subheader("2. Layout do posto para mapa de espaguete")
    layout_image_upload = st.file_uploader(
        "Imagem/layout do posto ou bloco para mapa de espaguete (opcional)",
        type=["png", "jpg", "jpeg"],
        help="Anexe uma imagem PNG/JPG do bloco, layout ou posição para apoiar a geração do mapa de espaguete."
    )

    st.markdown("---")

    # Seção 3: Opções da análise
    st.subheader("3. Opções da análise")
    st.caption("A análise será executada em qualidade máxima / produção.")
    option_cols = st.columns(4)
    with option_cols[0]:
        fill_standard = st.checkbox("Preencher abas Standard", value=True, help="Preenche as abas padrão do template com dados da análise")
    with option_cols[1]:
        insert_charts = st.checkbox("Incluir gráfico de balanceamento", value=False, help="Insere gráfico de distribuição de tempos no Excel")
    with option_cols[2]:
        insert_spaghetti = st.checkbox("Incluir mapa de espaguete", value=False, help="Insere mapa de movimentação do processo")
    with option_cols[3]:
        include_memory_files = st.checkbox(
            "Incluir arquivos de memória anexados nesta análise, se houver",
            value=bool(memory_payload["enabled"]),
            help="Usa a área auxiliar Memória da IA como complemento do RAG local.",
        )

    quality_mode = PRODUCTION_QUALITY_MODE
    preset = QUALITY_PRESETS[quality_mode]
    window_seconds = int(preset["window_seconds"])
    max_frames_per_window = int(preset["max_frames_per_window"])
    reprocess_low_confidence = bool(preset["reprocess_low_confidence"])
    extract_video_frames = False

    if os.getenv("APP_ENV", "local").strip().lower() == "dev":
        with st.expander("Configurações técnicas avançadas — uso de desenvolvimento", expanded=False):
            extract_video_frames = st.checkbox("Extrair frames do vídeo", value=False, help="Extrai e exibe frames para validação técnica")
            quality_mode = st.selectbox(
                "Qualidade da analise",
                options=list(QUALITY_PRESETS.keys()),
                index=0,
            )
            preset = QUALITY_PRESETS[quality_mode]
            window_seconds = st.number_input(
                "Duracao da janela (s)",
                min_value=5,
                max_value=60,
                value=int(preset["window_seconds"]),
                step=1,
            )
            max_frames_per_window = st.number_input(
                "Frames por janela",
                min_value=3,
                max_value=24,
                value=int(preset["max_frames_per_window"]),
                step=1,
            )
            reprocess_low_confidence = st.checkbox(
                "Reprocessar baixa confianca",
                value=bool(preset["reprocess_low_confidence"]),
            )

    layout_json_upload = None
    if insert_spaghetti:
        layout_json_upload = st.file_uploader(
            "Layout JSON do posto (opcional)",
            type=["json"],
            help="Envie JSON com coordenadas do layout para gerar o mapa de espaguete quando disponível.",
        )

    st.markdown("---")

    # Seção 4: Metadados
    st.subheader("4. Metadados")
    col_1, col_2, col_3 = st.columns(3)
    with col_1:
        departamento = st.text_input("Departamento", value="", help="Ex: FUNCTION AREA 5")
        bloco = st.text_input("Bloco", value="", help="Identificação do bloco/célula")
    with col_2:
        linha = st.text_input("Linha", value="", help="Identificação da linha")
        posto = st.text_input("Posto", value="", help="Identificação do posto")
    with col_3:
        processo = st.text_input("Processo", value="", help="Descrição do processo")
        data_analise = st.date_input("Data da análise", value=date.today(), help="Data em que a análise foi realizada")

    responsavel = st.text_input(
        "Responsável",
        value=_authenticated_responsavel_label(),
        help="Responsável pela operação ou análise",
    )
    takt_time_s = st.number_input("Takt time médio (segundos)", min_value=0.0, value=330.0, step=1.0, help="Tempo ciclo esperado")
    observacoes_usuario = st.text_area("Observações do usuário", value="", height=80, help="Notas adicionais relevantes para a análise")
    foco_inicio_processo = st.text_input(
        "Foco / início esperado do processo (opcional)",
        value="",
        help="Ex.: considerar o início quando a ferramenta é pega ou quando a peça entra no posto.",
    )
    fim_esperado_processo = st.text_input(
        "Fim esperado do processo (opcional)",
        value="",
        help="Ex.: considerar o fim após apontamento, liberação ou retorno do dispositivo.",
    )
    with st.expander("Contexto operacional adicional", expanded=False):
        variante_veiculo = st.text_input(
            "Configuração/variante do veículo (opcional)",
            value="",
            help="Ex.: 8x2, 6x4, 4x2.",
        )
        eixo_envolvido = st.text_input(
            "Eixo envolvido, se conhecido (opcional)",
            value="",
            help="Ex.: primeiro eixo, segundo eixo, terceiro eixo, último eixo.",
        )
        lado_envolvido = st.text_input(
            "Lado envolvido, se conhecido (opcional)",
            value="",
            help="Ex.: LD ou LE.",
        )
        observacoes_nomenclatura = st.text_area(
            "Observações de nomenclatura (opcional)",
            value="",
            height=70,
            help="Ex.: neste posto, o dispositivo de movimentação do pneu é chamado de VR de pneu.",
        )

    st.markdown("---")

    # Seção 5: Processamento
    st.subheader("5. Processamento")
    submitted = st.button("▶ Processar análise SPS", type="primary", use_container_width=False)

    if submitted:
        _handle_submit(
            video_upload=video_upload,
            template_upload=None,  # Não há mais upload de template
            layout_image_upload=layout_image_upload,
            layout_json_upload=layout_json_upload,
            knowledge_uploads=memory_payload["uploads"] if include_memory_files else [],
            memory_note=memory_payload["note"] if include_memory_files else "",
            persist_memory_for_future=bool(memory_payload["persist_for_future"]),
            provider_name=DEFAULT_PROVIDER,
            fill_standard=fill_standard,
            insert_charts=insert_charts,
            insert_spaghetti=insert_spaghetti,
            extract_video_frames=extract_video_frames,
            fps=DEFAULT_EXTRACTION_FPS,
            max_frames_value=DEFAULT_ANALYSIS_MAX_FRAMES,
            quality_mode=quality_mode,
            window_seconds=int(window_seconds),
            max_frames_per_window=int(max_frames_per_window),
            reprocess_low_confidence=bool(reprocess_low_confidence),
            departamento=departamento,
            linha=linha,
            bloco=bloco,
            posto=posto,
            processo=processo,
            responsavel=responsavel,
            data_analise=data_analise,
            takt_time_s=float(takt_time_s),
            observacoes_usuario=_compose_user_observations(
                observacoes_usuario,
                foco_inicio_processo,
                fim_esperado_processo,
                variante_veiculo,
                eixo_envolvido,
                lado_envolvido,
                observacoes_nomenclatura,
            ),
        )

    st.markdown("---")

    # Seção 6: Resultado da análise
    _render_video_status()

    review_state = st.session_state.get("pending_review")
    if review_state:
        st.subheader("6. Resultado da análise - Revisão de microetapas")
        _render_review_editor(review_state)
        return

    result = st.session_state.get("last_pipeline_result")
    if result:
        st.subheader("6. Resultado da análise")
        try:
            excel_path = Path(result["excel_path"])
            json_path = Path(result["json_path"])
            analysis = _read_analysis(json_path)
            _render_results(analysis, excel_path, json_path)
        except Exception as exc:  # pragma: no cover - erro exibido pela UI.
            st.error(f"Não foi possível carregar o resultado gerado: {exc}")


def _handle_submit(
    *,
    video_upload: Any,
    template_upload: Any,  # Será None sempre agora
    layout_image_upload: Any,
    layout_json_upload: Any,
    knowledge_uploads: list[Any],
    memory_note: str,
    persist_memory_for_future: bool,
    provider_name: str,
    fill_standard: bool,
    insert_charts: bool,
    insert_spaghetti: bool,
    extract_video_frames: bool,
    fps: float,
    max_frames_value: int,
    quality_mode: str,
    window_seconds: int,
    max_frames_per_window: int,
    reprocess_low_confidence: bool,
    departamento: str,
    linha: str,
    bloco: str,
    posto: str,
    processo: str,
    responsavel: str,
    data_analise: date,
    takt_time_s: float,
    observacoes_usuario: str,
) -> None:
    for key in ("last_pipeline_result", "last_excel_path", "last_json_path", "last_analysis", 
                "uploaded_video_path", "layout_image_path", "uploaded_knowledge_paths", "extracted_frames",
                "last_pipeline_metrics"):
        st.session_state.pop(key, None)
    st.session_state["correction_history"] = []

    if not departamento.strip() or not posto.strip() or not processo.strip() or not responsavel.strip():
        st.error("Preencha departamento, posto, processo e responsável antes de processar.")
        return
    if video_upload is None:
        st.error("Anexe o vídeo de entrada antes de processar a análise.")
        return

    # Salva vídeo
    saved_video_path: Path | None = None
    if video_upload is not None:
        try:
            saved_video_path = Path(save_uploaded_video(video_upload))
            st.session_state["uploaded_video_path"] = str(saved_video_path)
            st.session_state["last_video_upload"] = {
                "name": video_upload.name,
                "size_bytes": saved_video_path.stat().st_size,
                "path": str(saved_video_path),
            }
        except Exception as exc:
            st.error(f"Não foi possível salvar o vídeo enviado: {exc}")
            return

        # Extrai frames preview se solicitado
        if extract_video_frames:
            try:
                preview_frames = _extract_preview_frames(
                    saved_video_path,
                    fps=fps,
                    max_frames=max_frames_value or PREVIEW_MAX_FRAMES,
                )
                st.session_state["extracted_frames"] = [
                    {
                        "index": frame.index,
                        "timestamp_s": frame.timestamp_s,
                        "timestamp_formatado": frame.timestamp_formatado,
                        "path": frame.path,
                    }
                    for frame in preview_frames
                ]
                st.session_state["last_frame_preview"] = st.session_state["extracted_frames"]
            except Exception as exc:
                st.error(f"Não foi possível extrair frames do vídeo: {exc}")
                return

    # Salva imagem de layout se fornecida
    layout_image_path: str | None = None
    if layout_image_upload is not None:
        layout_image_path = save_uploaded_layout_image(layout_image_upload)
        if layout_image_path:
            st.session_state["layout_image_path"] = layout_image_path

    # Salva arquivos de conhecimento se fornecidos
    knowledge_paths: list[str] = []
    if knowledge_uploads:
        try:
            knowledge_paths = save_uploaded_knowledge_files(
                knowledge_uploads,
                persist_for_future=persist_memory_for_future,
            )
            st.session_state["uploaded_knowledge_paths"] = knowledge_paths
            if knowledge_paths:
                st.success(f"✓ {len(knowledge_paths)} arquivo(s) de conhecimento salvos com sucesso.")
        except Exception as exc:
            st.warning(f"Problema ao salvar arquivos de conhecimento: {exc}")
    manual_memory_path = save_manual_memory_note(memory_note, persist_for_future=persist_memory_for_future)
    if manual_memory_path:
        knowledge_paths.append(manual_memory_path)
        st.session_state["uploaded_knowledge_paths"] = knowledge_paths
        st.caption("ObservaÃ§Ã£o de memÃ³ria adicionada ao contexto desta anÃ¡lise.")

    # Valida credencial OpenAI
    if provider_name == "openai" and not _has_openai_api_key():
        st.error(OPENAI_API_KEY_ERROR_UI)
        return

    if provider_name == "openai" and saved_video_path is None:
        st.error("Análise real de vídeo exige vídeo anexado.")
        return

    # Usa template padrão (não há mais upload)
    template_path = get_default_template_path()
    if template_path is None or not template_path.exists():
        st.error("🚨 Template padrão Scania não encontrado. Verifique que o arquivo existe em: data/templates/PMGS_P1_Analise_SPS_CORRIGIDA_2026-03-18.xlsx")
        return
    
    try:
        _validate_template_workbook(template_path)
    except Exception as exc:
        st.error(f"Template padrão inválido: {exc}")
        return

    # Layout JSON (busca por padrão se spaghetti está marcado)
    layout_json_path: Path | None = None
    if insert_spaghetti and layout_json_upload is not None:
        layout_json_path = _save_uploaded_file(layout_json_upload, LAYOUTS_TEMP_DIR, LAYOUT_EXTENSIONS)
    elif insert_spaghetti:
        layout_json_path = _default_layout_path(posto)

    if insert_spaghetti and layout_json_path is None and layout_image_path is None:
        st.warning("⚠ Nenhum layout JSON ou imagem fornecido para o mapa de espaguete. O mapa pode não ser gerado corretamente.")

    video_path_for_pipeline = saved_video_path
    metadata = _build_metadata(
        departamento=departamento,
        linha=linha,
        bloco=bloco,
        posto=posto,
        processo=processo,
        responsavel=responsavel,
        data_analise=data_analise,
        takt_time_s=takt_time_s,
        observacoes_usuario=observacoes_usuario,
        fonte_video=str(saved_video_path) if saved_video_path is not None else None,
    )

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUTS_DIR / f"analise_sps_{_safe_slug(posto)}_{_timestamp()}.xlsx"
    max_frames = None if max_frames_value == 0 else max_frames_value

    try:
        progress_box = st.empty()
        progress_bar = st.progress(0)
        pipeline_metrics: dict[str, Any] = {
            "quality_mode": quality_mode,
            "window_seconds": window_seconds,
            "max_frames_per_window": max_frames_per_window,
            "reprocess_low_confidence": reprocess_low_confidence,
        }

        def _progress_callback(stage: str, payload: dict[str, Any]) -> None:
            pipeline_metrics[stage] = payload
            if stage == "knowledge":
                progress_box.info("1/10 Consultando memorias SPS e documentos anexados...")
                progress_bar.progress(6)
            elif stage == "metadata":
                progress_box.info("2/10 Lendo metadados do video...")
                progress_bar.progress(10)
            elif stage == "frames":
                progress_box.info("3/10 Extraindo frames representativos...")
                progress_bar.progress(22)
            elif stage == "timeline":
                progress_box.info("3/10 Criando indice temporal completo do video...")
                progress_bar.progress(24)
            elif stage == "windows":
                count = payload.get("count")
                if count is None:
                    progress_box.info("4/10 Criando janelas adaptativas...")
                else:
                    progress_box.info(f"4/10 Dividindo video em {count} janelas adaptativas...")
                progress_bar.progress(34)
            elif stage == "overview":
                progress_box.info("5/10 Gerando overview visual do video...")
                progress_bar.progress(45)
            elif stage == "window":
                current = int(payload.get("current", 0))
                total = max(1, int(payload.get("total", 1)))
                progress_box.info(f"6/10 Analisando janela {current} de {total}...")
                progress_bar.progress(min(75, 45 + int((current / total) * 30)))
            elif stage == "window_checkpoint":
                progress_box.info(f"6/10 Retomando janela {payload.get('window_index')} ja concluida em checkpoint...")
            elif stage == "reanalyze":
                progress_box.info("7/10 Reprocessando janelas de baixa confianca quando necessario...")
                progress_bar.progress(82)
            elif stage == "consolidate":
                progress_box.info("8/10 Consolidando microetapas e recalculando tempos...")
                progress_bar.progress(90)

        with st.spinner("Processando analise SPS por janelas..."):
            analysis = _run_analysis_only_compat(
                video_path=str(video_path_for_pipeline) if video_path_for_pipeline is not None else None,
                output_path=str(output_path),
                metadata=metadata,
                knowledge_root=str(KNOWLEDGE_ROOT),
                knowledge_paths=knowledge_paths,
                provider_name=provider_name,
                fps=fps,
                max_frames=max_frames,
                window_seconds=window_seconds,
                max_frames_per_window=max_frames_per_window,
                reprocess_low_confidence=reprocess_low_confidence,
                detail_window=str(QUALITY_PRESETS[quality_mode]["detail_window"]),
                quality_mode=quality_mode,
                progress_callback=_progress_callback,
            )
        progress_box.info("10/10 Validacao SPS concluida. Revise antes de gerar Excel.")
        progress_bar.progress(100)
        st.session_state["last_pipeline_metrics"] = pipeline_metrics

        st.session_state["pending_review"] = {
            "analysis": analysis.model_dump(mode="json"),
            "template_path": str(template_path),
            "output_path": str(output_path),
            "fill_standard": fill_standard,
            "standard_export_mode": "standard_consolidado",
            "insert_charts": insert_charts,
            "insert_spaghetti": insert_spaghetti,
            "layout_path": str(layout_json_path) if layout_json_path is not None else None,
            "layout_image_path": layout_image_path,
            "video_path": str(saved_video_path) if saved_video_path is not None else None,
            "metadata": metadata.model_dump(mode="json"),
            "knowledge_paths": knowledge_paths,
            "editor_key": f"microsteps_review_{_timestamp()}",
        }
        st.session_state["last_template_path"] = str(template_path)
        st.session_state["last_analysis"] = analysis.model_dump(mode="json")
        st.session_state["last_video_path"] = str(saved_video_path) if saved_video_path is not None else None
        st.session_state["last_metadata"] = metadata.model_dump(mode="json")
        st.session_state["last_knowledge_paths"] = knowledge_paths
        st.session_state["analysis_output_base_stem"] = output_path.stem
        st.success("✓ Análise gerada. Revise as microetapas antes de gerar o Excel final.")
    except Exception as exc:  # pragma: no cover - comportamento visual.
        st.error(f"Não foi possível processar a análise: {exc}")


def main() -> None:
    if st is None:
        raise RuntimeError("Streamlit nao esta instalado. Execute: pip install streamlit")
    _render_app()


if __name__ == "__main__":
    main()
