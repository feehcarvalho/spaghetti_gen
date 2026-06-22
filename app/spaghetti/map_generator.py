"""Geracao e insercao do mapa de espaguete no workbook SPS."""

from __future__ import annotations

import json
import logging
import os
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from app.schemas.analysis import OperationalAnalysis, SpaghettiMove


logger = logging.getLogger(__name__)

DEFAULT_SPAGHETTI_SHEET = "Diagrama de Espaguete"
DEFAULT_SPAGHETTI_ANCHOR = "B4"
DEFAULT_SPAGHETTI_WIDTH = 760
DEFAULT_SPAGHETTI_HEIGHT = 520


def load_layout(layout_path: str) -> dict:
    """Carrega e valida minimamente um layout JSON de posto."""

    path = Path(layout_path)
    if not path.exists():
        raise FileNotFoundError(f"Layout nao encontrado: {path}")

    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Layout deve ser um objeto JSON")
    if "locais" not in data or not isinstance(data["locais"], dict):
        raise ValueError("Layout deve conter objeto 'locais'")

    for name, point in data["locais"].items():
        if not isinstance(point, dict):
            raise ValueError(f"Local invalido no layout: {name}")
        if "x" not in point or "y" not in point:
            raise ValueError(f"Local sem coordenadas x/y no layout: {name}")
        try:
            float(point["x"])
            float(point["y"])
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Coordenadas invalidas para local: {name}") from exc

    return data


def _normalize_name(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    ascii_text = decomposed.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", ascii_text.lower()).strip()


def _layout_index(layout: dict) -> dict[str, tuple[str, dict[str, Any]]]:
    index: dict[str, tuple[str, dict[str, Any]]] = {}
    for name, point in layout.get("locais", {}).items():
        index[_normalize_name(name)] = (name, point)
        description = point.get("descricao")
        if isinstance(description, str) and description.strip():
            index[_normalize_name(description)] = (name, point)
    return index


def _resolve_point(
    location_name: str,
    index: dict[str, tuple[str, dict[str, Any]]],
) -> tuple[str, dict[str, Any]] | None:
    return index.get(_normalize_name(location_name))


def _append_alert(analysis: OperationalAnalysis, message: str) -> None:
    if message not in analysis.alertas_validacao:
        analysis.alertas_validacao.append(message)


def _valid_moves(
    analysis: OperationalAnalysis,
    index: dict[str, tuple[str, dict[str, Any]]],
) -> list[tuple[SpaghettiMove, tuple[str, dict[str, Any]], tuple[str, dict[str, Any]]]]:
    if analysis.spaghetti is None:
        _append_alert(analysis, "Mapa de espaguete nao gerado com movimentos: dados spaghetti ausentes.")
        return []

    valid: list[tuple[SpaghettiMove, tuple[str, dict[str, Any]], tuple[str, dict[str, Any]]]] = []
    for move in analysis.spaghetti.movimentos:
        origin = _resolve_point(move.origem, index)
        destination = _resolve_point(move.destino, index)
        if origin is None or destination is None:
            missing_parts = []
            if origin is None:
                missing_parts.append(f"origem '{move.origem}'")
            if destination is None:
                missing_parts.append(f"destino '{move.destino}'")
            _append_alert(
                analysis,
                f"Movimento spaghetti {move.ordem} ignorado: {', '.join(missing_parts)} ausente(s) no layout.",
            )
            continue
        valid.append((move, origin, destination))
    return valid


def _format_total_box(analysis: OperationalAnalysis, layout: dict, valid_count: int) -> str:
    lines = [f"Layout: {layout.get('layout_id', '')}", f"Movimentos desenhados: {valid_count}"]
    if analysis.spaghetti is not None:
        if analysis.spaghetti.total_passos_estimados is not None:
            lines.append(f"Passos estimados: {analysis.spaghetti.total_passos_estimados}")
        if analysis.spaghetti.distancia_total_m is not None:
            lines.append(f"Distancia total: {analysis.spaghetti.distancia_total_m:.1f} m")
    return "\n".join(line for line in lines if line)


def generate_spaghetti_map_image(
    analysis: OperationalAnalysis,
    layout_path: str,
    output_png: str,
) -> str:
    """Gera PNG do trajeto de espaguete usando somente pontos do layout."""

    layout = load_layout(layout_path)
    output = Path(output_png)
    output.parent.mkdir(parents=True, exist_ok=True)
    matplotlib_config_dir = output.parent / "matplotlib_cache"
    matplotlib_config_dir.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(matplotlib_config_dir))

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    width = float(layout.get("largura", 10))
    height = float(layout.get("altura", 8))
    index = _layout_index(layout)
    valid_moves = _valid_moves(analysis, index)

    fig, ax = plt.subplots(figsize=(8.2, 5.6), dpi=150)
    ax.set_title("Diagrama de Espaguete", fontsize=14, fontweight="bold")
    ax.set_xlim(0, width)
    ax.set_ylim(0, height)
    ax.set_aspect("equal", adjustable="box")
    ax.set_xlabel("Layout X")
    ax.set_ylabel("Layout Y")
    ax.grid(True, linestyle=":", linewidth=0.7, alpha=0.55)

    for move, origin, destination in valid_moves:
        _, origin_point = origin
        _, destination_point = destination
        x1 = float(origin_point["x"])
        y1 = float(origin_point["y"])
        x2 = float(destination_point["x"])
        y2 = float(destination_point["y"])
        ax.annotate(
            "",
            xy=(x2, y2),
            xytext=(x1, y1),
            arrowprops={
                "arrowstyle": "->",
                "color": "#C62828",
                "linewidth": 1.8,
                "alpha": 0.75,
                "shrinkA": 8,
                "shrinkB": 8,
            },
        )
        midpoint_x = (x1 + x2) / 2
        midpoint_y = (y1 + y2) / 2
        ax.text(
            midpoint_x,
            midpoint_y,
            str(move.ordem),
            ha="center",
            va="center",
            fontsize=8,
            fontweight="bold",
            color="#FFFFFF",
            bbox={"boxstyle": "circle,pad=0.25", "facecolor": "#1565C0", "edgecolor": "#FFFFFF"},
        )

    for name, point in layout["locais"].items():
        x = float(point["x"])
        y = float(point["y"])
        description = point.get("descricao") or name
        ax.scatter([x], [y], s=95, color="#1B5E20", edgecolors="#FFFFFF", linewidths=1.2, zorder=5)
        ax.text(
            x,
            y + 0.18,
            f"{name}\n{description}",
            ha="center",
            va="bottom",
            fontsize=7.5,
            color="#222222",
            bbox={"boxstyle": "round,pad=0.25", "facecolor": "#FFFFFF", "edgecolor": "#DDDDDD", "alpha": 0.9},
            zorder=6,
        )

    ax.text(
        0.99,
        0.02,
        _format_total_box(analysis, layout, len(valid_moves)),
        transform=ax.transAxes,
        ha="right",
        va="bottom",
        fontsize=8.5,
        bbox={"boxstyle": "round,pad=0.35", "facecolor": "#F7F7F7", "edgecolor": "#CCCCCC"},
    )
    fig.tight_layout()
    fig.savefig(output, format="png", bbox_inches="tight")
    plt.close(fig)

    return str(output)


def _spaghetti_anchor() -> str:
    return os.getenv("SPAGHETTI_MAP_ANCHOR", DEFAULT_SPAGHETTI_ANCHOR)


def insert_layout_image(
    workbook_path: str,
    image_path: str,
    sheet_name: str = DEFAULT_SPAGHETTI_SHEET,
) -> None:
    """Insere imagem de layout/posto como referência na aba de spaghetti."""
    workbook_file = Path(workbook_path)
    image_file = Path(image_path)

    if not workbook_file.exists():
        raise FileNotFoundError(f"Workbook nao encontrado: {workbook_file}")
    if not image_file.exists():
        raise FileNotFoundError(f"Imagem do layout nao encontrada: {image_file}")

    workbook = load_workbook(workbook_file, data_only=False, keep_links=True)
    if sheet_name not in workbook.sheetnames:
        logger.warning("Aba %s nao encontrada; imagem de layout nao inserida.", sheet_name)
        workbook.close()
        return

    sheet = workbook[sheet_name]
    image = ExcelImage(str(image_file))
    image.width = DEFAULT_SPAGHETTI_WIDTH
    image.height = DEFAULT_SPAGHETTI_HEIGHT
    sheet.add_image(image, _spaghetti_anchor())
    workbook.save(workbook_file)
    workbook.close()


def insert_spaghetti_map(
    workbook_path: str,
    image_path: str,
    sheet_name: str = DEFAULT_SPAGHETTI_SHEET,
) -> None:
    """Insere a imagem de espaguete em aba existente, sem apagar objetos."""

    workbook_file = Path(workbook_path)
    image_file = Path(image_path)

    if not workbook_file.exists():
        raise FileNotFoundError(f"Workbook nao encontrado: {workbook_file}")
    if not image_file.exists():
        raise FileNotFoundError(f"Imagem do spaghetti nao encontrada: {image_file}")

    workbook = load_workbook(workbook_file, data_only=False, keep_links=True)
    if sheet_name not in workbook.sheetnames:
        logger.warning("Aba %s nao encontrada; mapa de espaguete nao inserido.", sheet_name)
        workbook.close()
        return

    sheet = workbook[sheet_name]
    image = ExcelImage(str(image_file))
    image.width = DEFAULT_SPAGHETTI_WIDTH
    image.height = DEFAULT_SPAGHETTI_HEIGHT
    sheet.add_image(image, _spaghetti_anchor())
    workbook.save(workbook_file)
    workbook.close()
